import streamlit as st
import pandas as pd
import numpy as np
import os
from PIL import Image
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.ticker as ticker
from matplotlib.ticker import MaxNLocator
from streamlit_folium import st_folium
import folium
import plotly.graph_objects as go
# import squarify
import seaborn as sns
import plotly.express as px
import sys
import re
import openpyxl
from openpyxl.styles import Border, Side
from io import BytesIO
from numerize.numerize import numerize
import requests


#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
# PAGE structure
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
## ATENTION: CONFIRM IF I ONLY NEED TO DO THIS IN THE MAIN PAGE, NOT IN SECONDARY PAGES

im = Image.open("logo_web.png")
# st.set_page_config(page_title="RtR DATA EXPLORER", page_icon=im, layout="wide", initial_sidebar_state="expanded")
css_path = "style.css"

with open(css_path) as css:
    st.markdown( f'<style>{css.read()}</style>' , unsafe_allow_html= True)
##
##Hide footer streamlit
##
hide_streamlit_style = """
            <style>
            #MainMenu {visibility: hidden;}
            footer {visibility: hidden;}
            </style>
            """
st.markdown(hide_streamlit_style, unsafe_allow_html=True)
## Hide index when showing a table. CSS to inject contained in a string
hide_table_row_index = """
            <style>
            thead tr th:first-child {display:none}
            tbody th {display:none}
            </style>
            """
##
## Inject CSS with Markdown
##
st.markdown(hide_table_row_index, unsafe_allow_html=True)

#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
# Export Data & Filtering to only Partners*
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________

# Confirm why before i had df_len here. It is somehow related to the "buscador"
# @st.cache_data
# def load_data_cleaned_df_len():
#     from etl_process import df_len
#     return df_len

#Data Geo Country
@st.cache_resource
def load_countries():
        political_countries_url = "http://geojson.xyz/naturalearth-3.3.0/ne_50m_admin_0_countries.geojson"
        return political_countries_url
political_countries_url = load_countries()

#GI 2023
@st.cache_data
def load_data_cleaned_df_gi():
    from etl_process import df_gi
    return df_gi

@st.cache_data
def load_data_cleaned_df_gi_summary():
    from etl_process import df_gi_summary
    return df_gi_summary

@st.cache_data
def load_data_cleaned_formatted_date_gi():
    from etl_process import formatted_date_gi
    return formatted_date_gi

#PLEDGE 2023
@st.cache_data
def load_data_cleaned_df_pledge():
    from etl_process import df_pledge
    return df_pledge

@st.cache_data
def load_data_cleaned_df_pledge_summary():
    from etl_process import df_pledge_summary
    return df_pledge_summary

@st.cache_data
def load_data_cleaned_formatted_date_pledge():
    from etl_process import formatted_date_pledge
    return formatted_date_pledge

@st.cache_data
def load_df_pledge_summary_no_filtered():
    from etl_process import df_pledge_summary_no_filtered
    return df_pledge_summary_no_filtered


#PLAN 2023
@st.cache_data
def load_data_cleaned_df_2023_plan():
    from etl_process import df_2023_plan
    return df_2023_plan

@st.cache_data
def load_data_cleaned_df_2023_plan_summary():
    from etl_process import df_2023_plan_summary
    return df_2023_plan_summary

@st.cache_data
def load_data_cleaned_formatted_date_plan():
    from etl_process import formatted_date_plan
    return formatted_date_plan


#PLAN 2022
@st.cache_data 
def load_data_plan_2022():
    from etl_process import df_2022_plan
    return df_2022_plan

# PLAN CONSOLIDADO
@st.cache_data 
def load_data_plan_consolided():
    from etl_process import df_2022_2023_plan_consolidated
    return df_2022_2023_plan_consolidated

# #CAMPAIGN SUMMARY  #Not longer in use
# @st.cache_data
# def load_data_campaignoverview():
#     from etl_process import df_campaignoverall
#     return df_campaignoverall


#ADMINISTRATIVE
@st.cache_data
def load_data_df_partners_name():
    from etl_process import df_partners_name
    return df_partners_name

@st.cache_data
def load_data_geo_info():
    from etl_process import df_country 
    return df_country 

@st.cache_data
def load_data_cleaned_ac_atypes():
    from etl_process import ac_atypes
    return ac_atypes


## Calling the funcions and creating datasets
ac_atypes = load_data_cleaned_ac_atypes()
# df_campaignoverall = load_data_campaignoverview() #not longer in use
df_country =        load_data_geo_info()
df_partners_name =  load_data_df_partners_name()
df_2022_plan =      load_data_plan_2022()
df_gi =             load_data_cleaned_df_gi()
df_gi_summary =     load_data_cleaned_df_gi_summary()
formatted_date_gi = load_data_cleaned_formatted_date_gi()
df_pledge =         load_data_cleaned_df_pledge()
df_pledge_summary = load_data_cleaned_df_pledge_summary()
formatted_date_pledge = load_data_cleaned_formatted_date_pledge()
df_2023_plan =           load_data_cleaned_df_2023_plan()
df_2023_plan_summary =   load_data_cleaned_df_2023_plan_summary()
formatted_date_plan = load_data_cleaned_formatted_date_gi()
df_2022_2023_plan_consolidated = load_data_plan_consolided()
df_pledge_summary_no_filtered = load_df_pledge_summary_no_filtered()


#To Filter only Data Of RtR Current Partners:   # COnfirm this because it is already done in etl.py. In rigor, this should not be here. 
#concider to apply diferently somehow. 

# df_campaignoverall = df_campaignoverall[df_campaignoverall['Org_Type'] == 'Partner']  #Not longer in use
df_2022_plan = df_2022_plan[df_2022_plan['Org_Type'] == 'Partner']
df_gi = df_gi[df_gi['Org_Type'] == 'Partner']
df_gi_summary = df_gi_summary[df_gi_summary['Org_Type'] == 'Partner']
df_pledge = df_pledge[df_pledge['Org_Type'] == 'Partner']
df_pledge_summary = df_pledge_summary[df_pledge_summary['Org_Type'] == 'Partner']

df_2022_2023_plan_consolidated = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['Org_Type'] == 'Partner']
df_2023_plan_summary = df_2023_plan_summary[df_2023_plan_summary['Org_Type'] == 'Partner']


#CONFIRMATIONS
## ONLY THOSE RESPONDING IN 2023
#Identificar quienes respondieron solo en 2023
# st.write('GI 2023 responded')
# df_gi_summary['gi_2023_responded']

# st.write('PLEDGE')
# df_pledge_summary['Pledge2023responded']

# st.write('PLAN')
# df_2023_plan_summary['survey_reviewed_plan']

# ['InitName']
# df_gi_summary[['InitName','gi_2023_responded']]
# df_pledge_summary[['InitName','Pledge2023responded']]
# df_2023_plan_summary[['InitName','survey_reviewed_plan']]

# Assuming you have three dataframes: df1, df2, df3
# And 'SecondVar' is the name of the column that needs to be True

# Filter each dataframe where 'SecondVar' is True and then get unique 'InitName'
# unique_names_1 = df_gi_summary[df_gi_summary['gi_2023_responded'] == True]['InitName'].unique()
# unique_names_2 = df_pledge_summary[df_pledge_summary['Pledge2023responded'] == True]['InitName'].unique()
# unique_names_3 = df_2023_plan_summary[df_2023_plan_summary['survey_reviewed_plan'] == True]['InitName'].unique()

# # Combine all unique values into one list
# combined_unique_names = pd.Series(np.concatenate((unique_names_1, unique_names_2, unique_names_3)))

# # Get the unique values from the combined list
# all_unique_names = combined_unique_names.unique()

# # Count the number of unique 'InitName' values across the three datasets
# number_of_unique_names = len(all_unique_names)
# # number_of_unique_names

#ONLY SINGLE DATA SETS
# df_2022_plan

# df_2022_plan = df_2022_plan[df_2022_plan['Org_Type'] == 'Partner']
# df_2023_plan = df_2023_plan[df_2023_plan['Org_Type'] == 'Partner']

# st.write("df_campaignoverall") #Add TypeofOrg para poder hacer el filtro
# df_campaignoverall[['InitName','Org_Type']]
# # df_campaignoverall.columns

# # st.write("df_partners_name")
# # df_partners_name
# # df_partners_name.columns

# st.write("df_2022_plan")
# df_2022_plan[['InitName','Org_Type']]
# # df_2022_plan.columns

# st.write("df_gi")
# df_gi[['InitName','Org_Type']]
# # df_gi.columns

# st.write("df_gi_summary")
# df_gi_summary[['InitName','Org_Type']]
# # df_gi_summary.columns

# st.write("df_pledge")
# df_pledge[['InitName','Org_Type']]
# # df_pledge.columns

# st.write("df_pledge_summary")
# df_pledge_summary[['InitName','Org_Type']]
# # df_pledge_summary.columns

# st.write("df_2023_plan")
# df_2023_plan[['InitName','Org_Type']]
# # df_2023_plan.columns

# st.write("df_2023_plan_summary")
# df_2022_plan[['InitName','Org_Type']]
# # df_2022_plan.columns



#_____________________________________________________
#_____________________________________________________
#_____________________________________________________
#_____________________________________________________
# BLOCK DATA: GEOGRAFICAL DISTRIBUTION OF PLEDGES
#_____________________________________________________
#_____________________________________________________
#_____________________________________________________
#_____________________________________________________


def calculate_number_partners_reporting_countries_pledge(df_pledge):
    all_countries = [country.strip() for sublist in df_pledge['All_Countries'].str.split(';') for country in sublist if country.strip()]
    unique_countries = list(set(all_countries))
    df_pledge_unique_countries = pd.DataFrame(unique_countries, columns=['Country'])
    n_partners_reporting_countries = df_pledge[df_pledge['All_Countries'].str.strip() != '; ; ; ; ; ; ; ; ;'].shape[0]
    
    return df_pledge_unique_countries, n_partners_reporting_countries

df_pledge_unique_countries, n_partners_reporting_countries = calculate_number_partners_reporting_countries_pledge(df_pledge)


# Making a dataset with the number of partners per country to plot in work map

#PART 1
# Create a list of unique country names
countries = df_pledge['All_Countries'].str.split(';').explode().str.strip().unique()

# Initialize a dictionary to store the results
result_dict = {'country': [], 'frequency': [], 'Short name': []}

# For each country, count the frequency of users and concatenate the short names
for country in countries:
    # Count the frequency of users
    frequency = df_pledge['All_Countries'].str.contains(country.replace('(', '\(').replace(')', '\)')).sum()
    # Concatenate the short names separated by semicolons
    users = '; '.join(df_pledge.loc[df_pledge['All_Countries'].str.contains(country.replace('(', '\(').replace(')', '\)')), 'short_name'].unique())
    # Add the results to the dictionary
    result_dict['country'].append(country)
    result_dict['frequency'].append(frequency)
    result_dict['Short name'].append(users)

# Create a new DataFrame with the results
result_df1 = pd.DataFrame(result_dict)

# PART 2
# Create a copy of the original DataFrame and split the All_Countries column into separate countries
df2_pledge = df_pledge.copy()
df2_pledge['All_Countries'] = df2_pledge['All_Countries'].str.split(';')

# Explode the All_Countries column to create one row per country
df2_pledge = df2_pledge.explode('All_Countries')

# Group by country and count the number of users and the frequency of each country
result_df2_pledge = df2_pledge.groupby('All_Countries').agg({'short_name': lambda x: '; '.join(x.unique()), 'All_Countries': 'count'})

# Rename the columns
result_df2_pledge.rename(columns={'All_Countries': 'frequency2', 'short_name': 'Short name2'}, inplace=True)

# Merge the two DataFrames based on the country name
result_df = pd.merge(result_df1, result_df2_pledge, left_on='country', right_index=True, how='left')

# Fill NaN values in the merged DataFrame with 0
result_df.fillna(0, inplace=True)

# Concatenate the frequency and Short name columns from both codes into single columns
result_df['frequency'] = result_df['frequency'] + result_df['frequency2']
result_df['Short name'] = result_df[['Short name', 'Short name2']].apply(lambda x: '; '.join(filter(None, x)), axis=1)

# Drop the unnecessary columns
result_df.drop(['frequency2', 'Short name2'], axis=1, inplace=True)

# Sort the DataFrame by frequency in descending order
result_df.sort_values('frequency', ascending=False, inplace=True)

# Reset the index of the DataFrame
result_df.reset_index(drop=True, inplace=True)

# Rename the columns
result_df.rename(columns={'country': 'Country','frequency':'N Parters per country', 'Short name': 'Partners per country'}, inplace=True)
# Print the new DataFrame




#  PART 3
#dataframe to workwith (All data from the selection).
df2_pledge = df_pledge['All_Countries'].str.split(";", expand=True).apply(lambda x: x.str.strip())
df2_pledge = df2_pledge.stack().value_counts()
df2_pledge = df2_pledge.iloc[1:].sort_index().reset_index(name='Frecuency')
df2_pledge.rename(columns = {'index':'Country'}, inplace = True)

# Calculate basic percentage
df2_pledge['% country'] = (df2_pledge['Frecuency'] / df2_pledge['Frecuency'].sum()) * 100
# st.write(df2_pledge['% country'].sum())

# PART 4. Making one dataset
# Merge dataframes on 'Country' column and index
df_partners_per_country = pd.merge(df2_pledge, result_df, on='Country', how='left')

# Rename colums
df_partners_per_country = df_partners_per_country.rename(columns={'Frecuency': 'N Pledges', "% country": "% Pledges"})
# df_partners_per_country # Final Data Set to Chart. 


## Counting number of parters to indicate in future charts about countries (pledges)
#Sample size from selection to chart
sz = str(n_partners_reporting_countries)
df2 = pd.merge(df_partners_per_country, df_country, on='Country', how='right')
df2 = df2.dropna()


# _________________
#Getting Values from dataset to provide descriptions of maps [PLEDGE]
# _____________________

# COUNTRIES PLEDGE
# ______  
# gettin the top 3 countries   ## I might not be using this. 
# ___
s_df2 = df2.sort_values(by=['N Pledges'],ascending=False)
s_df2_best3 = s_df2.head(3)
list_best3 = s_df2_best3['Country'].unique()
list_best3 =  ', '.join(list_best3)

#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end

#getting a string with the first three countries.
list_best3 = replace_last(list_best3, ',', ' & ')

# ________
# gettion the 5% of countries top
# ______
num_countries = len(df2['N Pledges'])
top5_percent = int(np.ceil(num_countries * 0.05))
s_df2_best3 = s_df2.head(top5_percent)
list_best3 = s_df2_best3['Country'].unique()
list_best3 =  ', '.join(list_best3)

#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end

#getting a string with the 5% of countries top.
list_best_5_per = replace_last(list_best3, ',', ' & ')



# CONTINENTS PLEDGE
# ___________________
# gettin the best 3 continents and lowest 2
# ________________

# sum of % country in 'continent' and creating a new df
df3 = df2.groupby(['continent']).agg({'% Pledges': 'sum'})
df3 = df3.reset_index()
df3.rename(columns={'% Pledges': '% Continent'}, inplace=True)

df2 = pd.merge(df2, df3, on='continent', how='left')
df2 = df2.dropna()

s_df2 = df3.sort_values(by='% Continent', ascending=False)
s_df2_best3 = s_df2.head(3)
list_best3 = s_df2_best3['continent'].unique()
list_best3 =  ', '.join(list_best3)
#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end
#getting a string with the first three continents top.
list_best3_conti = replace_last(list_best3, ',', ' & ')

# _________________________
# #making a list of the two less presnce in 'continent'
s_df2 = df3.sort_values(by='% Continent', ascending=True)
s_df2_best3 = s_df2.head(2)
list_best3 = s_df2_best3['continent'].unique()
list_best3 =  ', '.join(list_best3)
#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end
#getting a string with the last three continents.
list_less_2_conti = replace_last(list_best3, ',', ' & ')


# REGIONS PLEDGE
# # ___________________
# # gettin the best 3 regions and lowest 2 regions
# # ________________

# sum of % country in 'region_wb' and creating a new df
df3 = df2.groupby(['region_wb']).agg({'% Pledges': 'sum'})
df3 = df3.reset_index()
df3.rename(columns={'% Pledges': '% Region'}, inplace=True)
df2 = pd.merge(df2, df3, on='region_wb', how='left')
df2 = df2.dropna()
# st.write(df2)

s_df2 = df3.sort_values(by='% Region', ascending=False)
s_df2_best3 = s_df2.head(3)
list_best3 = s_df2_best3['region_wb'].unique()
list_best3 =  ', '.join(list_best3)
#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end
#getting a string with the first three regions.
list_best3_region = replace_last(list_best3, ',', ' & ')

# _________________________
# #making a list of the three less presence in 'region'
s_df2 = df3.sort_values(by='% Region', ascending=True)
s_df2_best3 = s_df2.head(3)
list_best3 = s_df2_best3['region_wb'].unique()
list_best3 =  ', '.join(list_best3)
#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end
#getting a string with the last three regions.
list_less_2_region = replace_last(list_best3, ',', ' & ')


## SUB REGIONS PLEDGE
# # ___________________
# # gettin the best 3 Sub Regions and lowest 2  Sub regions
# # ________________

# sum of % country in 'subregion' and creating a new df
df3 = df2.groupby(['subregion']).agg({'% Pledges': 'sum'})
df3 = df3.reset_index()
df3.rename(columns={'% Pledges': '% Sub Region'}, inplace=True)
df2 = pd.merge(df2, df3, on='subregion', how='left')
df2 = df2.dropna()
# st.write(df2)

s_df2 = df3.sort_values(by='% Sub Region', ascending=False)
s_df2_best3 = s_df2.head(3)
list_best3 = s_df2_best3['subregion'].unique()
list_best3 =  ', '.join(list_best3)
#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end
#getting a string with the first three subregion.
list_best3_subregion = replace_last(list_best3, ',', ' & ')

# _________________________
# #making a list of the three less presence in 'region'
s_df2 = df3.sort_values(by='% Sub Region', ascending=True)
s_df2_best3 = s_df2.head(3)
list_best3 = s_df2_best3['subregion'].unique()
list_best3 =  ', '.join(list_best3)
#making a function to change the last comma with "&"
def replace_last(string, delimiter, replacement):
    start, _, end = string.rpartition(delimiter)
    return start + replacement + end
#getting a string with the last three subregion.
list_less_2_subregion = replace_last(list_best3, ',', ' & ')
# st.write(df2.columns)

####_________________________________________________________
####_________________________________________________________
#### END BLOCK DATA  GEOGRAFICAL DISTRIBUTION OF PLEDGES
####_________________________________________________________
####________________________________________________________-



####_________________________________________________________
####_________________________________________________________
#### BLOCK DATA  GEOGRAFICAL DISTRIBUTION OF PLAN
####_________________________________________________________
####________________________________________________________-


def calculate_number_partners_reporting_countries_plan(df_2022_2023_plan_consolidated):
    all_countries = [country.strip() for sublist in df_2022_2023_plan_consolidated['All_Countries'].str.split(';') for country in sublist if country.strip()]
    unique_countries = list(set(all_countries))
    df_2022_2023_plan_consolidated_unique_countries = pd.DataFrame(unique_countries, columns=['Country'])
    n_partners_reporting_countries = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['All_Countries'].str.strip() != '; ; ; ; ; ; ; ; ; ; ; ; ; ; '].shape[0]
    n_partners_reporting_countries = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['All_Countries'].str.strip() != '; ; ; ; '].shape[0]
    return df_2022_2023_plan_consolidated_unique_countries, n_partners_reporting_countries

df_2022_2023_plan_consolidated_unique_countries, n_partners_reporting_countries = calculate_number_partners_reporting_countries_plan(df_2022_2023_plan_consolidated)


# Making a dataset with the number of partners per country to plot in work map

#PART 1
# Create a list of unique country names
countries = df_2022_2023_plan_consolidated['All_Countries'].str.split(';').explode().str.strip().unique()

# Initialize a dictionary to store the results
result_dict = {'country': [], 'frequency': [], 'Short name': []}

# For each country, count the frequency of users and concatenate the short names
for country in countries:
    # Count the frequency of users
    frequency = df_2022_2023_plan_consolidated['All_Countries'].str.contains(country.replace('(', '\(').replace(')', '\)')).sum()
    # Concatenate the short names separated by semicolons
    users = '; '.join(df_2022_2023_plan_consolidated.loc[df_2022_2023_plan_consolidated['All_Countries'].str.contains(country.replace('(', '\(').replace(')', '\)')), 'short_name'].unique())
    # Add the results to the dictionary
    result_dict['country'].append(country)
    result_dict['frequency'].append(frequency)
    result_dict['Short name'].append(users)

# Create a new DataFrame with the results
result_countries_df_plan_consolidated1_plan_consolidated = pd.DataFrame(result_dict)

# PART 2
# Create a copy of the original DataFrame and split the All_Countries column into separate countries
df2_2022_2023_plan_consolidated = df_2022_2023_plan_consolidated.copy()
df2_2022_2023_plan_consolidated['All_Countries'] = df2_2022_2023_plan_consolidated['All_Countries'].str.split(';')

# Explode the All_Countries column to create one row per country
df2_2022_2023_plan_consolidated = df2_2022_2023_plan_consolidated.explode('All_Countries')

# Group by country and count the number of users and the frequency of each country
result_countries_df_plan_consolidated2_2022_2023_plan_consolidated = df2_2022_2023_plan_consolidated.groupby('All_Countries').agg({'short_name': lambda x: '; '.join(x.unique()), 'All_Countries': 'count'})

# Rename the columns
result_countries_df_plan_consolidated2_2022_2023_plan_consolidated.rename(columns={'All_Countries': 'frequency2', 'short_name': 'Short name2'}, inplace=True)

# Merge the two DataFrames based on the country name
result_countries_df_plan_consolidated = pd.merge(result_countries_df_plan_consolidated1_plan_consolidated, result_countries_df_plan_consolidated2_2022_2023_plan_consolidated, left_on='country', right_index=True, how='left')

# Fill NaN values in the merged DataFrame with 0
result_countries_df_plan_consolidated.fillna(0, inplace=True)

# Concatenate the frequency and Short name columns from both codes into single columns
result_countries_df_plan_consolidated['frequency'] = result_countries_df_plan_consolidated['frequency'] + result_countries_df_plan_consolidated['frequency2']
result_countries_df_plan_consolidated['Short name'] = result_countries_df_plan_consolidated[['Short name', 'Short name2']].apply(lambda x: '; '.join(filter(None, x)), axis=1)

# Drop the unnecessary columns
result_countries_df_plan_consolidated.drop(['frequency2', 'Short name2'], axis=1, inplace=True)

# Sort the DataFrame by frequency in descending order
result_countries_df_plan_consolidated.sort_values('frequency', ascending=False, inplace=True)

# Reset the index of the DataFrame
result_countries_df_plan_consolidated.reset_index(drop=True, inplace=True)

# Rename the columns
result_countries_df_plan_consolidated.rename(columns={'country': 'Country','frequency':'N Parters per country', 'Short name': 'Partners per country'}, inplace=True)
# Print the new DataFrame

result_countries_df_plan_consolidated = result_countries_df_plan_consolidated[result_countries_df_plan_consolidated['Country'] != ""]

n_countries_plan_consolidated = result_countries_df_plan_consolidated['Country'].nunique()
# st.write("Number of Countries Plan Consolidated")
# n_countries_plan_consolidated


####_________________________________________________________
####_________________________________________________________
#### END BLOCK DATA  GEOGRAFICAL DISTRIBUTION OF PLAN
####_________________________________________________________
####________________________________________________________-






####__________________________________________________________________________________________________________________________________________________________________
####__________________________________________________________________________________________________________________________________________________________________
####__________________________________________________________________________________________________________________________________________________________________
###  Magnitud Metrics Calculations
####__________________________________________________________________________________________________________________________________________________________________
####__________________________________________________________________________________________________________________________________________________________________
####__________________________________________________________________________________________________________________________________________________________________

st.write(
    """
    <style>
    [data-testid="stMetricDelta"] svg {
        display: none;
    }
    </style>
    """,
    unsafe_allow_html=True,)


## caculating diferences between 2022 and 2023
def calculate_percentage_change(old_value, new_value):
    if old_value == 0:
        raise ValueError("The old value cannot be zero for percentage calculation")
    percentage_change = ((new_value - old_value) / old_value) * 100
    return percentage_change

#CAMPAIGN GOAL
rtr_goal_individuals = 4000000000 

## GENERAL LEVEL
n_all_rtrpartners_2022 = 30 
n_all_rtrpartners_2023 = ((df_partners_name["Org_Type"] == "Partner") | (df_partners_name["Org_Type"] == "Partner New")).sum()  #34 
percentage_increase_number_partnes_all_campaign = round(calculate_percentage_change(n_all_rtrpartners_2022, n_all_rtrpartners_2023))
# percentage_increase_number_partnes_all_campaign
n_reporting_partners = (df_partners_name["Org_Type"] == "Partner").sum()  ## THIS VARIABLE IS USED AS 100%. OTHERS PARTNERS HAVE NOT STARTED REPORTING PROCESS
n_newpartners2023 = n_all_rtrpartners_2023 - n_reporting_partners

n_rtrmember = (df_partners_name["Org_Type"] == "Member").sum()
n_rtrnotpartner = (df_partners_name["Org_Type"] == "No longer a Partner").sum()

#GI
n_partner_gi = df_gi_summary[(df_gi_summary["Org_Type"] == "Partner") & (df_gi_summary["gi_overall_index"] > 0)].shape[0]
n_member_gi = df_gi_summary[(df_gi_summary["Org_Type"] == "Member") & (df_gi_summary["gi_overall_index"] > 0)].shape[0]
n_initiatives_gi_pending = df_gi_summary[((df_gi_summary["Org_Type"] == "Partner") | (df_gi_summary["Org_Type"] == "Member")) & (df_gi_summary["gi_overall_index"] == 0)].shape[0]

#PLEDGE
n_partner_pledge_2023    = df_pledge_summary[(df_pledge_summary["Org_Type"] == "Partner") & (df_pledge_summary["PledgeOverallIdx"] > 0)].shape[0]
n_member_pledge_2023     = df_pledge_summary[(df_pledge_summary["Org_Type"] == "Member") & (df_pledge_summary["PledgeOverallIdx"] > 0)].shape[0]
n_initiatives_pledge_pending = n_reporting_partners - n_partner_pledge_2023
percentaje_reporting_partner_pledge_2023 = round((n_partner_pledge_2023 / n_reporting_partners)*100)
# n_partner_pledge_2023
# n_reporting_partners
# percentaje_reporting_partner_pledge_2023
# n_initiatives_pledge_pending = df_gi_summary[((df_pledge_summary["Org_Type"] == "Partner") | (df_pledge_summary["Org_Type"] == "Member")) & (df_pledge_summary["PledgeOverallIdx"] == 0)].shape[0]

#PLAN CONSOLIDATED (CONSOLIDATED)
n_partner2022_2023_plan = df_2022_2023_plan_consolidated[(df_2022_2023_plan_consolidated["Org_Type"] == "Partner")]["InitName"].nunique()
n_member2022_2023_plan = df_2022_2023_plan_consolidated[(df_2022_2023_plan_consolidated["Org_Type"] == "Member")]["InitName"].nunique()
n_total_plans_consolidated = df_2022_2023_plan_consolidated.shape[0]
percentaje_reporting_partner_plan_consolidated = round((n_partner2022_2023_plan / n_reporting_partners)*100)

dollars_movilized_all_plans = df_2022_2023_plan_consolidated['US_Dollars_Mobilized'].sum()
dollars_mobilized_all_plans_n_plans = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['US_Dollars_Mobilized'] > 0]['InitName'].count()
dollars_mobilized_all_plans_n_partners = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['US_Dollars_Mobilized'] > 0]['InitName'].nunique()
percentaje_partner_reporting_adaptation_finance_consolidated = round((dollars_mobilized_all_plans_n_partners / n_reporting_partners)*100)

## Getting the numbers of Initiatives plans in each primary beneficiary (CONSOLIDATED)
n_2022_2023_plans_individuals = df_2022_2023_plan_consolidated['IndivPlan'].notna().sum()
n_2022_2023_plans_companies = df_2022_2023_plan_consolidated['CompPlan'].notna().sum()
n_2022_2023_plans_regions = df_2022_2023_plan_consolidated['RegionsPlan'].notna().sum()
n_2022_2023_plans_cities = df_2022_2023_plan_consolidated['CitiesPlan'].notna().sum()
n_2022_2023_plans_natsyst = df_2022_2023_plan_consolidated['NatSysPlan'].notna().sum()

#Number of plans (CONSOLIDATED)
#GetTinG the total number of partners that have any reported metrics in all data_consolidated
numbers_provid_list = ['n_direc_indiv','n_indiv_companies','n_indiv_regions','n_indiv_cities','n_hecta_natsys']
# Create a condition for non-NA values and non-zero values
condition = (df_2022_2023_plan_consolidated[numbers_provid_list].notna()) & (df_2022_2023_plan_consolidated[numbers_provid_list] != 0)
# Check if any columns in the row meet the condition and assign the result to 'reportmetrics'
df_2022_2023_plan_consolidated['reportmetrics'] = condition.any(axis=1)
n_plans_reporting_metrics_plan_consolidated = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['reportmetrics'] == True].shape[0]


#Number of individuals Plan consolidated
n_individuals_plan_consolidated = 1869492849  #Given by JC

#Number of hectares plan
n_hectareas_plan_consolidated = round(df_2022_2023_plan_consolidated['n_hecta_natsys'].sum()*0.75)  ## Same metodology as in pledge number of hectareas
# n_hectareas_plan_consolidated
n_partners_reporting_n_hectareas_plan_consolidated = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['n_hecta_natsys'] > 0]['n_hecta_natsys'].nunique()
# n_partners_reporting_n_hectareas_plan_consolidated


# List of variables to use in plan consolidated: 
# n_plans_reporting_metrics_plan_consolidated  # Numbers of plans whit information in metrics about individuals
# n_2022_2023_plans_individuals # Number of plans per primary beneficiay
# n_2022_2023_plans_companies # Number of plans per primary beneficiay
# n_2022_2023_plans_regions # Number of plans per primary beneficiay
# n_2022_2023_plans_cities # Number of plans per primary beneficiay
# n_2022_2023_plans_natsyst # Number of plans per primary beneficiay

# n_direct_individuals_plan_consolid = df_2022_2023_plan_consolidated['n_direc_indiv'].sum()*0.75
# n_partners_reportingnumberof_direct_individuals2023 = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['n_direc_indiv'].notna() & (df_2022_2023_plan_consolidated['n_direc_indiv'] != 0)]['n_direc_indiv'].shape[0]

# n_indiv_companies_plan_consolid = df_2022_2023_plan_consolidated['n_indiv_companies'].sum()*0.75
# n_partners_reportingnumberofindiv_companies2023 = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['n_indiv_companies'].notna() & (df_2022_2023_plan_consolidated['n_indiv_companies'] != 0)]['n_indiv_companies'].shape[0]

# n_indiv_regions_plan_consolid = df_2022_2023_plan_consolidated['n_indiv_regions'].sum()*0.75
# n_partners_reportingnumberofindiv_regions2023 = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['n_indiv_regions'].notna() & (df_2022_2023_plan_consolidated['n_indiv_regions'] != 0)]['n_indiv_regions'].shape[0]

# n_indiv_natsyst_plan_consolid = 0
# n_partners_reportingnumberofindiv_regions2023 = df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['n_indiv_regions'].notna() & (df_2022_2023_plan_consolidated['n_indiv_regions'] != 0)]['n_indiv_regions'].shape[0]

# df_2022_2023_plan_consolidated.columns
# n_direct_individuals_plan_consolidated =
# n_indv_companies_plan_consolidated =
# n_indv_regions_plan_consolidated =
# n_indv_cites_plan_consolidated =
# n_indv_natsyst_plan_consolidated =
# n_companies_plan_consolidated =
# n_regions_plan_consolidated =
# n_cites_plan_consolidated =
# n_natsyst_plan_consolidated =


## Plan 2023 only
n_partner_plan_only2023 = df_2023_plan_summary[(df_2023_plan_summary["Org_Type"] == "Partner") & (df_2023_plan_summary["plan_overall_idx"] > 0)]["InitName"].nunique()
n_member_plan_only2023 = df_2023_plan_summary[(df_2023_plan_summary["Org_Type"] == "Member") & (df_2023_plan_summary["plan_overall_idx"] > 0)]["InitName"].nunique()
# n_initiatives_plan2023_pending = (n_reporting_partners + n_rtrmember)-(n_partner_plan_only2023 + n_member_plan_only2023)

partner_n_plan_only2023    = df_2023_plan_summary[(df_2023_plan_summary["Org_Type"] == "Partner") & (df_2023_plan_summary["plan_overall_idx"] > 0)].shape[0]
member_n_plan_only2023     = df_2023_plan_summary[(df_2023_plan_summary["Org_Type"] == "Member") & (df_2023_plan_summary["plan_overall_idx"] > 0)].shape[0]

## Members 2022
# n_members2022 = df_gi_summary['Members2022Total'].sum()
n_members2022 = 2085   ##Info Data Explorer. 
# n_partnersreportinmembers2022 =  np.count_nonzero(df_gi_summary['Members2022Total'])
n_partnersreportinmembers2022 = 23

n_members2023 = df_gi_summary['Members2023Total'].sum()
n_partnersreportinmembers2023 = np.count_nonzero(df_gi_summary['Members2023Total']) 

## Plan 2022 only
n_plans_in_2022 = df_2022_plan.shape[0]
n_partners_reporting_plan_in_2022 = df_2022_plan['InitName'].nunique()
# n_plans_in_2022
# n_partners_reporting_plan_in_2022


## DETAILS LEVEL

## PLEDGES
## Getting the numbers of Initiatives pledge in each beneficiary
partners_report_pledge_d_individuals2023 = df_pledge_summary['DirIndivPldg'].notna().sum()
partners_repor_pledge_companies2023 = df_pledge_summary['CompPldg'].notna().sum()
partners_repor_pledge_regions2023 = df_pledge_summary['RegPldg'].notna().sum()
partners_repor_pledge_cities2023 = df_pledge_summary['CitiesPldg'].notna().sum()
partners_repor_pledge_nat_syst2023 = df_pledge_summary['NatSysPldg'].notna().sum()

## Identification of number of individuals (ALL PLEDGE)
n_direct_individuals2022 = df_pledge_summary_no_filtered['DirIndiv2022Num'].sum()*0.75
n_partners_reportingnumberof_direct_individuals2022 = df_pledge_summary[df_pledge_summary['DirIndiv2022Num'].notna() & (df_pledge_summary['DirIndiv2022Num'] != 0)]['DirIndiv2022Num'].shape[0]

n_direct_individuals2022 = df_pledge_summary_no_filtered['DirIndiv2022Num'].sum()*0.75
n_partners_reportingnumberof_direct_individuals2022 = df_pledge_summary[df_pledge_summary['DirIndiv2022Num'].notna() & (df_pledge_summary['DirIndiv2022Num'] != 0)]['DirIndiv2022Num'].shape[0]

n_direct_individuals2023 = df_pledge_summary['DirIndiv2023Num'].sum()*0.75
n_partners_reportingnumberof_direct_individuals2023 = df_pledge_summary[df_pledge_summary['DirIndiv2023Num'].notna() & (df_pledge_summary['DirIndiv2023Num'] != 0)]['DirIndiv2023Num'].shape[0]

n_indiv_companies2023 = df_pledge_summary['CompsNumIndiv'].sum()*0.75
n_partners_reportingnumberofindiv_companies2023 = df_pledge_summary[df_pledge_summary['CompsNumIndiv'].notna() & (df_pledge_summary['CompsNumIndiv'] != 0)]['CompsNumIndiv'].shape[0]

n_indiv_regions2023 = df_pledge_summary['NumIndivRegs'].sum()*0.75
n_partners_reportingnumberofindiv_regions2023 = df_pledge_summary[df_pledge_summary['NumIndivRegs'].notna() & (df_pledge_summary['NumIndivRegs'] != 0)]['NumIndivRegs'].shape[0]

n_indiv_cities2023 = df_pledge_summary['NumIndivCities'].sum()*0.75
n_partners_reportingnumberofindiv_cities2023 = df_pledge_summary[df_pledge_summary['NumIndivCities'].notna() & (df_pledge_summary['NumIndivCities'] != 0)]['NumIndivCities'].shape[0]

n_indiv_natsyt2022 = df_pledge_summary_no_filtered['NumIndivNatSys'].sum()*0.75
n_partners_reportingnumberofindiv_natsyst2023 = df_pledge_summary[df_pledge_summary_no_filtered['NumIndivNatSys'].notna() & (df_pledge_summary_no_filtered['NumIndivNatSys'] != 0)]['NumIndivNatSys'].shape[0]

# oficial number of individuals here
total_numbers_individuals_2023 = n_direct_individuals2023 + n_indiv_companies2023 + n_indiv_regions2023 + n_indiv_cities2023 + n_indiv_natsyt2022
total_numbers_individuals_pledge_oficial_IR = 3166185040  #Calculated

#GetTinG the total number of partners that have any reported metrics in 2023
numbers_provid_list = ['DirIndiv2023Num','CompsNumIndiv','NumIndivRegs','NumIndivCities','NumIndivNatSys']
# Create a condition for non-NA values and non-zero values
condition = (df_pledge_summary[numbers_provid_list].notna()) & (df_pledge_summary[numbers_provid_list] != 0)
# Check if any columns in the row meet the condition and assign the result to 'reportmetrics'
df_pledge_summary['reportmetrics'] = condition.any(axis=1)
n_parters_reporting_metrics_pledge = df_pledge_summary[df_pledge_summary['reportmetrics'] == True].shape[0]


# counting numbers of beneficiaries (pledge)
Comps2022Num_pledge = df_pledge_summary_no_filtered['Comps2022Num'].sum()
Comps2022Num_pledge_adj = df_pledge_summary_no_filtered['Comps2022Num'].sum()*0.75
n_partners_reportingnumberofcompanies2022_pledge = df_pledge_summary_no_filtered[df_pledge_summary_no_filtered['Comps2022Num'].notna() & (df_pledge_summary_no_filtered['Comps2022Num'] != 0)]['Comps2022Num'].shape[0]

Comps2023Num_pledge = df_pledge_summary['Comps2023Num'].sum()
Comps2023Num_pledge_adj = round(df_pledge_summary['Comps2023Num'].sum() * 0.75)
n_partners_reportingnumberofcompanies2023_pledge = df_pledge_summary[df_pledge_summary['Comps2023Num'].notna() & (df_pledge_summary['Comps2023Num'] != 0)]['Comps2023Num'].shape[0]

Regs2022Nums_pledge = df_pledge_summary_no_filtered['Regs2022Nums'].sum()
Regs2022Nums_pledge_adj = df_pledge_summary_no_filtered['Regs2022Nums'].sum()*0.75
n_partners_reportingnumberofregions2022_pledge = df_pledge_summary_no_filtered[df_pledge_summary_no_filtered['Regs2022Nums'].notna() & (df_pledge_summary_no_filtered['Regs2022Nums'] != 0)]['Regs2022Nums'].shape[0]

Regs2023Nums_pledge = df_pledge_summary['Regs2023Nums'].sum()
# Regs2023Nums_pledge_adj = df_pledge_summary['Regs2023Nums'].sum()*0.75
Regs2023Nums_pledge_adj = round(df_pledge_summary['Regs2023Nums'].sum() * 0.75)
n_regions_regions4 = 100   #Adjusted 
n_partners_reportingnumberofregions2023_pledge = df_pledge_summary[df_pledge_summary['Regs2023Nums'].notna() & (df_pledge_summary['Regs2023Nums'] != 0)]['Regs2023Nums'].shape[0]

Cities2022Num_pledge = df_pledge_summary_no_filtered['Cities2022Num'].sum()
Cities2022Num_pledge_adj = df_pledge_summary_no_filtered['Cities2022Num'].sum()*0.75
n_partners_reportingnumberofcities2022_pledge = df_pledge_summary_no_filtered[df_pledge_summary_no_filtered['Cities2022Num'].notna() & (df_pledge_summary_no_filtered['Cities2022Num'] != 0)]['Cities2022Num'].shape[0]

Cities2023Num_pledge = df_pledge_summary['Cities2023Num'].sum()
# Cities2023Num_pledge_adj = df_pledge_summary['Cities2023Num'].sum()*0.75
Cities2023Num_pledge_adj = round(df_pledge_summary['Cities2023Num'].sum() * 0.75)
n_cities_citiesrtr = 86 #Adjusted 
n_partners_reportingnumberofcities2023_pledge = df_pledge_summary[df_pledge_summary['Cities2023Num'].notna() & (df_pledge_summary['Cities2023Num'] != 0)]['Cities2023Num'].shape[0]

NatSysHectaTotal2022_pledge = df_pledge_summary_no_filtered['NatSysHectaTotal2022'].sum()
NatSysHectaTotal2022_pledge_adj = df_pledge_summary_no_filtered['NatSysHectaTotal2022'].sum()*0.75
n_partners_reportingnumberofnaturalsyst2022_pledge = df_pledge_summary_no_filtered[df_pledge_summary_no_filtered['NatSysHectaTotal2022'].notna() & (df_pledge_summary_no_filtered['NatSysHectaTotal2022'] != 0)]['NatSysHectaTotal2022'].shape[0]

NumHectNatSys2023_pledge = df_pledge_summary['NumHectNatSys2023'].sum()
# NumHectNatSys2023_pledge_adj = df_pledge_summary['NumHectNatSys2023'].sum()*0.75
NumHectNatSys2023_pledge_adj = round(df_pledge_summary['NumHectNatSys2023'].sum() * 0.75)
n_partners_reportingnumberofnaturalsyst2023_pledge = df_pledge_summary[df_pledge_summary['NumHectNatSys2023'].notna() & (df_pledge_summary['NumHectNatSys2023'] != 0)]['NumHectNatSys2023'].shape[0]


## caculating diferences between 2022 and 2023
def calculate_percentage_change(old_value, new_value):
    if old_value == 0:
        raise ValueError("The old value cannot be zero for percentage calculation")
    percentage_change = ((new_value - old_value) / old_value) * 100
    return percentage_change

percentage_increase_resilience_hectares_pledge = round(calculate_percentage_change(NatSysHectaTotal2022_pledge_adj, NumHectNatSys2023_pledge_adj))
percentage_increase_resilience_individuals_pledge = round(calculate_percentage_change(n_direct_individuals2022, total_numbers_individuals_pledge_oficial_IR))
percentage_increase_n_plan2022vs2023 = round(calculate_percentage_change(n_plans_in_2022, n_total_plans_consolidated ))
percentage_increase_n_partners_with_plan2022vs2023 = round(calculate_percentage_change(n_partners_reporting_plan_in_2022, n_partner2022_2023_plan))
# n_plans_in_2022
# n_total_plans_consolidated
# percentage_increase_n_partners_with_plan2022vs2023


# # pendientes
# num_people_potential_increase = 999
# num_hectares_potential_increase = 999
# percentage_progress_people = 999
# percentage_progress_hectares = 999

# dollars_movilized_all_plans
# dollars_mobilized_all_plans_n_partners
# dollars_mobilized_all_plans_n_plans

## Comparation Pledge vs Plan
percentaje_individuals_pledge_vs_plan = round((n_individuals_plan_consolidated/total_numbers_individuals_pledge_oficial_IR )*100)
percentaje_hectareas_nat_syst_pledge_vs_plan = round((n_hectareas_plan_consolidated/NumHectNatSys2023_pledge_adj )*100)
percentaje_countries_pledge_vs_plan = round((n_countries_plan_consolidated/df_pledge_unique_countries.shape[0])*100)




##__________________________________________________________________________________________________________________________________________________________________
##__________________________________________________________________________________________________________________________________________________________________
##__________________________________________________________________________________________________________________________________________________________________
#### END Magnitud Metrics Calculations
##__________________________________________________________________________________________________________________________________________________________________
##__________________________________________________________________________________________________________________________________________________________________
##__________________________________________________________________________________________________________________________________________________________________


####_____________________________________
####_____________________________________
#### Presentation of metrics_____________
####_____________________________________
####_____________________________________

st.title("RtR CAMPAIGN OVERVIEW")

# st.markdown('#### SUMMARY (2023) 15 nov.')
# summary_text = f"""
# Since its inception in 2021, the Race to Resilience campaign has undergone thirteen rounds of application reviews and now includes {n_reporting_partners} partner organizations (see Appendix I), operating across {df_pledge_unique_countries.shape[0]} countries on all continents, addressing all systems covered by the Sharm-el-Sheikh Adaptation Agenda.

# To date, partners have reported a total of {n_members2023} collaborating members, underscoring a vast network and a collective endeavor to fortify global resilience.

# Progress is contingent upon each partner’s capabilities and their campaign membership duration. Among all partners, {n_partner_pledge_2023} have articulated a pledge, cumulatively aiming to bolster the resilience of {numerize(total_numbers_individuals_pledge_oficial_IR)} individuals and {numerize(NumHectNatSys2023_pledge_adj)} hectares of natural systems—an upswing of {numerize(percentage_increase_resilience_individuals_pledge)}% and {percentage_increase_resilience_hectares_pledge}%, respectively, from the previous year.

# At the forefront, the countries with the highest count of pledged actions by RtR Partners—comprising the top 5%—include: {list_best_5_per}.

# Regionally, {list_best3_region} stand out as frontrunners in the frequency of pledges from RtR Partners. On the other hand, {list_less_2_region} present less pledges.

# Additionally, {n_partner2022_2023_plan} partners have initiated plans to achieve their commitments, potentially increasing the resilience of {num_people_potential_increase} people and {num_hectares_potential_increase} hectares of natural systems. This represents progress of {percentage_progress_people}% and {percentage_progress_hectares}% towards their pledged goals, respectively.

# This progression underlines the dynamic expansion and the growing influence of the campaign in fostering climate resilience worldwide. Moreover, {dollars_mobilized_all_plans_n_partners} partners through {dollars_mobilized_all_plans_n_plans} resilience-buidling plans have collectively mobilized more than {numerize(dollars_movilized_all_plans)} US Dollars in financial resources.
# """
# st.markdown(summary_text)


# st.markdown('#### TEXT #1.')
# short_description = f"""
# Nearly three years since its launch, the campaign remains committed to its mission of transforming pledges into tangible impacts. Out of {n_all_rtrpartners_2023} RtR Partners, {n_reporting_partners} have participated in the current reporting process. Therefore, this report provides a comprehensive overview of the progress achieved so far by these partners, representing over {n_members2023} members. Alongside these entities, {numerize(Comps2023Num_pledge_adj)} companies, {numerize(n_cities_citiesrtr)} cities, and {numerize(n_regions_regions4)} regions are implementing resilience actions across {df_pledge_unique_countries.shape[0]} countries. Together, these initiatives have pledged to enhance the resilience of {numerize(total_numbers_individuals_pledge_oficial_IR)} billion people by 2030. Additionally,  according to the reports of the partners and members' they are mobilizing substantial financial resources, amounting to ${numerize(dollars_movilized_all_plans)} billion USD. This effort not only benefits people but also nature covering {numerize(NumHectNatSys2023_pledge_adj)} million hectares.
# """
# st.markdown(short_description)

# st.markdown('#### TEXT #2')


long_description = f"""
### Executive Summary

As the first-ever Global Stocktake is set to conclude at the UN Climate Change Conference COP28 and the definition of a Global Goal on Adaptation makes significant advances, the Race to Resilience Campaign presents its contribution to the progress on adaptation delivered by the Non-Party Stakeholders. The overarching goal: putting people and nature first,making 4 billion people more resilient to climate change by 2030.
The strength of the Campaign lies on the joint ambition shared by the partners and members that implement locally-led adaptation actions to benefit the most vulnerable, supported by a credible, comprehensive and transparent progress tracking framework that unifies all partners and members in turning pledges into action. 

To date, the combined pledges from Campaign partners seek to strengthen the resilience of {numerize(total_numbers_individuals_pledge_oficial_IR)} billion people by 2030. This ambition has increased {numerize(percentage_increase_resilience_individuals_pledge)}% compared to 2022.  The pledges target {numerize(NumHectNatSys2023_pledge_adj)} million hectares of natural systems to be protected, restored or managed by 2030 and deliver benefits to people and the planet. 

In addition, for the first time the Campaign demonstrates progress on the cumulative results achieved against the pledges. In Partners’s Plans, there are {numerize(n_individuals_plan_consolidated)} individuals estimated to increase resilience ({percentaje_individuals_pledge_vs_plan}% of the Pledge), {numerize(n_hectareas_plan_consolidated)} hectares of natural systems where adaptation measures are planned to be implemented ({percentaje_hectareas_nat_syst_pledge_vs_plan}% of the Pledge), and {numerize(dollars_movilized_all_plans)} billion USD mobilised to accelerate implementation. The progress highlights the gap between the partners’ pledges and plans, as well as the extent of effort needed to fulfill these pledges. These are all significant contributions of Non-Party Stakeholders in reducing vulnerability, enhancing adaptive capacity, and increasing resilience.

Transparent reporting is crucial for demonstrating climate adaptation progress - to build trust,  strengthen accountability, and inform evidence-based policies. While speed of progress varies based on each partner's capacity and the year in which they joined, to date {numerize(percentaje_reporting_partner_pledge_2023)}% of reporting partners have submitted a pledge, {numerize(percentaje_reporting_partner_plan_consolidated)}% of partners submitted an action plan and almost {numerize(percentaje_partner_reporting_adaptation_finance_consolidated)}% of partners reported the adaptation finance mobilised. The Campaign’s Data Explorer presents the complete database of the global actions and analysis of the results achieved to date.

Much more is needed to deliver changes at the speed and scale that our Planet needs. But, this report shows how NPS are increasingly stepping up to accelerate inclusive and equitable climate action to meaningfully change the pace and scale needed. Since its launch in 2021, the campaign has grown and counts now with a total of {n_all_rtrpartners_2023} partner initiatives, an increase of {numerize(percentage_increase_number_partnes_all_campaign)}% in relation to 2022, and with a total of {n_members2023} collaborating members. Furthermore, new subnational governments have joined the Race through their Flagship Initiatives: with a cumulated {n_cities_citiesrtr} cities from Cities Race to Resilience and a cumulated {n_regions_regions4} regions from RegionsAdapt. The adaptation solutions pledged by partners and members are located in {df_pledge_unique_countries.shape[0]} countries on every continent, across {Cities2023Num_pledge_adj} cities and {Regs2023Nums_pledge_adj} regions. In Partners’ plans, the {percentaje_countries_pledge_vs_plan}% of those countries are considered, gap that will be reduced as Partners increase the number of their plans in the future. 
The adaptation solutions delivered span across all priority systems included in the Sharm-el-Sheikh Adaptation Agenda and informed its first implementation report.
"""
st.markdown(long_description)



st.markdown('#### RtR CURRENTLY FEATURES (2023)')
st.markdown("**STAGE 1: APPLIED & HAS BEEN ACCEPTED TO RTR CAMPAIGN**")
col1, col2, col3,col4 = st.columns(4)

# col1.metric("Initiatives",n_reporting_partners)
col1.metric("Partners (Reporting Cicle 2023)",n_reporting_partners)
col2.metric("Members",n_rtrmember) 
col3.metric("Not longer Partner",n_rtrnotpartner) 

st.markdown("* **Countries pledged for operation by RtR Partner Initiatives***")
col1, col2, col3 = st.columns(3)
col1.metric("Countries 2022",173,"(out of 18 partners)")
col2.metric("Countries 2023",df_pledge_unique_countries.shape[0],"(out of "+str(n_partners_reporting_countries)+" partners)")
st.caption("*Differences in the number of countries arise because initiatives that are no longer partners are not counted. Additionally, during the 2023 reporting process, Partner Initiatives may have indicated fewer countries due to enhancements in their metrics systems over the past year.")

st.markdown("* **Member organizations represented by RtR Partner Initiatives.**")
col1, col2, col3 = st.columns(3)
col1.metric("Members organizations 2022",n_members2022,"(out of "+str(n_partnersreportinmembers2022)+" partners by 2022)")##
col2.metric("Members organizations 2023",n_members2023,"(out of "+str(n_partnersreportinmembers2023)+" partners)")##
st.caption("*Differences in member counts are due to initiatives no longer being partners and thus not being included. In the 2023 reporting process, Partner Initiatives might have reported fewer members because they were required to explicitly identify organizations that align with the following definition: *In RtR, a member of an RtR Partner is defined as: an implementer, collaborator (partner organization), endorser, or donor - entities that directly contribute to the pledge, plan, and resilience actions reported by your initiative to the campaign, and have consented to be linked to the RtR through the partner initiative.* During the 2022 reporting process, they were not as strict with this definition and included all related organizations, regardless of their consent to be linked to RtR. Additionally, a significant number of RtR Initiatives have not yet reported their member counts for the 2023 process. As a result, this number is expected to increase in subsequent metric updates")

st.markdown('#### RtR INITIATIVES REPORTING PROGRESS')
st.markdown("**STAGE 1.2: REPORTING GENERAL INFORMATION**")
col1, col2, col3 = st.columns(3)
col1.metric("Partners reporting General Information*",n_partner_gi)
# col2.metric("Members reporting  General Information*",n_member_gi)
col2.metric("Initiatives pending General Information",n_initiatives_gi_pending) 
st.caption("*This doesn't indicate a 100% complete report. Some Initiatives still need to finish certain modules of the reporting tool.")

st.markdown("**STAGE 2: PLEDGE STATEMENTS**")
col1, col2, col3 = st.columns(3)
col1.metric("Partners reporting Pledge Statements*",n_partner_pledge_2023)
# col2.metric("Members reporting Pledge Statements*",n_member_pledge_2023)
col2.metric("Initiatives pending Pledge Statements",n_initiatives_pledge_pending)
st.caption("*This doesn't indicate a 100% complete report. Some Initiatives still need to finish certain modules of the reporting tool.")

st.markdown("**STAGE 3: RESILIENCE-BUILDING PLAN**")
col1, col2, col3 = st.columns(3)
col1.metric("Partners' reporting plan",n_partner2022_2023_plan)
col2.metric("Total Plans (2022-2023)",n_total_plans_consolidated)

# col2.metric("Members' reporting 2022 & 2023 Plans",n_member2022_2023_plan) 
col1, col2, col3 = st.columns(3)
col1.metric("Partners Updating Plan in 2023",n_partner_plan_only2023)
# col2.metric("Members' reporting  Plans 2023",n_member_plan_only2023 ) 
# col2.metric("Initiatives pending Plans 2023",n_initiatives_plan2023_pending) ##not going anymore

# dollars_movilized_all_plans
# dollars_mobilized_all_plans_n_partners
# dollars_mobilized_all_plans_n_plans

col2.metric("Total Updated Plan in 2023*",partner_n_plan_only2023)
# col2.metric("Members' N° Plans 2023*",member_n_plan_only2023) 
col1.metric("US Dolars Movilized by Partners",numerize(dollars_movilized_all_plans),"(out of "+str(dollars_mobilized_all_plans_n_plans)+" plans from "+str(dollars_mobilized_all_plans_n_partners)+" Partners)")

st.caption("*In 2023, the RtR Technical Secretariat and the Methodological Advisory Group (MAG) collaborated closely with RtR Initiatives during monthly sessions. They initiated a Metric Enhancement Process to refine the 2022 Plan Reporting Tool. After COP27, the integration of the Sharm El Sheikh Adaptation Agenda became essential for resilience building plan submissions. This resulted in a more streamlined reporting process, grounded in a Theory of Action, which emphasized both the magnitude and depth of resilience. As a result, RtR Initiatives were asked to use the Enhanced 2023 Resilience Building Plan Reporting Tool for their submissions. This led to a more robust metrics reporting, thereby enhancing the overall RtR Reporting process.")

st.markdown("#### NUMBER OF INDIVIDUALS PLEDGED & PLANNED TO BE MADE MORE RESILIENT TO CLIMATE CHANGE BY 2030")
st.markdown("AGGREGATED METRICS PLEDGE")
col1, col2 = st.columns(2)
col1.metric("N° Direct Individuals Pledged 2022*",numerize(n_direct_individuals2022),"(out of 17 partners)")
# col2.metric("N° Total Individuals Pledged 2023",numerize(total_numbers_individuals_2023),"(out of "+str(n_parters_reporting_metrics_pledge)+" partners)")
col2.metric("N° Total Individuals Pledged 2023 (IR)",numerize(total_numbers_individuals_pledge_oficial_IR),"(out of "+str(n_parters_reporting_metrics_pledge)+" partners)")
st.caption("*Metric presented in the RtR Data Explorer for COP27.")

# st.markdown("AGGREGATED METRICS PLAN")
# col1, col2 = st.columns(2)
# col1.metric("N° Direct Individuals (Plan)*",numerize(n_2022_2023_plans_individuals),"(out of "+str(n_plans_reporting_metrics_plan_consolidated)+" plans)")



st.markdown("DISAGGREGATED METRICS 2023 (N° Individuals Pledged)**")
col1, col2,col3 = st.columns(3)
col1.metric("N° Direct Individuals Pledged 2023",numerize(n_direct_individuals2023),"(out of "+str(n_partners_reportingnumberof_direct_individuals2023)+" partners)")
col2.metric("N° Individuals through Companies 2023",numerize(n_indiv_companies2023),"(out of "+str(n_partners_reportingnumberofindiv_companies2023)+" partners)")
col3.metric("N° Individuals through Regions 2023",numerize(n_indiv_regions2023),"(out of "+str(n_partners_reportingnumberofindiv_regions2023)+" partners)")
col1.metric("N° Individuals through Cities 2023",numerize(n_indiv_cities2023),"(out of "+str(n_partners_reportingnumberofindiv_cities2023)+" partners)")
col2.metric("N° Individuals through Natural Systems 2023",numerize(n_indiv_natsyt2022),"(out of "+str(n_partners_reportingnumberofindiv_natsyst2023)+" partners)")
st.caption("**All metrics have been adjusted for double counting based on 2022 criteria.")

st.markdown("#### NUMBER OF OTHER PRIMARY BENEFICIARIES PLEDGED TO BE MADE MORE RESILIENT TO CLIMATE CHANGE BY 2030.")
st.markdown("**COMPANIES***")
col1, col2,col3 = st.columns(3)
col1.metric("N° Companies Pledged 2022 (Simple Sum)",numerize(Comps2022Num_pledge),"(out of "+str(n_partners_reportingnumberofcompanies2022_pledge)+" partners)")
col2.metric("N° Companies Pledged 2022 (Adjusted)**",numerize(Comps2022Num_pledge_adj),"(out of "+str(n_partners_reportingnumberofcompanies2022_pledge)+" partners)")
col1.metric("N° Companies Pledged 2023 (Simple Sum)",numerize(Comps2023Num_pledge),"(out of "+str(n_partners_reportingnumberofcompanies2023_pledge)+" partners)")
col2.metric("N° Companies Pledged 2023 (Adjusted)**",numerize(Comps2023Num_pledge_adj),"(out of "+str(n_partners_reportingnumberofcompanies2023_pledge)+" partners)")
st.caption("*Metrics regarding the number of companies pledged were not presented in the RtR Data Explorer for COP27")
st.caption("**Metrics have been adjusted for double counting based on 2022 criteria.")

st.markdown("**REGIONS**")
col1, col2,col3 = st.columns(3)
col1.metric("N° Regions Pledged 2022 (Simple Sum)",numerize(Regs2022Nums_pledge),"(out of "+str(n_partners_reportingnumberofregions2022_pledge)+" partners)")
col2.metric("N° Regions Pledged 2022 (Adjusted)*",numerize(Regs2022Nums_pledge_adj),"(out of "+str(n_partners_reportingnumberofregions2022_pledge)+" partners)")
col1.metric("N° Regions Pledged 2023 (Simple Sum)",numerize(Regs2023Nums_pledge),"(out of "+str(n_partners_reportingnumberofregions2023_pledge)+" partners)")
col2.metric("N° Regions Pledged 2023 (Adjusted)**",numerize(Regs2023Nums_pledge_adj),"(out of "+str(n_partners_reportingnumberofregions2023_pledge)+" partners)")
st.caption("*Metric presented in the RtR Data Explorer for COP27.")
st.caption("**Metrics have been adjusted for double counting based on 2022 criteria.")

st.markdown("**CITIES**")
col1, col2,col3 = st.columns(3)
col1.metric("N° Cities Pledged 2022 (Simple Sum)",numerize(Cities2022Num_pledge),"(out of "+str(n_partners_reportingnumberofcities2022_pledge)+" partners)")
col2.metric("N° Cities Pledged 2022 (Adjusted)*",numerize(Cities2022Num_pledge_adj),"(out of "+str(n_partners_reportingnumberofcities2022_pledge)+" partners)")
col1.metric("N° Cities Pledged 2023 (Simple Sum)",numerize(Cities2023Num_pledge),"(out of "+str(n_partners_reportingnumberofcities2023_pledge)+" partners)")
col2.metric("N° Cities Pledged 2023 (Adjusted)**",numerize(Cities2023Num_pledge_adj),"(out of "+str(n_partners_reportingnumberofcities2023_pledge)+" partners)")
st.caption("*Metric presented in the RtR Data Explorer for COP27.")
st.caption("**Metrics have been adjusted for double counting based on 2022 criteria.")

st.markdown("**NATURAL SYSTEMS**")
col1, col2,col3 = st.columns(3)
col1.metric("N° Hectares of Natural Systems Pledged 2022 (Simple Sum)",numerize(NatSysHectaTotal2022_pledge ),"(out of "+str(n_partners_reportingnumberofnaturalsyst2022_pledge)+" partners)")
col2.metric("N° Hectares of Natural Systems Pledged 2022 (Adjusted)*",numerize(NatSysHectaTotal2022_pledge_adj),"(out of "+str(n_partners_reportingnumberofnaturalsyst2022_pledge)+" partners)")
col1.metric("N° Hectares of Natural Systems Pledged 2023 (Simple Sum)",numerize(NumHectNatSys2023_pledge),"(out of "+str(n_partners_reportingnumberofnaturalsyst2023_pledge)+" partners)")
col2.metric("N° Hectares of Natural Systems Pledged 2023 (Adjusted)**",numerize(NumHectNatSys2023_pledge_adj),"(out of "+str(n_partners_reportingnumberofnaturalsyst2023_pledge)+" partners)")
st.caption("*Metric presented in the RtR Data Explorer for COP27.")
st.caption("**Metrics have been adjusted for double counting based on 2022 criteria.")

####_____________________________________
####_____________________________________
#### END Presentation of metrics_________
####_____________________________________
####_____________________________________




#####______________________________________________________________________________________
#####______________________________________________________________________________________
#####______________________________________________________________________________________
##### Mapping RtR Partners' Global Reach
#####______________________________________________________________________________________
#####______________________________________________________________________________________
#####______________________________________________________________________________________


# st.subheader("CHARTING THE RACE: MAPPING RtR PARTNERS' GLOBAL REACH")

# PART 1) MAP HEADQUARTERS
# st.markdown("### HEADQUARTERS")
# #Data set from here to frontend
# df2_gi = df_gi['q29'].value_counts().reset_index()
# df2_gi.columns = ['Country', 'Frequency']
# df3_gi = df_gi[['InitName','q26','q29','short_name']]

# # Function to create the new DataFrame
# def create_country_initiatives_df(df_gi,df_country):
#     # Ensure 'q29' is a string and split it into separate rows for each country
#     df_gi['q29'] = df_gi['q29'].astype(str)
#     countries_expanded = df_gi['q29'].str.split(';', expand=True).stack().reset_index(level=1, drop=True).rename('Country')

#     # Join back to df_gi for the 'short_name' and 'q9'
#     df_gi_expanded = df_gi.join(countries_expanded).reset_index()

#     # Remove any empty or null country entries
#     df_gi_expanded = df_gi_expanded[df_gi_expanded['Country'].str.strip().astype(bool)]

#     # Group by the country and aggregate
#     df_country_initiatives = df_gi_expanded.groupby('Country').agg(
#         n_initiatives=pd.NamedAgg(column='Country', aggfunc='size'),
#         org_short_names=pd.NamedAgg(column='short_name', aggfunc=lambda x: '; '.join(x.dropna().unique())),
#         org_full_names=pd.NamedAgg(column='InitName', aggfunc=lambda x: '\n'.join(['- ' + name for name in x.dropna().unique()]))
#     ).reset_index()
    
#     df_country_initiatives = df_country_initiatives [df_country_initiatives ['Country'] != "nan"]
#     df_country_lat_lon = df_country[['Country','lat','lon']]
#     df_country_initiatives = df_country_initiatives.merge(df_country_lat_lon, on='Country', how='left')
#     return df_country_initiatives

# df4_gi_hq_by_country = create_country_initiatives_df(df_gi, df_country)
# # df4_gi_hq_by_country

# #WORLDMAP HEADQUARTERS
# numbers_partners_hq = df4_gi_hq_by_country['n_initiatives'].sum()
# world_map_HQ = folium.Map(location=(0, 0), zoom_start=1, tiles="cartodb positron") 

# for index, row in df4_gi_hq_by_country.iterrows():
#     lat = row["lat"]
#     lon = row["lon"]
#     if not pd.isna(lat) and not pd.isna(lon):  # skip rows with non-numerical lat/lon values
#         folium.Marker(
#             location=[lat, lon],
#             # popup=row['q9'],
#             # tooltip=f"<b>N° Headquarters:</b> {row['n_initiatives']}"
#             tooltip=f"<b>Country:</b> {row['Country']}<br><b>N° Headquarters:</b> {row['n_initiatives']}<br><b>Initiatives:</b> {row['org_short_names']}"
#             # tooltip=f"<b>Country:</b> {row['Country']}<b>N° Headquarters:</b> {row['n_initiatives']}<br><b>Number of partners per country:</b> {row['N Parters per country']}<br><b>Partners per country:</b> {row['Partners per country']}"
#         ).add_to(world_map_HQ)
        
# folium.Choropleth(
#     geo_data=political_countries_url,
#     data=df2_gi,
#     columns=['Country', 'Frequency'],
#     key_on="feature.properties.sovereignt",
#     fill_color="RdPu",
#     fill_opacity=0.7,
#     line_opacity=0.2,
#     legend_name="Number of RtR Initiatives' Headquarters by Country",
#     highlight=True,
# ).add_to(world_map_HQ)
# folium.LayerControl().add_to(world_map_HQ)

# st_folium(world_map_HQ,width=725, returned_objects=[])
# st.text("Out of "+str(numbers_partners_hq)+" RtR Initiatives (Source: General Information Survey)")
# with st.expander("Check Table of Head Quarters Info"):
#     st.write(df4_gi_hq_by_country)
# END PART 1) MAP HEADQUARTERS







# ## PART 2) PLEDGES DISTRIBUTION
# tab1, tab2 = st.tabs(["World Map" ,"Tree Map" ])
# #___________________________________________
# # WORLD MAP    #https://github.com/jefftune/pycountry-convert/tree/master/pycountry_convert
# ##https://realpython.com/python-folium-web-maps-from-data/
# #___________________________________________

# # _________________________________________
# # FINAL GEO DATA FRAME
# # _________________________________________

# new_order = ['Country','adm0_a3','N Pledges','% Pledges','N Parters per country','Partners per country','continent',
# '% Continent','region_wb','% Region','subregion','% Sub Region','country_to_plot','Frecuency','% country',
# 'Total population in millions','ND_GAIN Index Country','lat','lon',]

# df2 = df2.reindex(columns=new_order)

# @st.cache_data
# def load_countries():
#     political_countries_url = "http://geojson.xyz/naturalearth-3.3.0/ne_50m_admin_0_countries.geojson"
#     return political_countries_url
# political_countries_url = load_countries()

# world_map_full = folium.Map(location=(0, 0), zoom_start=1, tiles="cartodb positron") 

# for index, row in df2.iterrows():
#     lat = row["lat"]
#     lon = row["lon"]
#     if not pd.isna(lat) and not pd.isna(lon):  # skip rows with non-numerical lat/lon values
#         folium.Marker(
#             location=[lat, lon],
#             popup=row['country_to_plot'],
#             tooltip=f"<b>Country:</b> {row['Country']}<br><b>Pledges (All Primary Beneficiaries):</b> {row['N Pledges']}<br><b>Number of Partners</b> {row['N Parters per country']}<br><b>Names of Partners:</b> {row['Partners per country']}"
#         ).add_to(world_map_full)

# folium.Choropleth(
#     geo_data=political_countries_url,
#     data=df2,
#     columns=["country_to_plot", "N Pledges"],
#     key_on="feature.properties.name",
#     fill_color="RdPu",
#     fill_opacity=0.7,
#     line_opacity=0.2,
#     legend_name="Total pledges across all primary beneficiaries in RtR Pledge Statements*",
#     highlight=True,
#     name="Frequency of pledges by country.",

# ).add_to(world_map_full)
# folium.LayerControl().add_to(world_map_full)


# with tab1:
#     st_folium(world_map_full,width=725, key="map1", returned_objects=[])
#     st.text('Out of '+ sz +' RtR Partners (Source: RtR Pledge Statement Survey 2023)')
#     with st.expander("Navigating this worldmap: A User's Guide"):
#         st.write("""
# This interactive map provides an overview of the countries where RtR Partner have made pledges, along with relevant information such as the number of pledges, number and name of partners per country.

# To use the map:

# - Zoom in and out: Use the plus (+) and minus (-) buttons in the top left corner of the map, or scroll up and down on your mouse.

# - Move the map: Click and drag on the map to move it in any direction.

# - View country information: Click on any marker on the map to view a popup with information about that country, such as the country name, number of pledges, number of partners per country, and partners per country.

# - View number of pledges: The countries are color-coded based on the number of pledges in each country. To view the legend, click on the button on the top right corner of the map that looks like layers stacked on top of each other.

# - The grey color indicates that no information is yet available on a country.

# """)
    
#     st.markdown(
#     """
#     **INSIGHTS FROM THE WORLD MAP: KEY OBSERVATIONS AND FINDINGS**
    
#     The world map displays the consolidated pledges dedicated to all primary beneficiaries: Individuals, Companies, Regions, Cities, and Natural Systems, as outlined in the RtR Pledge Statements. An individual RtR Partner can produce a single pledge statement, catering to one or multiple beneficiary categories. Thus, if a partner commits to all categories, the total pledge count represented here would tally to 5 — one for each beneficiary.
    
#     The countries topping the list with the highest number of pledged actions by RtR Partners, representing the uppermost 5%, include: """ + list_best_5_per + """.

#     By continental distribution, the predominance of pledged actions by RtR Partners is noted in """ + list_best3_conti + """. Conversely, """ + list_less_2_conti + """ witness a reduced volume of pledges.
    
#     From a regional perspective, """ + list_best3_region + """ emerge as leaders in the frequency of RtR Partners' pledges. Conversely, """ + list_less_2_region + """ exhibit a lower pledge activity by these partners.
    
#     Zooming into subregional details, """ + list_best3_subregion + """ are the focal points for pledges by RtR Partners. In contrast, """ + list_less_2_subregion + """ experience fewer pledge commitments.
#     """)


# #___________________________________________
# # TREEMAP CONTINENTS
# #___________________________________________

# df2.rename(columns={'continent': 'Continents'}, inplace=True)
# df2.rename(columns={'% Continent': 'Relative presence by continent (%)'}, inplace=True)

# fig_treemap = px.treemap(df2, path=[px.Constant("World"),'Continents','region_wb','subregion','adm0_a3'], values = '% Pledges', color='Relative presence by continent (%)', color_continuous_scale='RdPu')
# fig_treemap.update_traces(root_color="#FF37D5")
# fig_treemap.update_layout(title_text='Treemap - Resilience Action Pledges by Continent, Region, Sub Region and Country')
# fig_treemap.update_layout(margin = dict(t=50, l=25, r=25, b=25), width=1000)
# fig_treemap.add_annotation(x=1, y=0,
#                 text='Out of '+ sz +' RtR Partners (Source: RtR Pledge Statement Survey).',showarrow=False,
#                 yshift=-20)
# fig_treemap.data[0].hovertemplate = '%{value:.1f}%'
# fig_treemap.update_traces(marker_line_color='grey')

# with tab2:
#     st.plotly_chart(fig_treemap)
#     with st.expander("Navigating this treemap: A User's Guide"):
#         st.write("""
#     The **treemap** is a data visualization that *demonstrates hierarchical relationships using nested rectangles*. Each rectangle represents a category, and its size corresponds to a value or proportion of the data. In this particular treemap, the data is organized by continents, regions, and countries.

#     At the highest level of the treemap, there is a rectangle representing the entire world, labeled as "World."
    
#     This treemap *support interactivity* by allowing users to click on a rectangle and zoom in to see nested categories. For example, if you click on a rectangle representing a continent, you might see rectangles representing regions within that continent. By clicking on a rectangle representing a region, you might see rectangles representing subregions within that country.

#     This interactivity makes it easier for users to explore and understand hierarchical relationships within the data. It allows users to zoom in and out of the treemap and view the data from different perspectives. Additionally, users can hover over each rectangle to see more detailed information about the category and its associated data value.
# """)
    
#     st.markdown(
#     """
#     **INSIGHTS FROM THE TREE MAP: KEY OBSERVATIONS AND FINDINGS**
    
#     The treemap provides a visual representation of the consolidated pledges dedicated to all primary beneficiaries: Individuals, Companies, Regions, Cities, and Natural Systems. These pledges are captured in the RtR Pledge Statements. Within the treemap, each rectangle corresponds to a specific geographical area, with the size denoting the percentage of associated pledges. Additionally, the color variations may represent different metrics or categories.

#     Each RtR Partner can issue a single pledge statement, which may cater to one or several primary beneficiaries. For instance, if a partner commits to all of these beneficiaries, their contributions within the treemap would be represented as five distinct pledges — one for each beneficiary. It's from this distribution that the percentage of all pledges is calculated.

#     On a continental scale, the predominance of pledged actions by RtR Partners is denoted by the larger rectangles for """ + list_best3_conti + """. In contrast, the smaller rectangles for """ + list_less_2_conti + """ signify a decreased volume of pledges.

#     Examining the regions, those with substantial pledge activities by RtR Partners are represented by more expansive rectangles, namely """ + list_best3_region + """. Conversely, """ + list_less_2_region + """ are depicted by smaller rectangles, indicating lesser pledge activities.

#     Delving into the subregional specifics, the treemap emphasizes """ + list_best3_subregion + """ with larger rectangles, signifying increased pledge commitments by RtR Partners. On the opposite spectrum, the more modest rectangles for """ + list_less_2_subregion + """ denote fewer pledge commitments.

#     Nested within the subregional rectangles are the associated countries. Notably, the countries that emerge prominently, showcased by the largest rectangles and denoting the highest pledge counts by RtR Partners, include: """ + list_best_5_per + """.
#     """
# )
# ## END PART 2) PLEDGES DISTRIBUTION



# PART 3) EXPERIMENTAL CHARTS - NO PIN
# st.write("#### EXPERIMENTAL CHARTS - NO PIN")

# st.markdown("### World Map (no pin)")
# world_map_clean = folium.Map(location=(0, 0), zoom_start=1, tiles="cartodb positron") 

# # for index, row in df2.iterrows():
# #     lat = row["lat"]
# #     lon = row["lon"]
# #     if not pd.isna(lat) and not pd.isna(lon):  # skip rows with non-numerical lat/lon values
# #         folium.Marker(
# #             location=[lat, lon],
# #             popup=row['country_to_plot'],
# #             tooltip=f"<b>Country:</b> {row['Country']}<br><b>Pledges (All Primary Beneficiaries):</b> {row['N Pledges']}<br><b>Number of Partners</b> {row['N Parters per country']}<br><b>Names of Partners:</b> {row['Partners per country']}"
# #         ).add_to(world_map_clean)

# folium.Choropleth(
#     geo_data=political_countries_url,
#     data=df2,
#     columns=["country_to_plot", "N Pledges"],
#     key_on="feature.properties.name",
#     fill_color="RdPu",
#     fill_opacity=0.7,
#     line_opacity=0.2,
#     legend_name="Total pledges across all primary beneficiaries in RtR Pledge Statements*",
#     highlight=True,
#     name="Frequency of pledges by country.",

# ).add_to(world_map_clean)
# folium.LayerControl().add_to(world_map_clean)

# st_folium(world_map_clean,width=725, key="map2", returned_objects=[])
# st.text('Out of '+ sz +' RtR Partners (Source: RtR Pledge Statement Survey 2023)')
    
# END PART 3) EXPERIMENTAL CHARTS - NO PIN

##___________________________________________
##___________________________________________
##___________________________________________
## END BLOCK MAPS
##___________________________________________
##___________________________________________
##___________________________________________





##___________________________________________
##___________________________________________
##___________________________________________
## CHARTS COMPARATION GOAL - PLEDGE - PLAN
##___________________________________________
##___________________________________________
##___________________________________________

st.markdown("## ADVANCES OF THE CAMPAIGN")

#CHART LEVEL ADVANCE CAMPAIGN
st.markdown("### INDIVIDUALS")
# INDIVIDUALS
rtr_goal_individuals_numerized = str(numerize(rtr_goal_individuals))
total_numbers_individuals_pledge_oficial_IR_numerized = str(numerize(total_numbers_individuals_pledge_oficial_IR))
n_individuals_plan_consolidated_numerized = str(numerize(n_individuals_plan_consolidated))

# Calculate the percentages relative to the campaign goal
percent_goal = (rtr_goal_individuals / rtr_goal_individuals) * 100
percent_pledged = (total_numbers_individuals_pledge_oficial_IR / rtr_goal_individuals) * 100
percent_planned = (n_individuals_plan_consolidated / rtr_goal_individuals) * 100

# Set colors from the RdPu palette
colors = sns.color_palette("RdPu", n_colors=3)

# Set the size of the figure
fig_width = 6  # 
fig_height = 1.5  # 

# Create the figure and bar chart
fig_advances_campaign_individuals, ax = plt.subplots(figsize=(fig_width, fig_height))

# Plotting the bars
bars = ax.bar(['Goal','Pledge', 'Plan'],
              [rtr_goal_individuals,total_numbers_individuals_pledge_oficial_IR, n_individuals_plan_consolidated],
              color=colors[:3])

# Adding values to the columns  (absolute numbers)
for bar, numerized in zip(bars,
                                   [rtr_goal_individuals_numerized, total_numbers_individuals_pledge_oficial_IR_numerized, n_individuals_plan_consolidated_numerized]):
    height = bar.get_height()
    ax.text(bar.get_x() + bar.get_width()/2., height + 1,  # Slightly above the bar
            f'{numerized}', ha='center', va='bottom')


# Set labels and title
ax.set_ylabel('Billions of Individuals')
ax.set_xlabel('RtR Campaign')

# Adjust the layout to make room for the x-axis label
plt.subplots_adjust(bottom=0.25)

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# Show the figure in the Streamlit app
st.pyplot(fig_advances_campaign_individuals)





st.markdown("### HECTARES OF NATURAL SYSTEMS")
##HECTARES
hectares_pledge_numerize =  str(numerize(NumHectNatSys2023_pledge_adj))
hectares_plan_numerize =  str(numerize(n_hectareas_plan_consolidated))

# Calculate the percentages relative to the campaign goal
percent_pledged = (NumHectNatSys2023_pledge_adj / NumHectNatSys2023_pledge_adj) * 100
percent_planned = (n_hectareas_plan_consolidated / NumHectNatSys2023_pledge_adj) * 100

# Set colors from the RdPu palette
colors = sns.color_palette("RdPu", n_colors=2)

# Set the size of the figure
fig_width = 6  # 
fig_height = 1.5  # 

# Create the figure and bar chart
fig_advances_campaign_hectares, ax = plt.subplots(figsize=(fig_width, fig_height))

# Plotting the bars
bars = ax.bar(['Pledge', 'Plan'],
              [NumHectNatSys2023_pledge_adj, n_hectareas_plan_consolidated],
              color=colors[:2])

# Adding values to the columns, combining percentages and absolute numbers
for bar, percent, numerized in zip(bars,
                                   [percent_pledged, percent_planned], 
                                   [hectares_pledge_numerize, hectares_plan_numerize]):
    height = bar.get_height()
    ax.text(bar.get_x() + bar.get_width()/2., height + 1,  # Slightly above the bar
            f'{percent:.1f}%\n({numerized})', ha='center', va='bottom')

# Set labels and title
ax.set_ylabel('Million of Hectares')
ax.set_xlabel('RtR Campaign')

# Adjust the layout to make room for the x-axis label
plt.subplots_adjust(bottom=0.25)

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# Show the figure in the Streamlit app
st.pyplot(fig_advances_campaign_hectares)

##___________________________________________
##___________________________________________
##___________________________________________
## END CHARTS COMPARATION GOAL - PLEDGE - PLAN
##___________________________________________
##___________________________________________
##___________________________________________




##___________________________________________
##___________________________________________
##___________________________________________
## CHARTS PARTNERS - ALL 
##___________________________________________
##___________________________________________
##___________________________________________

# Calculate the percentages relative to the total number of partners
percent_reporting = (n_reporting_partners / n_all_rtrpartners_2023) * 100
percent_new = (n_newpartners2023 / n_all_rtrpartners_2023) * 100

# Set colors from the RdPu palette
colors = sns.color_palette("RdPu", n_colors=2)

# Define the size of the figure
fig_width = 8  # Width of the figure
fig_height = 1.5  # Height of the figure

# Create the figure and horizontal bar chart with defined size
fig_all_partners, ax = plt.subplots(figsize=(fig_width, fig_height))

# Plotting the horizontal bars for reporting and new partners on the same bar
bars = ax.barh(['Partners'], [n_reporting_partners], color=colors[0], label='Reporting Partners')
ax.barh(['Partners'], [n_newpartners2023], left=[n_reporting_partners], color=colors[1], label='New Partners')

# Adding labels inside each section of the bar with absolute number and percentage
# For Reporting Partners
midpoint_reporting = n_reporting_partners / 2
ax.text(midpoint_reporting, 0, f'{n_reporting_partners} Partners Participating in the 2023 Reporting Process\n({percent_reporting:.1f}%)',
        va='center', ha='center', color='black')

# For New Partners
midpoint_new = n_reporting_partners + (n_newpartners2023 / 2)
ax.text(midpoint_new, 0, f'{n_newpartners2023} \nNew Partners\n({percent_new:.1f}%)',
        va='center', ha='center', color='black')

# Set labels
ax.set_xlabel('Number of Partners')
ax.set_yticklabels([])  # Hide the y-axis labels

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# Add annotation text
annotation_text = f"N = {n_all_rtrpartners_2023} Partners"
fig_all_partners.text(0.95, 0.0, annotation_text, ha='right', va='bottom', fontsize=10, color='black')  # Adjust coordinates and alignment

# Adjust layout to fit and show the bar chart properly
plt.tight_layout()

# Show the figure in the Streamlit app
st.write("Partners")
st.pyplot(fig_all_partners)
##___________________________________________
##___________________________________________
##___________________________________________
## END CHART PARTNERS - ALL 
##___________________________________________
##___________________________________________
##___________________________________________














# # #___________________________________________
# # #___________________________________________
# # #___________________________________________
# # #___________________________________________
# # # PARTNERS REPORTING CHARTS
# # #___________________________________________
# # #___________________________________________
# # #___________________________________________
# # #___________________________________________

st.subheader("RtR PARTNERS REPORTING PROGRESS")

tab1, tab2, tab3, tab4 = st.tabs(["All Stages", "General Information", "Pledge Statements","Plans"])

## LEVEL 1) ALL STAGES
# CAMPAIGN OVERVIEW: NUMBER OF PARTNERS IN DIFERENT STAGES OF THE RACE: ## RE-DO INCORPORATING TO CAMPAIGN OVERALL THE PARTNERS THAT ARE NOT PART OF THE 2023 REPORTING CICLE. 

# Variables 
# n_reporting_partners
# n_partner_gi
# n_partner_pledge_2023
# n_partner2022_2023_plan

# Calculate the percentages
percent_partner_gi = (n_partner_gi / n_reporting_partners) * 100
percent_partner_pledge_2023 = (n_partner_pledge_2023 / n_reporting_partners) * 100
percent_partner2022_2023_plan = (n_partner2022_2023_plan / n_reporting_partners) * 100
# dollars_mobilized_all_plans_n_partners
# percentaje_partner_reporting_adaptation_finance_consolidated



# Set darker colors from the RdPu palette
colors = sns.color_palette("RdPu", n_colors=4)

# Set the size of the figure
fig_width = 8  # 
fig_height = 2  # 

# Create the figure and bar chart
fig_partners_reporting_tools, ax = plt.subplots(figsize=(fig_width, fig_height))

# Plotting the bars
bars = ax.bar(['General\nInformation', 'Pledge\nStatements', 'Resilience-Building\nPlans','Adaptation Finance\nMovilized'],
              [percent_partner_gi, percent_partner_pledge_2023, percent_partner2022_2023_plan,percentaje_partner_reporting_adaptation_finance_consolidated],
              color=colors)

# Adding percentage and absolute number labels on top of each bar
for bar, absolute in zip(bars, [n_partner_gi, n_partner_pledge_2023, n_partner2022_2023_plan, dollars_mobilized_all_plans_n_partners]):
    yval = bar.get_height()
    plt.text(bar.get_x() + bar.get_width()/2, yval + 1, f'{yval:.1f}% ({absolute})', ha='center', va='bottom')

# Set labels and title
ax.set_ylabel('Percentage (%) Partners')
ax.set_xlabel('RtR Reporting Tools', labelpad=15)

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# st.pyplot(fig_partners_reporting_tools)


with tab1:
    st.pyplot(fig_partners_reporting_tools)
        










# LEVEL 2: GENERAL INFORMATION
# #RANKING CHART #1 GENERAL INFORMATION
def create_fig1_gi():
    data = df_gi_summary

    # Set a base figure height and an increment per partner
    base_height = 2
    height_per_partner = 0.1
    fig_height = base_height + height_per_partner * len(data['InitName'])

    fig1_gi, ax = plt.subplots(figsize=(10, fig_height))
    left = np.zeros(len(data['InitName']))

    # Use colormap
    cmap = cm.get_cmap('RdPu')  # Use RdPu colormap
    colors = cmap(data['gi_overall_index'].values / data['gi_overall_index'].values.max())

    # Adjust the bar height here
    bars = ax.barh(data['InitName'], data['gi_overall_index'], color=colors, height=0.5)  
    # ax.set_xlabel('2023 GIS Completeness Index (%)')  # Indicate percentage in the label
    # ax.set_ylabel('Partner Names')
    ax.set_title('2023 GI Completeness Index (%) per Partner')
    ax.tick_params(axis='y', labelsize=8)  # Adjust the size of the y-axis labels (partner names)
        
    # Formatting x-tick labels to display as percentages
    vals = ax.get_xticks()
    ax.set_xticklabels(['{:.0f}%'.format(x) for x in vals])

    # Loop over the bars and add the value as text
    for bar in bars:
        width = bar.get_width()
        ax.text(width,  # Set the x position of the text
                bar.get_y() + bar.get_height() / 2,  # Set the y position of the text
                f'{width:.1f}%',  # Set the text to be displayed
                ha='left',  # horizontal alignment
                va='center',
                fontsize=8)  # vertical alignment

    # Remove top and right spines
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    # Set y-limits to remove excess blank space
    ax.set_ylim(-0.5, len(data['InitName']) - 0.5)
    ax.set_xlim(0, 100)
    ax.set_xticklabels([])
    ax.tick_params(axis='x', length=0)
        
    # Use tight_layout to fit all elements nicely
    plt.tight_layout()
    
    return fig1_gi
fig1_gi = create_fig1_gi()


#Chart2
#Horizontal Chart 2_Multi Index_General Information
general_survey_index_chart_list = ['InitName','partner_info%',
                                   'members_info%','metrics_info%',
                                   'saa_info%','action_cluster_info%',
                                   'direct_impact_info%',]

# Define your data
data = df_gi_summary[general_survey_index_chart_list]

# Make colormap
cmap = plt.get_cmap('RdPu', len(general_survey_index_chart_list)-1)  # minus 1 because 'q9' is not an index

# Set the figure size
fig2_gi, ax = plt.subplots(figsize=(10, len(data['InitName'])/2))  # Adjust the vertical size according to number of partners

# Initial left position is zero
left = np.zeros(len(data['InitName']))

# Loop over all columns except for 'q9' and create a bar for each
for i, column in enumerate(general_survey_index_chart_list[1:]):
        bars = ax.barh(data['InitName'], data[column], left=left, color=cmap(i), label=column)
        left += data[column]

# Add legend and labels
leg = ax.legend(loc='upper left', bbox_to_anchor=(1, 1))  # Moves the legend outside the plot
ax.set_xlabel('Index Value')
ax.set_ylabel('Partner Names')
ax.set_title('Distribution of Different Indexes for Each Partner')

# Set y-limits to remove excess blank space
ax.set_ylim(-0.5, len(data['InitName']) - 0.5)

# Show the plot in Streamlit
# st.pyplot(fig2_gi)

with tab2:
    st.pyplot(fig1_gi)
    st.pyplot(fig2_gi)



## LEVEL 3: PLEDGE STATEMENTS
# #RANKING CHART #3.1 PLEDGE STATEMENT: % OF COMPLETNESS PLEDGE
data = df_pledge_summary

# Set a base figure height and an increment per partner
base_height = 2
height_per_partner = 0.1
fig_height = base_height + height_per_partner * len(data['InitName'])

fig1_pledge, ax = plt.subplots(figsize=(10, fig_height))
left = np.zeros(len(data['InitName']))

# Use colormap
cmap = cm.get_cmap('RdPu')  # Use RdPu colormap
colors = cmap(data['PledgeOverallIdx'].values / data['PledgeOverallIdx'].values.max())

# Adjust the bar height here
bars = ax.barh(data['InitName'], data['PledgeOverallIdx'], color=colors, height=0.5)  
# ax.set_xlabel('2023  Pledge completeness Index (%)')  # Indicate percentage in the label
# ax.set_ylabel('Partner Names')
ax.set_title('2023 PLEDGE Completeness Index (%) per Partner')
ax.tick_params(axis='y', labelsize=8)  # Adjust the size of the y-axis labels (partner names)
    
# Formatting x-tick labels to display as percentages
vals = ax.get_xticks()
ax.set_xticklabels(['{:.0f}%'.format(x) for x in vals])

# Loop over the bars and add the value as text
for bar in bars:
    width = bar.get_width()
    ax.text(width,  # Set the x position of the text
            bar.get_y() + bar.get_height() / 2,  # Set the y position of the text
            f'{width:.1f}%',  # Set the text to be displayed
            ha='left',  # horizontal alignment
            va='center',
            fontsize=8)  # vertical alignment

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# Set y-limits to remove excess blank space
ax.set_ylim(-0.5, len(data['InitName']) - 0.5)
ax.set_xlim(0, 100)
ax.set_xticklabels([])
ax.tick_params(axis='x', length=0)
    
# Use tight_layout to fit all elements nicely
plt.tight_layout()
# st.pyplot(fig1_pledge)


# #CHART 3.2: DETAILS OF COMPLETENESS BY DIFFERENT DIMENSIONS (PLEDGE)
pledge_overall_index_list = ['InitName','ResponderInfoIdx','PrimBenefIdx',
                             'LocalLedComIdx','FinancesIdx',
                             'MetricsExplnIdx',
                             'ClimRiskIdx',
                             'DiagAssessIdx','MemConsultIdx',
                             'PlanConfIdx',]

# Define your data
data = data[pledge_overall_index_list]
    
# Make colormap
cmap = plt.get_cmap('RdPu', len(pledge_overall_index_list)-1)  # minus 1 because 'q9' is not an index

# Set the figure size
fig2_pledge, ax = plt.subplots(figsize=(10, len(data['InitName'])/2))  # Adjust the vertical size according to number of partners

# Initial left position is zero
left = np.zeros(len(data['InitName']))

# Loop over all columns except for 'q9' and create a bar for each
for i, column in enumerate(pledge_overall_index_list[1:]):
        bars = ax.barh(data['InitName'], data[column], left=left, color=cmap(i), label=column)
        left += data[column]

# Add legend and labels
leg = ax.legend(loc='upper left', bbox_to_anchor=(1, 1))  # Moves the legend outside the plot
ax.set_xlabel('Index Value')
ax.set_ylabel('Partner Names')
ax.set_title('Pledge Statement Distribution of Different Indexes for Each Partner')

# Set y-limits to remove excess blank space
ax.set_ylim(-0.5, len(data['InitName']) - 0.5)

# Show the plot in Streamlit
# st.pyplot(fig2_pledge)


# CHART #3.3 DISTRIBUTION OF PLEDGES AMONG DIFFERENT BENEFICIARIES [ABSOLUTE NUMBERS - FULL DETAILS]
def gradient(start_hex, end_hex, n):
    """Generate a gradient of colors between two hex colors."""
    s = np.array([int(start_hex[i:i+2], 16) for i in (1, 3, 5)])
    e = np.array([int(end_hex[i:i+2], 16) for i in (1, 3, 5)])
    return [f"#{int(i[0]):02x}{int(i[1]):02x}{int(i[2]):02x}" for i in np.array([tuple(j) for j in np.linspace(s, e, n)])]

def plot_chart():
    data = {
        'Primary Beneficiaries': ['Individuals', 'Companies', 'Regions', 'Cities', 'Natural Systems'],
        'Pledge Reported': [partners_report_pledge_d_individuals2023, partners_repor_pledge_companies2023, partners_repor_pledge_regions2023, partners_repor_pledge_cities2023, partners_repor_pledge_nat_syst2023],
        'Numbers Reported': [n_partners_reportingnumberof_direct_individuals2023, n_partners_reportingnumberofcompanies2023_pledge, n_partners_reportingnumberofregions2023_pledge, n_partners_reportingnumberofcities2023_pledge, n_partners_reportingnumberofnaturalsyst2023_pledge],
        'Individuals Benefited': [0, n_partners_reportingnumberofindiv_companies2023, n_partners_reportingnumberofindiv_regions2023, n_partners_reportingnumberofindiv_cities2023, n_partners_reportingnumberofindiv_natsyst2023]
    }

    labels = data['Primary Beneficiaries']
    barWidth = 0.25
    r1 = np.arange(len(data['Pledge Reported']))
    r2 = [x + barWidth for x in r1]
    r3 = [x + barWidth for x in r2]

    colors = gradient("#FF37D5", "#112E4D", 3)

    fig_pledge_full_detail, ax = plt.subplots(figsize=(12, 8))

    # Create bars  ##REVIEW THE NAMES OF THE LABELS
    bars1 = ax.barh(r1, data['Pledge Reported'], color=colors[0], height=barWidth, label='Pledge on Primary Beneficiary') 
    bars2 = ax.barh(r2, data['Numbers Reported'], color=colors[1], height=barWidth, label='Numbers of Primary Beneficiary Reported')
    bars3 = ax.barh(r3, data['Individuals Benefited'], color=colors[2], height=barWidth, label='Individuals Benefited Reported')

    # Display the absolute numbers next to the bars
    for idx, bars in enumerate([bars1, bars2, bars3]):
        for bar, label in zip(bars, labels):
            width = bar.get_width()
            # Avoid placing '0' for Individuals group in the Individuals Benefited bar set
            if not (width == 0 and label == "Individuals" and idx == 2):
                ax.text(width + 0.5, bar.get_y() + bar.get_height()/2, '%d' % int(width), ha='center', va='bottom')

    # Title & Subtitle
    # plt.title('2023 RtR Pledges Overview')
    # plt.xlabel("Number of RtR Initiatives Reporting Pledges for Each Primary Beneficiary")

    # Y axis
    plt.yticks([r + barWidth for r in range(len(data['Pledge Reported']))], labels)

    plt.tight_layout()
    plt.legend()
    # st.pyplot(fig)
    return fig_pledge_full_detail 
fig_pledge_full_detail = plot_chart()

# CHART #3.4 Pledge Overview Simple  [ABSOLUTE NUMBERS]
def gradient(start_hex, end_hex, n):
    """Generate a gradient of colors between two hex colors."""
    s = np.array([int(start_hex[i:i+2], 16) for i in (1, 3, 5)])
    e = np.array([int(end_hex[i:i+2], 16) for i in (1, 3, 5)])
    return [f"#{int(i[0]):02x}{int(i[1]):02x}{int(i[2]):02x}" for i in np.array([tuple(j) for j in np.linspace(s, e, n)])]

def plot_chart_single():
    data = {
        'Primary Beneficiaries': ['Individuals', 'Companies', 'Regions', 'Cities', 'Natural Systems'],
        'Pledge Reported': [partners_report_pledge_d_individuals2023, partners_repor_pledge_companies2023, partners_repor_pledge_regions2023, partners_repor_pledge_cities2023, partners_repor_pledge_nat_syst2023],
    }

    # Sort data alphabetically by 'Primary Beneficiaries'
    sorted_indices = np.argsort(data['Primary Beneficiaries'])
    sorted_labels = np.array(data['Primary Beneficiaries'])[sorted_indices]
    sorted_values = np.array(data['Pledge Reported'])[sorted_indices]

    barWidth = 0.6
    r1 = np.arange(len(sorted_labels))

    # Choose your color gradient here or set a single color
    colors = gradient("#FF37D5", "#112E4D", len(sorted_labels))

    fig_pledges_primary_beneficiary, ax = plt.subplots(figsize=(11, 5))
  
    # Create vertical bars
    bars1 = ax.bar(r1, sorted_values, color=colors, width=barWidth)
    
    # Display the absolute numbers above the bars
    for bar in bars1:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, height + 0.5, '%d' % int(height), ha='center', va='bottom', fontsize=14)

    plt.ylabel("Number of Partners", fontsize=16)
    plt.xlabel("Beneficiary Categories (Pledge)", fontsize=16)
    
    # plt.gcf().subplots_adjust(bottom=0.6) 
    # annotation_text = f"Out of {n_partner_pledge_2023} Partners reporting Pledge Statements"
    # fig_pledges_primary_beneficiary.text(0.95, -0.08, annotation_text, ha='right', va='bottom', fontsize=12, color='black', transform=fig_pledges_primary_beneficiary.transFigure)  # Adjust Y coordinate slightly above the bottom

    # X axis
    plt.xticks(r1, sorted_labels, fontsize=14, rotation=0)  # Rotate labels for better visibility
    ax.tick_params(axis='y', labelsize=14)
    ax.tick_params(axis='x', length=5)
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)
    
    # X axis: Setting integer ticks
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True)) 
    ax.set_ylim(0, 29)
    
    plt.tight_layout()
    plt.legend(fontsize=12)
    return fig_pledges_primary_beneficiary 

# # Usage
fig_pledges_primary_beneficiary = plot_chart_single()
# st.write('Pledge Overview Correcto Numeros Obsolutos')
# st.pyplot(fig_pledges_primary_beneficiary)


## CHART 3.5 PLEDGE IN PORCENTAJE

def gradient(start_hex, end_hex, n):
    """Generate a gradient of colors between two hex colors."""
    s = np.array([int(start_hex[i:i+2], 16) for i in (1, 3, 5)])
    e = np.array([int(end_hex[i:i+2], 16) for i in (1, 3, 5)])
    return [f"#{int(i[0]):02x}{int(i[1]):02x}{int(i[2]):02x}" for i in np.array([tuple(j) for j in np.linspace(s, e, n)])]


n_total = n_reporting_partners
def plot_chart_single():
    # Your data goes here:
    data = {
        'Primary Beneficiaries': ['Individuals', 'Companies', 'Regions', 'Cities', 'Natural Systems'],
        'Pledge Reported': [partners_report_pledge_d_individuals2023, partners_repor_pledge_companies2023, partners_repor_pledge_regions2023, partners_repor_pledge_cities2023, partners_repor_pledge_nat_syst2023],
    }

    # Convert absolute values to percentages
    data['Pledge Reported'] = [(value / n_total) * 100 for value in data['Pledge Reported']]

    # Sort data alphabetically by 'Primary Beneficiaries'
    sorted_indices = np.argsort(data['Primary Beneficiaries'])
    sorted_labels = np.array(data['Primary Beneficiaries'])[sorted_indices]
    sorted_values = np.array(data['Pledge Reported'])[sorted_indices]

    barWidth = 0.6
    r1 = np.arange(len(sorted_labels))

    fig_pledges_beneficiaries_percentg, ax = plt.subplots(figsize=(11, 5))
    
    colors = gradient("#FF37D5", "#112E4D", 5)

    # Create vertical bars
    bars1 = ax.bar(r1, sorted_values, color=colors, width=barWidth)  # Choose a color or use your gradient function

    # Display the percentages above the bars
    for bar in bars1:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, height + 0.5, '%.1f%%' % height, ha='center', va='bottom', fontsize=14)

    plt.ylabel("% Partners", fontsize=16)
    plt.xlabel("Beneficiary Categories (Pledge)", fontsize=16)

    # X axis
    plt.xticks(r1, sorted_labels, fontsize=14, rotation=0)  # Rotate labels for better visibility
    ax.tick_params(axis='y', labelsize=14)
    ax.tick_params(axis='x', length=5)
    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    # Adjust the y-axis limits and ticks to accommodate percentages
    ax.set_ylim(0, 100)
    ax.yaxis.set_major_formatter(ticker.PercentFormatter())

    plt.tight_layout()
    plt.legend(fontsize=12)
    return fig_pledges_beneficiaries_percentg 

fig_pledges_beneficiaries_percentg = plot_chart_single()
# st.write('Pledge Overview Percentage ')
# st.pyplot(fig_pledges_beneficiaries_percentg)


with tab3:
    st.write("LEVEL OF COMPLETNESS")
    st.pyplot(fig1_pledge)
    st.pyplot(fig2_pledge)
    st.write("DISTRIBUTION OF PLEDGES BY BENEFICIARIES")
    st.pyplot(fig_pledges_beneficiaries_percentg)
    st.pyplot(fig_pledges_primary_beneficiary)
    st.pyplot(fig_pledge_full_detail)
    


# LEVEL 4) PLAN (ALL: CONSOLIDATED & ONLY 2023 INFO)

## CHART 4.1 PLAN CONSOLIDATED (2022 AND 2023) FULL
### Chart Number of Plan Consolidated: After the dataset of 2023 and 2022 are united based on 2023 as primary information
# Chart of different 2022 and 2023 number of charts
# Grouping and plotting data
grouped = df_2022_2023_plan_consolidated.groupby(['InitName', 'PlanYear']).size().unstack().fillna(0)

fig_plan2022_2023_full, ax = plt.subplots(figsize=(10, 7))
width = 0.4
ind = np.arange(len(grouped))

ax.barh(ind - width/2, grouped.get('2022', 0), width, color='#112E4D', label='2022')
ax.barh(ind + width/2, grouped.get('2023', 0), width, color='#FF37D5', label='2023')

ax.set_xlabel('Number of Resilience-Building Plans')
ax.set_title('Resilience-Building Plans by RtR Initiatives for 2022 and 2023')
ax.set_yticks(ind)
ax.set_yticklabels(grouped.index)
ax.legend()
ax.xaxis.set_major_locator(MaxNLocator(integer=True))
ax.set_axisbelow(True)
ax.xaxis.grid(color='gray', linestyle='dashed', which='both', linewidth=0.5)
plt.tight_layout()
# st.pyplot(fig_plan2022_2023_full)


## CHART 4.2 PLAN CONSOLIDATED (UPDATED)
## Chart of only one colums with the number of plan per partner in 2023
# Chart of different 2022 and 2023 number of charts
# Grouping and plotting data
grouped = df_2022_2023_plan_consolidated.groupby(['InitName', 'PlanYear']).size().unstack().fillna(0)

fig_plan2022_2023_simple_consolidated, ax = plt.subplots(figsize=(10, 7))
width = 0.4
ind = np.arange(len(grouped))

ax.barh(ind - width/2, grouped.get('2022', 0), width, color='#112E4D', label='Submitted in 2022')
ax.barh(ind + width/2, grouped.get('2023', 0), width, color='#FF37D5', label='Updated in 2023')

ax.set_xlabel('Number of Resilience-Building Plans')
ax.set_title('Resilience-Building Plans by RtR Initiatives')
ax.set_yticks(ind)
ax.set_yticklabels(grouped.index)
ax.legend()
ax.xaxis.set_major_locator(MaxNLocator(integer=True))
ax.set_axisbelow(True)
ax.xaxis.grid(color='gray', linestyle='dashed', which='both', linewidth=0.5)
plt.tight_layout()
# st.pyplot(fig_plan2022_2023_simple_consolidated)


# n_plans_reporting_metrics_plan_consolidated  # Numbers of plans whit information in metrics about individuals
# n_2022_2023_plans_individuals # Number of plans per primary beneficiay
# n_2022_2023_plans_companies # Number of plans per primary beneficiay
# n_2022_2023_plans_regions # Number of plans per primary beneficiay
# n_2022_2023_plans_cities # Number of plans per primary beneficiay
# n_2022_2023_plans_natsyst # Number of plans per primary beneficiay


# Chart #4.3 Plan Overview Simple   
def gradient(start_hex, end_hex, n):
    """Generate a gradient of colors between two hex colors."""
    s = np.array([int(start_hex[i:i+2], 16) for i in (1, 3, 5)])
    e = np.array([int(end_hex[i:i+2], 16) for i in (1, 3, 5)])
    return [f"#{int(i[0]):02x}{int(i[1]):02x}{int(i[2]):02x}" for i in np.array([tuple(j) for j in np.linspace(s, e, n)])]


def plot_chart_single_plan():
    # Your data goes here:
    data = {
        'Primary Beneficiaries': ['Individuals', 'Companies', 'Regions', 'Cities', 'Natural Systems'],
        'Plan Reported': [n_2022_2023_plans_individuals, n_2022_2023_plans_companies, n_2022_2023_plans_regions, n_2022_2023_plans_cities, n_2022_2023_plans_natsyst],
    }

    # Sort data alphabetically by 'Primary Beneficiaries'
    sorted_indices = np.argsort(data['Primary Beneficiaries'])
    sorted_labels = np.array(data['Primary Beneficiaries'])[sorted_indices]
    sorted_values = np.array(data['Plan Reported'])[sorted_indices]

    barWidth = 0.6
    r1 = np.arange(len(sorted_labels))

    # Choose your color gradient here or set a single color
    colors = gradient("#FF37D5", "#112E4D", len(sorted_labels))

    fig_plans_primary_beneficiary, ax = plt.subplots(figsize=(11, 5))

    # Create bars
    bars1 = ax.bar(r1, sorted_values, color=colors, width=barWidth)

    # Display the absolute numbers above the bars
    for bar in bars1:
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2, height + 0.5, '%d' % int(height), ha='center', va='bottom', fontsize=14)

    plt.xlabel("Beneficiary Categories (Plan)", fontsize=16)
    plt.ylabel("Number of Plan Submissions", fontsize=16)

    # plt.gcf().subplots_adjust(bottom=0.6)
    # annotation_text = f"Out of {n_total_plans_consolidated} plans from {n_partner2022_2023_plan} Partners (Source: Plan Survey 2022 and 2023)"
    # fig_plans_primary_beneficiary.text(0.95, -0.08, annotation_text, ha='right', va='bottom', fontsize=12, color='black', transform=fig_plans_primary_beneficiary.transFigure)

    # X axis
    plt.xticks(r1, sorted_labels, fontsize=14, rotation=0)
    ax.tick_params(axis='y', labelsize=14)
    ax.tick_params(axis='x', length=5)
    ax.spines['right'].set_visible(False)   
    ax.spines['top'].set_visible(False)

    # Y axis: Setting integer ticks
    ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True))
    # ax.set_ylim(0, 29)

    plt.tight_layout()
    plt.legend(fontsize=12)
    return fig_plans_primary_beneficiary 

# Usage
fig_plans_primary_beneficiary = plot_chart_single_plan()
# st.write('RtR Plans on Primary Beneficiary - Overview (simple)')
# st.pyplot(fig_plans_primary_beneficiary)


## #RANKING CHART #4.4 PLAN (ONLY 2023)
data = df_2023_plan_summary

base_height = 2
height_per_partner = 0.1
fig_height = base_height + height_per_partner * len(data['plan_short_name'])

fig1_plan, ax = plt.subplots(figsize=(10, fig_height))
left = np.zeros(len(data['plan_short_name']))

# Use colormap
cmap = cm.get_cmap('RdPu')  # Use RdPu colormap
colors = cmap(data['plan_overall_idx'].values / data['plan_overall_idx'].values.max())


# Adjust the bar height here
bars = ax.barh(data['plan_short_name'], data['plan_overall_idx'], color=colors, height=0.5)  
# ax.set_xlabel('2023 GIS Completeness Index (%)')  # Indicate percentage in the label
# ax.set_ylabel('Partner Names')
ax.set_title('2023 PLAN Completeness Index (%) per Partner')
ax.tick_params(axis='y', labelsize=8)  # Adjust the size of the y-axis labels (partner names)
    
# Formatting x-tick labels to display as percentages
vals = ax.get_xticks()
ax.set_xticklabels(['{:.0f}%'.format(x) for x in vals])

# Loop over the bars and add the value as text
for bar in bars:
    width = bar.get_width()
    ax.text(width,  # Set the x position of the text
            bar.get_y() + bar.get_height() / 2,  # Set the y position of the text
            f'{width:.1f}%',  # Set the text to be displayed
            ha='left',  # horizontal alignment
            va='center',
            fontsize=8)  # vertical alignment

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# Set y-limits to remove excess blank space
ax.set_ylim(-0.5, len(data['plan_short_name']) - 0.5)
ax.set_xlim(0, 100)
ax.set_xticklabels([])
ax.tick_params(axis='x', length=0)
    
# Use tight_layout to fit all elements nicely
plt.tight_layout()
# st.pyplot(fig1_plan)


# #Horizontal Chart 4.5_Multi Index (ONLY 2023)
data = df_2023_plan_summary
list_to_plan_chart_all_index = ['plan_short_name','plan_identif_idx','description_problem_plan_idx','saa_idx','ra_idx','p_benefic_overall_idx_plan','governance_plan_idx','barriers_assumpt_plan_idx']

# Define your data
data = data[list_to_plan_chart_all_index]
    
# Make colormap
cmap = plt.get_cmap('RdPu', len(list_to_plan_chart_all_index)-1)  # minus 1 because 'q9' is not an index

# Set the figure size
fig2_plan, ax = plt.subplots(figsize=(10, len(data['plan_short_name'])/2))  # Adjust the vertical size according to number of partners

# Initial left position is zero
left = np.zeros(len(data['plan_short_name']))

# Loop over all columns except for 'q9' and create a bar for each
for i, column in enumerate(list_to_plan_chart_all_index[1:]):
        bars = ax.barh(data['plan_short_name'], data[column], left=left, color=cmap(i), label=column)
        left += data[column]

# Add legend and labels
leg = ax.legend(loc='upper left', bbox_to_anchor=(1, 1))  # Moves the legend outside the plot
ax.set_xlabel('Index Value')
ax.set_ylabel('Partner Names')
ax.set_title('PLAN Distribution of Different Indexes for Each Partner')

# Set y-limits to remove excess blank space
ax.set_ylim(-0.5, len(data['plan_short_name']) - 0.5)

# Show the plot in Streamlit
# st.pyplot(fig2_plan)



## CHART % PLANES GENERAL 
# name of the chart: fig_plans_percentaje_submissions

# Variables
# n_reporting_partners  #total reporting partners (all)
# n_partner2022_2023_plan #total partners reporting plan (all)
n_no_plan_reporting_partners = n_reporting_partners - n_partner2022_2023_plan
# n_partner_plan_only2023 #Partners reporting and updating plans in 2023 
n_partner_plan_only2022 = n_partner2022_2023_plan - n_partner_plan_only2023

# # Variables needed
# n_reporting_partners
# n_no_plan_reporting_partners
# n_partner2022_2023_plan
# n_partner_plan_only2023
# n_partner_plan_only2022

# n_no_plan_reporting_partners
# n_partner_plan_only2023
# n_partner_plan_only2022


# n_no_plan_reporting_partners
# n_partner2022_2023_plan


## CHART % PARTNERS REPORTING PLAN (SIMPLE)

# Calculate the percentages
percent_no_plan = (n_no_plan_reporting_partners / n_reporting_partners) * 100
percent_plan = (n_partner2022_2023_plan / n_reporting_partners) * 100

# Set darker colors from the RdPu palette
bar_colors = sns.color_palette("RdPu", n_colors=3)
bar_height = 0.1  # The height of the bar to make it narrow

# Set the size of the figure
fig_width = 8  # You can adjust this as needed
fig_height = 2  # You can adjust this as needed

# Create the figure with a defined size
fig_plans_percentaje_submissions_simple, ax = plt.subplots(figsize=(fig_width, fig_height))

# Plotting the stacked horizontal bar with the 'height' parameter
bars = ax.barh(['Partners'], [percent_plan], color=bar_colors[2], label='Plan(s) Submitted', height=bar_height)
ax.barh(['Partners'], [percent_no_plan], left=[percent_plan], color=bar_colors[0], label='No Plan Submitted', height=bar_height)

# Position the legend in the upper right corner
ax.legend(loc='upper right')    

text_position = bar_height / 50

ax.text(percent_plan / 2, text_position, f'{percent_plan:.1f}% ({n_partner2022_2023_plan})', va='center', ha='center', color='black')
ax.text(percent_plan + percent_no_plan / 2, text_position, f'{percent_no_plan:.1f}% ({n_no_plan_reporting_partners})', va='center', ha='center', color='black')

# Set labels and title
ax.set_xlabel('Percentage (%)')

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# ax.set_title('Plan Submission Percentages')

# Show the figure in the Streamlit app
# st.pyplot(fig_plans_percentaje_submissions_simple)






## CHART % PARTNERS REPORTING PLAN (DETAILED)
# Calculate the percentages
percent_no_plan = (n_no_plan_reporting_partners / n_reporting_partners) * 100  # No plan submitted
percent_plan2022 = (n_partner_plan_only2022 / n_reporting_partners) * 100     # Plan submitted in 2022
percent_plan_updated2023 = (n_partner_plan_only2023 / n_reporting_partners) * 100 # Plans submitted in 2023

# Set darker colors from the RdPu palette
bar_colors = sns.color_palette("RdPu", n_colors=3)
bar_height = 0.1  # The height of the bar to make it narrow

# Create the figure with a defined size
fig_width = 9
fig_height = 2
fig_plans_percentaje_submissions_detailed, ax = plt.subplots(figsize=(fig_width, fig_height))

# Plotting the stacked horizontal bar
bars = ax.barh(['Partners'], [percent_plan2022], color=bar_colors[2], label='Submitting Plans in 2022', height=bar_height)
ax.barh(['Partners'], [percent_plan_updated2023], left=[percent_plan2022], color=bar_colors[1], label='Plans Updated/Submitted in 2023', height=bar_height)
ax.barh(['Partners'], [percent_no_plan], left=[percent_plan2022 + percent_plan_updated2023], color=bar_colors[0], label='No Plans Submitted', height=bar_height)

# Adjust layout to make space for the legend below the chart
plt.subplots_adjust(bottom=0.25)  # Adjust this value as needed

# Position the legend below the plot area
ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.35), fancybox=True, shadow=False, ncol=3)


# Adding value labels with both percentage and absolute number
text_position = bar_height / 50
ax.text(percent_plan2022 / 2, text_position, f'{percent_plan2022:.1f}% ({n_partner_plan_only2022})', va='center', ha='center', color='black')
ax.text(percent_plan2022 + percent_plan_updated2023 / 2, text_position, f'{percent_plan_updated2023:.1f}% ({n_partner_plan_only2023})', va='center', ha='center', color='black')
ax.text(percent_plan2022 + percent_plan_updated2023 + percent_no_plan / 2, text_position, f'{percent_no_plan:.1f}% ({n_no_plan_reporting_partners})', va='center', ha='center', color='black')

# Set labels
ax.set_xlabel('Percentage (%)')

# Remove top and right spines
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# Show the figure in the Streamlit app
# st.pyplot(fig_plans_percentaje_submissions_detailed)



with tab4:
    st.write('CONSOLIDATION OF PLANS 2022 AND 2023')
    st.pyplot(fig_plans_percentaje_submissions_simple)
    st.pyplot(fig_plans_percentaje_submissions_detailed)
    # st.pyplot(fig_plan2022_2023_full)
    st.pyplot(fig_plan2022_2023_simple_consolidated)
    st.pyplot(fig_plans_primary_beneficiary)

    st.write('DETAILS ONLY ON UPDATED 2023 PLANS')
    st.pyplot(fig1_plan)
    st.pyplot(fig2_plan)

# # # ______________________________________________
# # # ______________________________________________
# # # END BLOCK GI PLEDGE PLAN DESCRIPTION
# # # ______________________________________________
# # # ______________________________________________










# ###__________________________________________________________________________________
# ###__________________________________________________________________________________
# ## PLAN ACTION TYPES PLAN CONSOLIDADO
# ###__________________________________________________________________________________
# ###__________________________________________________________________________________

# list_action_areas = ['x14','x15','x16','x17','x18','x19','x20','x21','x22',]
# # df_2022_2023_plan_consolidated[list_action_areas] 

# update_action_types = {'x14':'Climate risk and vulnerability assessments, disclosure and monitoring actions',
# 'x15':'Access to early warning systems and development of early actions.',
# 'x16':'Preparedness with contingency plans and emergency response',
# 'x17':'Establishment of effective governance to manage climate risks accompanied by human and institutional capacity-building',
# 'x18':'Nature-based solutions used to reduce risks',
# 'x19':'Climate-proofing of infrastructure and services',
# 'x21':'Sharing of knowledge and best practices on climate risk management',
# 'x20':'Risk transfer: insurance and social protection instruments',
# 'x22':'Increase in the volume, quality and access of public and private finance to invest in resilience',
# }

# # Counting occurrences where each action area is not False
# ac_counts = {
#     'x14': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x14'] != False]['x14'].count(),
#     'x15': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x15'] != False]['x15'].count(),
#     'x16': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x16'] != False]['x16'].count(),
#     'x17': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x17'] != False]['x17'].count(),
#     'x18': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x18'] != False]['x18'].count(),
#     'x19': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x19'] != False]['x19'].count(),
#     'x20': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x20'] != False]['x20'].count(),
#     'x21': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x21'] != False]['x21'].count(),
#     'x22': df_2022_2023_plan_consolidated[df_2022_2023_plan_consolidated['x22'] != False]['x22'].count()
# }
# # Creating a DataFrame from the counts
# action_types_df = pd.DataFrame(list(ac_counts.items()), columns=['Action Type', 'Count'])

# # Mapping action areas to descriptive text
# action_types_df['Action Type'] = action_types_df['Action Type'].map(update_action_types)
# action_types_df = action_types_df.sort_values(by='Count', ascending=True)


# # #chart 4
# # Define a function to split long labels into multiple lines
# def split_label(label, max_chars=60):
#     words = label.split()
#     split_labels = ['']
#     current_line = 0

#     for word in words:
#         # Check if adding the next word would exceed the max_chars limit
#         if len(split_labels[current_line]) + len(word) > max_chars:
#             # Check if we are already at 2 lines, if so, append the rest to the last line
#             if current_line == 1:
#                 split_labels[current_line] += ' ' + word
#             else:
#                 # Move to the next line if not
#                 current_line += 1
#                 split_labels.append(word)
#         else:
#             # Add the word with a space if the line is not empty
#             if split_labels[current_line]:
#                 split_labels[current_line] += ' ' + word
#             else:
#                 split_labels[current_line] += word

#     return '\n'.join(split_labels)

# # Apply the function to the labels
# action_types_df['Action Type Three Lines'] = action_types_df['Action Type'].apply(split_label)


# # Creating the horizontal bar chart
# fig_height = len(action_types_df) * 0.6  # Adjust the figure height as needed
# fig, ax = plt.subplots(figsize=(10, fig_height))

# # Set the color of the bars using the 'color' parameter
# bars = ax.barh(action_types_df['Action Type'], action_types_df['Count'], color='#FF37D5')
# ax.set_xlabel('Number of Plans')
# ax.set_ylabel('Marrakech Action Types')
# ax.set_title('Number of Plans by Marrakech Action Types')

# # Set the labels for y-axis with split text
# # Make sure 'Action Type Three Lines' is created in your DataFrame, similar to the split_label function you mentioned
# ax.set_yticks(action_types_df.index)
# ax.set_yticklabels(action_types_df['Action Type Three Lines'], fontsize=10)

# # Adjust the layout to prevent overlapping labels
# plt.tight_layout()

# # annotation_text = f"Out of {n_total_plans_consolidated} plans from {n_partner2022_2023_plan} Partners (Source: Plan Survey 2022 and 2023)"
# # fig.text(0.95, 0.00, annotation_text, ha='right', va='bottom', fontsize=8, color='black', transform=fig.transFigure)

# # Add values on each bar
# for bar in bars:
#     label_x_pos = bar.get_width()
#     label_y_pos = bar.get_y() + bar.get_height() / 2
#     ax.text(label_x_pos, label_y_pos, s=f'  {int(bar.get_width())}', va='center', ha='left', fontsize=8)

# # Display the plot in Streamlit
# st.pyplot(fig)

###__________________________________________________________________________________
###__________________________________________________________________________________
## END ACTION TYPES PLAN CONSOLIDADO
###__________________________________________________________________________________
###__________________________________________________________________________________





# #____________________________________________
# #___________________________________________
# #___________________________________________
# #___________________________________________
# # # BLOQUE PRIORITY GROUPS
# # #___________________________________________
# #___________________________________________
# #___________________________________________

# # Chart #3 Priority Groups Action GI
# st.subheader("PRIORITY GROUPS")
# st.caption("**Q6. Do the actions of your initiative related to RtR target any of these priority groups? (select all that apply)**")

# # list of columns with Priority Groups
# pg_list = ['q34', 'q35', 'q36', 'q37', 'q38', 'q39', 'q40', 'q41', 'q42']

# # Use .any() function to check if any value in the row is True (i.e., is not NaN), then sum up these True values
# pg_all = df_gi[pg_list].notna().any(axis=1).sum()
# pg_others = df_gi['q43'].count()
# pg_noneofabove = df_gi['q33'].count()

# st.caption("Q6. Number of Partners reporting priority groups within the provided list: "+str(pg_all))
# st.caption("Q6. Number of Partners reporting 'Others' as priority groups: "+str(pg_others))
# st.caption("Q6. Number of Partners reporting  'None of the Above' as priority groups: "+str(pg_noneofabove))

# ## CHART PRIORITY GROUP %

# list = ['q34', 'q35', 'q36', 'q37', 'q38', 'q39', 'q40', 'q41', 'q42', 'q43']
# df_gi2 = df_gi[list]
# total = n_reporting_partners  # Total number

# # Count the values and calculate percentages
# pg0 = (df_gi2["q34"].count() / total) * 100  # Women and girls
# pg1 = (df_gi2["q35"].count() / total) * 100  # LGBTQIA+
# pg2 = (df_gi2["q36"].count() / total) * 100  # Elderly
# pg3 = (df_gi2["q37"].count() / total) * 100  # Children and Youth
# pg4 = (df_gi2["q38"].count() / total) * 100  # Disabled
# pg5 = (df_gi2["q39"].count() / total) * 100  # Indigenous or traditional communities
# pg6 = (df_gi2["q40"].count() / total) * 100  # Racial, ethnic and/or religious minorities
# pg7 = (df_gi2["q41"].count() / total) * 100  # Refugees
# pg8 = (df_gi2["q42"].count() / total) * 100  # Low income communities
# pg9 = (df_gi2["q43"].count() / total) * 100  # Other


# s_df_gi2 = pd.DataFrame(dict(
# 	r_values=[pg0, pg1, pg2, pg3, pg4, pg5, pg6, pg7, pg8, pg9],
# 	theta_values=['Women and girls','LGBTQIA+','Elderly','Children and Youth','Disabled','Indigenous or traditional communities','Racial, ethnic and/or religious minorities','Refugees','Low income communities','Other']))

# # Data
# r_values = [pg0, pg1, pg2, pg3, pg4, pg5, pg6, pg7, pg8,pg9]
# theta_values = ['Women and girls','LGBTQIA+','Elderly','Children and Youth','Disabled','Indigenous or traditional communities','Racial, ethnic and/or religious minorities','Refugees','Low income communities','Other']

# # Create the plot
# fig_pp = go.Figure(data=go.Scatterpolar(
# 	r=r_values,
# 	theta=theta_values,
# 	mode='lines',
# 	fill='toself',
# 	line=dict(color='#FF37D5', width=1)
# ))

# fig_pp.update_layout(
#     polar=dict(
#         radialaxis=dict(
#             tickfont=dict(color='#112E4D'),
#             range=[0, 100],
#             title=dict(  # Adding the title for the radial axis
#                 text="(%) Partners",  # The title text
#                 font=dict(size=14, color='black')  # Customize the font size and color
#             )
#         ),
#         angularaxis=dict(
#             tickfont=dict(size=12, color='black')
#         )
#     )
# )
# # # Update layout
# # fig_pp.update_layout(
# # 	polar=dict(
# # 		radialaxis=dict(
# # 			tickfont=dict(color='#112E4D'),
# # 			range=[0, 29]
# # 		)
# # 	),
# # 	# title="RtR Partners' Targeted Priority Groups",
# 	# annotations=[go.layout.Annotation(
# 	# 	x=1,
# 	# 	y=0,
# 	# 	text='Out of '+str(n_pg)+' RtR Partners reporting about the Priority Groups',
# 	# 	showarrow=False,
# 	# 	yshift=-60
# 	# )]
# # )
# st.write("RtR Partners' Targeted Priority Groups (%) OFICIAL")
# st.write(fig_pp)

# # #___________________________________________
# # #___________________________________________
# # #___________________________________________
# # #___________________________________________
# # # END BLOCK PRIORITY GROUPS
# # #___________________________________________
# # #___________________________________________
# # #___________________________________________
# # #___________________________________________

















# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# # # ______________________________________________
# # # ______________________________________________
# # # BLOCK SAA_______________________________
# # # ______________________________________________
# # # ______________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________


# st.header("SECTION 4. SHARM-EL-SHEIKH ADAPTATION AGENDA (SAA) IMPACT SYSTEMS AND CROSS CUTTING ENABLERS")
# st.markdown("This section is devoted to the RtR Partner's involvement in the SAA Impact Systems and any associated cross-cutting enablers.")
# # st.caption("DATA: "+type_info)
# st.subheader("4.1 - RtR PARTNER RELATED TO SAA IMPACT SYSTEMS AND CROSS CUTTING ENABLERS")


# # Data set to plot making: 
# impact_systems = {
# 'Finance: Cross-cutting': 'q147',
# 'Planning and Policy: Cross-cutting': 'q137',
# 'Water and Nature': 'q85',
# 'Infrastructure': 'q122',
# 'Coastal and Oceans': 'q111',
# 'Human Settlements': 'q98',
# 'Health': 'q135',
# 'Food and Agriculture': 'q76'
# }

# saa_to_explanation = {
# 	'q76': 'q77', 
# 	'q85': 'q86',
# 	'q98': 'q99',
# 	'q111':'q112',
# 	'q122':'q123',
# 	'q135':'q136',
# 	'q137':'q138',
# 	'q147':'q148'
# }

# impact_system_list = []
# partner_list = []
# relationship_list = []
# explanation_list = []

# for impact_system, column_name in impact_systems.items():
# 	partners = df_gi['InitName']
# 	relationships = df_gi[column_name]
# 	if column_name in saa_to_explanation:
# 		explanations = df_gi[saa_to_explanation[column_name]]
# 	else:
# 		explanations = [pd.NA] * len(partners)
	
# 	impact_system_list.extend([impact_system] * len(partners))
# 	partner_list.extend(partners)
# 	relationship_list.extend(relationships)
# 	explanation_list.extend(explanations)

# new_df_gi = pd.DataFrame({
# 	'Partner Name': partner_list,
# 	'SAA Impact Systems': impact_system_list,
# 	'Relationship': relationship_list,
# 	'Explanation (optional)': explanation_list
# })

# new_df_gi['SAA Impact Systems'] = pd.Categorical(new_df_gi['SAA Impact Systems'], categories=impact_systems.keys(), ordered=True)
# new_df_gi['relation_yes'] = new_df_gi['Relationship'].apply(lambda x: 'Yes' if pd.notna(x) and x != 'Unsure' and x != 'Not Related' else pd.NA)

# #Data Set SAA Summary with RtR Partners & Number of SAA Taskforces
# n_partner_with_saa = new_df_gi[new_df_gi['relation_yes'] == 'Yes']
# n_partner_with_saa = n_partner_with_saa['Partner Name'].value_counts().reset_index()
# n_partner_with_saa.rename(columns={'count': 'N° SAA Taskforces'}, inplace=True)
# n_partner_with_saa.rename(columns={'Partner Name': 'RtR PARTNER NAME'}, inplace=True)
# summary_partner_saa = n_partner_with_saa
# new_order_saa = ['RtR PARTNER NAME','N° SAA Taskforces']


# summary_partner_saa = summary_partner_saa[new_order_saa].sort_values('N° SAA Taskforces', ascending=False)
# summary_partner_saa.rename(columns={'SURVEY RESPONSE STATUS': 'Reliability','RtR PARTNER NAME':'Name'}, inplace=True)
# summary_partner_saa['Related to SAA'] = np.where(summary_partner_saa['N° SAA Taskforces'].notna(), "Yes", "Pending Confirmation")
# # st.write(summary_partner_saa)
	
# # Compute counts of unique values
# value_counts = summary_partner_saa['Related to SAA'].value_counts()
# # Compute percentages
# percentages = value_counts / len(summary_partner_saa) * 100
# # Combine the counts and percentages into a DataFrame for display
# stats_SAA_df_gi = pd.DataFrame({
# 	'Is the Partner related to an SAA Taskforce?': value_counts.index,
# 	'Count': value_counts.values,
# 	'Percentage (%)': percentages.values
# })
# stats_SAA_df_gi = stats_SAA_df_gi.reset_index(drop=True)


	
# #Making a SUNBURST Chart
# # Create a dictionary for assigning specific colors to the unique values in 'Related to SAA'
# color_dict = {"Yes": "#FF37D5", "Pending Confirmation": "#DCE3E5"}
# # Dataset to chart
# counts = summary_partner_saa.groupby(['Related to SAA', 'Name']).size().reset_index(name='Count')
# # Create sunburst chart
# fig = px.sunburst(counts, path=['Related to SAA', 'Name'], values='Count', color='Related to SAA',
# 				color_discrete_map=color_dict)
# # Add title
# fig.update_layout(title="RtR Partners related to SAA Impact Systems")

# #Displaying the number of Partners that has a confirmed relationship with SAA Taskforces
# number_partner_saa = summary_partner_saa['N° SAA Taskforces'].notna().sum()
# number_partner_saa_pending = summary_partner_saa['N° SAA Taskforces'].isna().sum()

# col1, col2, col3= st.columns(3)
# col1.metric('RtR Partners related to SAA Taskforces', str(number_partner_saa))
# # col2.metric('Pending Confirmation', str(number_partner_saa_pending))

# col1, col2= st.columns(2)
# # col1.plotly_chart(fig)
# col2.write(stats_SAA_df_gi)
# # csv = convert_df_gi(stats_SAA_df_gi)
# # col2.download_button(
# # 						label="Download data as CSV",
# # 						data=csv,
# # 						file_name='RtR_SAA_relation_'+datetofiles+'.csv',
# # 						mime='text/csv',
# # 					)


# st.subheader("4.2 - NUMBER OF SAA IMPACT SYSTEMS PER RtR PARTNER")

# #Making a Chart for the Number of SAA Taskforces

# #putting 0 to all NA to display them on the chart
# saa_partner_chart = summary_partner_saa
# saa_partner_chart['N° SAA Taskforces'] = saa_partner_chart['N° SAA Taskforces'].fillna(0)


# # Set the desired height for the chart
# fig_height = len(saa_partner_chart) * 0.2

# # Create the horizontal bar chart
# fig, ax = plt.subplots(figsize=(8, fig_height))
# bars = ax.barh(saa_partner_chart['Name'], saa_partner_chart['N° SAA Taskforces'], height=0.6)

# # Set the plot title and axis labels
# ax.set_title("Number of SAA Impact Systems per RtR Partner", fontsize=10)
# ax.set_xlabel('Number of SAA Impact Systems', fontsize=8)
# ax.set_ylabel('RtR Partner', fontsize=8)

# # Set the font size for x-axis labels
# plt.xticks(fontsize=6)
# # Set the font size for partner names
# plt.yticks(fontsize=6)

# # Add values on each bar
# for bar in bars:
# 	width = bar.get_width()
# 	label_y_pos = bar.get_y() + bar.get_height() / 2
# 	if not np.isnan(width):  # Check if width is not NaN
# 		if width == 0:
# 			ax.text(width + 0.1, label_y_pos, "Pending Confirmation", fontsize=6, va='center', ha='left')
# 		else:
# 			ax.text(width + 0.1, label_y_pos, s=f'{int(width)}', fontsize=6, va='center', ha='left')
		
# # Remove top and right spines
# ax.spines['right'].set_visible(False)
# ax.spines['top'].set_visible(False)

# # Set y-limits to remove excess blank space
# ax.set_ylim(-0.5, len(saa_partner_chart['Name']) - 0.5)

# # Invert the y-axis to display partners from top to bottom
# ax.invert_yaxis()

# # Use tight_layout to fit all elements nicely
# plt.tight_layout()
		
# # Show the plot in Streamlit
# st.pyplot(fig)

# st.caption('**TABLE #5:** NUMBER OF SAA IMPACT SYSTEMS PER RTR PARTNER')
# with st.expander("Explore Data: Number of SAA Taskforces per RtR Partner "):
# 	st.write("Number of SAA Taskforces per RtR Partner")
# 	st.write(summary_partner_saa)
# 	# csv = convert_df_gi(summary_partner_saa)
# 	# st.download_button(
# 	# 				label="Download data as CSV",
# 	# 				data=csv,
# 	# 				file_name='RtR_SAA_NumberPerPartner_'+datetofiles+'.csv',
# 	# 				mime='text/csv',
# 	# 			)


# st.subheader("4.3 - RtR PARTNER'S TYPE OF RELATIONSHIP WITH SAA iMPACT SYSTEMS")
# st.markdown("#### 4.3.1 - Relationship Types of RtR Partners with SAA Impact Systems")  # SHOWN IN 2023 REPORT


# Original one (NOT USED BECAUSE ABSOLUT NUMBERS)

# #Making DataSet with all SAA Taskforces Details for Chart
# new_df_gi_filtered = new_df_gi[new_df_gi['relation_yes'] == 'Yes']
# grouped_data_all = new_df_gi.groupby(['SAA Impact Systems', 'Relationship'], sort=False).size().unstack().reset_index()
# grouped_data_details = new_df_gi_filtered.groupby(['SAA Impact Systems', 'Relationship'], sort=False).size().unstack()
# total_partners = n_reporting_partners

# # Set the color palette
# sns.set_palette("RdPu")
# # Create the plot
# # fig_details, ax = plt.subplots(figsize=(10, len(grouped_data_details) / 2))
# fig_details, ax = plt.subplots(figsize=(6, len(grouped_data_details) / 2))
# # Create the horizontal bar chart
# grouped_data_details.plot(kind='barh', stacked=True, ax=ax)
# # Set the plot title and axis labels
# # ax.set_title("Number of RtR Partners related to each SAA Impact Systems [Details]")
# ax.set_xlabel('Number of RtR Partners')
# ax.set_ylabel('SAA Priority Systems')
# ax.set_xlim([0, 29])

# # Add annotation text at the bottom right of the figure
# # annotation_text = f"Out of {number_partner_saa} partners (Source: General Information Survey 2023)"
# # # Place the text at the bottom right of the figure
# # plt.gcf().subplots_adjust(bottom=0.2)  # Adjust the bottom of the figure to make room for the text.
# # fig_details.text(0.90, 0.00, annotation_text, ha='right', va='bottom', fontsize=12, color='black', transform=fig_details.transFigure)

# # Add individual and total values on the bars
# for p in ax.patches:
#     left, bottom, width, height = p.get_bbox().bounds
#     ax.text(left+width/2, bottom+height/2, int(width), ha='center', va='center')
# totals = grouped_data_details.sum(axis=1)
# for i, total in enumerate(totals):
#     ax.text(total + 0.2, i, f'Total: {int(total)} Partners', va='center')

# ax.spines['right'].set_visible(False)
# ax.spines['top'].set_visible(False)
# # ax.spines['left'].set_visible(False)
# # ax.spines['bottom'].set_visible(False)

# # Move the legend to the right of the plot area
# ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))

# # Show the plot in Streamlit
# st.write('"Number of RtR Partners related to each SAA Impact Systems [Details] (ORIGINAL)"')
# st.pyplot(fig_details)



# # Chart % OFICIAL
# # Making DataSet with all SAA Taskforces Details for Chart
# new_df_gi_filtered = new_df_gi[new_df_gi['relation_yes'] == 'Yes']
# grouped_data_all = new_df_gi.groupby(['SAA Impact Systems', 'Relationship'], sort=False).size().unstack().reset_index()
# grouped_data_details = new_df_gi_filtered.groupby(['SAA Impact Systems', 'Relationship'], sort=False).size().unstack()

# # Set the color palette
# sns.set_palette("RdPu")

# # Create the plot
# fig_details, ax = plt.subplots(figsize=(6, len(grouped_data_details) / 2))

# # Create the horizontal bar chart
# grouped_data_details.plot(kind='barh', stacked=True, ax=ax)

# # Set the plot title and axis labels
# ax.set_xlabel('Number of RtR Partners')
# ax.set_ylabel('SAA Priority Systems')
# ax.set_xlim([0, 29])

# # Add individual and total values on the bars
# for p in ax.patches:
#     left, bottom, width, height = p.get_bbox().bounds
#     ax.text(left + width/2, bottom + height/2, int(width), ha='center', va='center')

# # Calculate the total percentages using 29 as 100%
# totals = grouped_data_details.sum(axis=1)
# total_base = n_reporting_partners

# for i, total in enumerate(totals):
#     percentage = (total / total_base) * 100  # Calculate the percentage
#     ax.text(total + 0.2, i, f'{int(total)} ({percentage:.1f}%)', va='center')

# # Adjust legend and spines
# ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
# ax.spines['right'].set_visible(False)
# ax.spines['top'].set_visible(False)

# # Show the plot in Streamlit
# st.write('"Percentage of RtR Partners related to each SAA Impact Systems [Details]"')
# st.pyplot(fig_details)      ## <--- This one is oficial





# st.caption('**TABLE #6:** RELATIONSHIP WITH SAA IMPACT SYSTEMS')
# with st.expander("Explore Data: RtR Partner's Relationship with SAA impact Systems"):
# 	st.markdown("RtR Partner's Relationship with SAA impact Systems")
# 	st.write(grouped_data_all)
# 	#    convert_df_gi(grouped_data_all)
# 	# st.download_button(
# 	# 				label="Download data as CSV",
# 	# 				data=csv,
# 	# 				file_name='RtR_SAA_TypeOfRelation_'+datetofiles+'.csv',
# 	# 				mime='text/csv',
# 	# 			)
# 	st.caption("Note: 'Unsure' entries should be confirmed directly with RtR Partners.")


# st.markdown("#### 4.3.2 - Partner's Explanations of the relationship with SAA Impact Systems")

# saa_partners_toplot = new_df_gi_filtered

# saa_partners_toplot['concat'] = saa_partners_toplot['Partner Name'].astype(str) + saa_partners_toplot['SAA Impact Systems'].astype(str)

# # Calculate counts for each unique value
# value_counts = saa_partners_toplot['concat'].value_counts()
# # Calculate percentages
# total = len(saa_partners_toplot)
# value_percentages = (value_counts / total) * 100
# # If you want to add this percentage to your dataframe for each corresponding value
# saa_partners_toplot['% Partner'] = saa_partners_toplot['concat'].map(value_percentages)

# # Calculate occurrence of each SAA Taskforce
# taskforce_counts = saa_partners_toplot['SAA Impact Systems'].value_counts()
# # Calculate the percentage for each SAA Taskforce
# total = len(saa_partners_toplot)
# taskforce_percentage = (taskforce_counts / total) * 100
# # Map the percentage back to the original dataframe
# saa_partners_toplot['%SAA Taskforce'] = saa_partners_toplot['SAA Impact Systems'].map(taskforce_percentage)


# def add_line_breaks(text, char_limit=100):
# 	if text == "nan":  # if it's a NaN converted to string
# 		return ""
# 	words = text.split()
# 	lines = []
# 	current_line = ""
# 	for word in words:
# 		if len(current_line + word) > char_limit:
# 			lines.append(current_line.strip())
# 			current_line = ""
# 		current_line += word + " "
# 	lines.append(current_line.strip())
# 	return '<br>'.join(lines)

# # Convert all values in the column to string before applying the function
# saa_partners_toplot['Explanation (formatted)'] = saa_partners_toplot['Explanation (optional)'].astype(str).apply(add_line_breaks)


# fig_treemap = px.treemap(saa_partners_toplot, 
# 						path=[px.Constant("RTR PARTNER RELATIONS TO SAA IMPACT SYSTEMS"), 
# 							'SAA Impact Systems', 'Relationship', 'Partner Name', 'Explanation (formatted)'], 
# 						values='% Partner', 
# 						color='%SAA Taskforce', 
# 						color_continuous_scale='RdPu',
# 						custom_data=['Explanation (formatted)'])

# fig_treemap.update_traces(root_color="#FF37D5", 
# 						marker_line_color='grey', 
# 						hovertemplate='%{customdata[0]}<extra></extra>'
# 						)

# fig_treemap.data[0].hovertemplate = '%{value:.1f}%'

# fig_treemap.update_layout(title_text="RTR PARTNER RELATIONS TO SAA IMPACT SYSTEMS",
# 						margin=dict(t=50, l=25, r=25, b=25),
# 						width=1000)

# st.plotly_chart(fig_treemap)

# new_df_gi_filtered = new_df_gi_filtered[['Partner Name','SAA Impact Systems','Relationship','Explanation (optional)',]]


# st.caption("**TABLE #7** FULL DETAILS AND EXPLANATIONS OF RTR PARTNERS' RELATIONSHIP WITH SAA IMPACT SYSTEMS")
# with st.expander("Explore Data: Partner's Explanations of the relationship with SAA Impact Systems "):
# 	st.write(new_df_gi_filtered)
# 	# csv = convert_df_gi(new_df_gi_filtered)
# 	# st.download_button(
# 	# 					label="Download data as CSV",
# 	# 					data=csv,
# 	# 					file_name='RtR_SAA_TypeOfRelationHow_'+datetofiles+'.csv',
# 	# 					mime='text/csv',
# 	# 				)


# st.subheader("4.4 - RTR PARTNER RELATIONS TO SAA IMPACT SYSTEMS' OUTCOME TARGETS")

# outcome_to_explanation = {
# 	#Food and Agriculture Systems
# 	'q78':'q78',
# 	'q79': 'q82', 
# 	'q80': 'q83', 
# 	'q81': 'q84',
# 	#Water and Natural Systems
# 	'q87':'q87',
# 	'q88': 'q93',
# 	'q89': 'q94',
# 	'q90': 'q95',
# 	'q91': 'q96',
# 	'q92': 'q97',
# 	#Human Settlements
# 	'q100': 'q100',
# 	'q101': 'q106',
# 	'q102': 'q107',
# 	'q103': 'q108',
# 	'q104': 'q109',
# 	'q105': 'q110',
# 	#Coastal and Oceanic Systems 
# 	'q113': 	'q113',
# 	'q114': 	'q118',
# 	'q115': 	'q119',
# 	'q116': 	'q120',
# 	'q117': 	'q121',
# 	#Infrastructure Systems
# 	'q124': 	'q124',
# 	'q125': 	'q130',
# 	'q126': 	'q131',
# 	'q127': 	'q132',
# 	'q128': 	'q133',
# 	'q129': 	'q134', 
# 	#Cross-Cutting Enabler 'Planning'
# 	'q138': 	'q138',
# 	'q139': 	'q143',
# 	'q140': 	'q144',
# 	'q141': 	'q145',
# 	'q142': 	'q146',
# 	#Cross-Cutting Enabler 'Finance'
# 	'q149': 	'q149',
# 	'q150': 	'q153',
# 	'q151': 	'q154',
# 	'q152': 	'q155',
# }

# task_force_to_outcome = {
# 	'Food and Agriculture Systems': ['q78','q79', 'q80', 'q81'],
# 	'Water and Natural Systems': ['q87','q88', 'q89', 'q90', 'q91', 'q92'],
# 	'Human Settlements Systems': ['q100','q101', 'q102', 'q103', 'q104', 'q105'],
# 	'Coastal and Oceanic Systems': ['q113','q114', 'q115', 'q116', 'q117'],
# 	'Infrastructure Systems': ['q124','q125', 'q126', 'q127', 'q128', 'q129'],
# 	'Planning & Policy: Cross-cutting': ['q138','q139', 'q140', 'q141', 'q142'],
# 	'Finance: Cross-cutting': ['q149','q150', 'q151', 'q152']
# }

# final_df_gi_outcomes = pd.DataFrame(columns=['RtR Partner Name', 'SAA Impact Systems', 'Outcome Target', 'Explanation'])

# for task_force, outcomes in task_force_to_outcome.items():
# 	for outcome in outcomes:
# 		explanation = outcome_to_explanation[outcome]
# 		temp_df_gi = df_gi[['InitName', outcome, explanation]]
# 		temp_df_gi.columns = ['RtR Partner Name', 'Outcome Target', 'Explanation']
# 		temp_df_gi['SAA Impact Systems'] = task_force
# 		final_df_gi_outcomes = pd.concat([final_df_gi_outcomes, temp_df_gi])

# final_df_gi_outcomes = final_df_gi_outcomes.dropna(how='all')
# final_df_gi_outcomes = final_df_gi_outcomes[~pd.isna(final_df_gi_outcomes['Outcome Target'])]


# # Calculate occurrence of each RtR Partner Name
# partner_counts = final_df_gi_outcomes['RtR Partner Name'].value_counts()
# # Calculate the percentage for each RtR Partner Name
# total = len(final_df_gi_outcomes)
# partner_percentage = (partner_counts / total) * 100
# # Map the percentage back to the original dataframe
# final_df_gi_outcomes['%RtR Partner Name'] = final_df_gi_outcomes['RtR Partner Name'].map(partner_percentage)
	
	
# # Calculate occurrence of each SAA Taskforce
# taskforce_counts = final_df_gi_outcomes['SAA Impact Systems'].value_counts()
# # Calculate the percentage for each SAA Taskforce
# total = len(final_df_gi_outcomes)
# taskforce_percentage = (taskforce_counts / total) * 100
# # Map the percentage back to the original dataframe
# final_df_gi_outcomes['%SAA Taskforce'] = final_df_gi_outcomes['SAA Impact Systems'].map(taskforce_percentage)


# # Calculate occurrence of each Outcome Target per SAA Taskforce
# outcome_counts = final_df_gi_outcomes.groupby(['SAA Impact Systems', 'Outcome Target']).size().reset_index(name='counts')
# # Calculate total occurrences per SAA Taskforce
# total_counts = final_df_gi_outcomes.groupby(['SAA Impact Systems']).size().reset_index(name='total')
# # Merge the two dataframes to compute the percentage
# merged = outcome_counts.merge(total_counts, on='SAA Impact Systems')
# merged['%OutcomeTarget_by_taskforce'] = (merged['counts'] / merged['total']) * 100
# # Merge back with the original dataframe
# final_df_gi_outcomes = final_df_gi_outcomes.merge(merged[['SAA Impact Systems', 'Outcome Target', '%OutcomeTarget_by_taskforce']], on=['SAA Impact Systems', 'Outcome Target'], how='left')

# # Calculate occurrence of each SAA Taskforce
# taskforce_counts = final_df_gi_outcomes['Outcome Target'].value_counts()
# # Calculate the percentage for each SAA Taskforce
# total = len(final_df_gi_outcomes)
# taskforce_percentage = (taskforce_counts / total) * 100
# # Map the percentage back to the original dataframe
# final_df_gi_outcomes['%OutcomeTarget_all'] = final_df_gi_outcomes['Outcome Target'].map(taskforce_percentage)

# # st.write(final_df_gi_outcomes)

# #MAKING THE TREE MAP
# final_df_gi_to_treemap = final_df_gi_outcomes[['SAA Impact Systems', 'Outcome Target','%SAA Taskforce', '%OutcomeTarget_all']]
# final_df_gi_to_treemap = final_df_gi_to_treemap.drop_duplicates()


# # note = "NOTE: (1) No Health System outcomes affect SAA percentages. (2) Multiple RtR ties shift Task Force %s."
# note = "NOTE: (1) Health Systems aren't in the SAA percentages due to no outcome targets. (2) RtR Partners' multiple Task Force ties can shift the percentages."

# fig_treemap = px.treemap(final_df_gi_to_treemap, 
# 						path=[px.Constant("RTR PARTNER RELATIONS TO SAA IMPACT SYSTEMS' OUTCOME TARGETS"), 'SAA Impact Systems', 'Outcome Target',], 
# 						values='%OutcomeTarget_all', 
# 						color='%SAA Taskforce', 
# 						color_continuous_scale='RdPu')

# fig_treemap.update_traces(root_color="#FF37D5")
# fig_treemap.update_layout(title_text="RTR PARTNER RELATIONS TO SAA IMPACT SYSTEMS' OUTCOME TARGETS",
# 						margin=dict(t=50, l=25, r=25, b=25),
# 						width=1000)
# fig_treemap.add_annotation(x=1, y=0,
# 			text=note,showarrow=False,
# 			yshift=-20)

# # It looks like you want to show the values as percentages with one decimal place.
# fig_treemap.data[0].hovertemplate = '%{value:.1f}%'
# fig_treemap.update_traces(marker_line_color='grey')
# st.plotly_chart(fig_treemap)




# final_df_gi_outcomes = final_df_gi_outcomes[['RtR Partner Name','SAA Impact Systems','Outcome Target','Explanation']]

# # st.write(final_df_gi_outcomes.columns)



# st.caption('**TABLE #8** SAA IMPACT SYSTEMS OUTCOMES TARGETS')
# with st.expander("Explore Data: RtR Partner relation to SAA impact Systems' Outcomes Targets "):
# 	st.write("RtR Partner relation to SAA impact Systems' Outcomes Targets")
# 	st.write(final_df_gi_outcomes)
# 	# csv = convert_df_gi(final_df_gi_outcomes)
# 	# st.download_button(
# 	# 				label="Download data as CSV",
# 	# 				data=csv,
# 	# 				file_name='RtR_GeneralInfoSurvey_SAA_'+datetofiles+'.csv',
# 	# 				mime='text/csv',
# 	# 			)
	

# st.subheader("4.5 - RtR PARTNER & SAA RELATIONSHIPS BY IMPACT SYSTEMS")

# st.markdown("#### 4.5.1 - FOOD AND AGRICULTURE SYSTEMS")
# st.caption("**Q24. How do the Food and Agriculture Systems relate to the work of your organization?**")
# foodandagriculture_saa = df_gi['q76'].count()
# foodandagriculture_saa_direct_related = len(df_gi[df_gi['q76'] == 'Directly related'])
# foodandagriculture_saa_indirect_related = len(df_gi[df_gi['q76'] == 'Indirectly related'])
# foodandagriculture_saa_not_related = len(df_gi[df_gi['q76'] == 'Not Related'])
# foodandagriculture_saa_unsure = len(df_gi[df_gi['q76'] == 'Unsure'])
# st.caption("Q24. Number of partners 'Directly related': "+str(foodandagriculture_saa_direct_related))
# st.caption("Q24. Number of partners 'Indirectly related': "+str(foodandagriculture_saa_indirect_related))
# st.caption("Q24. Number of partners 'Not Related': "+str(foodandagriculture_saa_not_related))
# st.caption("Q24. Number of partners 'Unsure': "+str(foodandagriculture_saa_unsure))
# st.caption("**Q25. If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Food and Agriculture Systems? (OPTIONAL)**")
# # foodandagriculture_saa_explanation = df_gi['q77'].count()
# st.caption("Q25. Number of partners responding the question: "+str(foodandagriculture_saa_explanation))

# saa_taskforce = 'Food and Agriculture Systems'

# with st.expander("Explore More in "+saa_taskforce ):
# 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# 	st.write('Number of RtR Partners related to '+saa_taskforce)
# 	st.write(grouped_data_all_filtered)
	
# 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# 	st.write('RtR Partner relation to '+saa_taskforce)
# 	st.write(new_df_gi_filtered_2)
	
# 	st.write('RtR Partner relation to '+saa_taskforce+ "' Outcomes Targets" )
# 	final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# 	st.write(final_df_gi_outcome_target)
	



# # st.markdown("#### 4.5.2 - WATER AND NATURE SYSTEMS")
# # # st.caption("**Q28. How do the Water and Nature Systems relate to the work of your organization?**")
# # # water_natsyst_saa = df_gi['q85'].count()
# # # water_natsyst_saa_direct_related = len(df_gi[df_gi['q85'] == 'Directly related'])
# # # water_natsyst_saa_indirect_related = len(df_gi[df_gi['q85'] == 'Indirectly related'])
# # # water_natsyst_saa_not_related = len(df_gi[df_gi['q85'] == 'Not Related'])
# # # water_natsyst_saa_unsure = len(df_gi[df_gi['q85'] == 'Unsure'])
# # # st.caption("Q28. Number of partners 'Directly related': "+str(water_natsyst_saa_direct_related))
# # # st.caption("Q28. Number of partners 'Indirectly related': "+str(water_natsyst_saa_indirect_related))
# # # st.caption("Q28. Number of partners 'Not Related': "+str(water_natsyst_saa_not_related))
# # # st.caption("Q28. Number of partners 'Unsure': "+str(water_natsyst_saa_unsure))
# # # st.caption("**Q29 If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Water and Nature Systems? (OPTIONAL)**")
# # # water_natsyst_saa_explanation = df_gi['q86'].count()
# # # st.caption("Q29. Number of partners responding the question: "+str(water_natsyst_saa_explanation))


# # saa_taskforce = 'Water and Natural Systems'

# # with st.expander("Explore More in "+saa_taskforce ):
# # 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('Number of RtR Partners related to '+saa_taskforce)
# # 	st.write(grouped_data_all_filtered)
	
# # 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('RtR Partner relation to '+saa_taskforce)
# # 	st.write(new_df_gi_filtered_2)
	
# # 	st.write('RtR Partner relation to '+saa_taskforce+ "' Outcomes Targets" )
# # 	final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# # 	st.write(final_df_gi_outcome_target)
	


# # st.markdown("#### 4.5.3 - HUMAN SETTLEMENT SYSTEMS")
# # # st.caption("**Q32. How do the Human Settlement Systems relate to the work of your organization?**")
# # # humansettlement_saa = df_gi['q98'].count()
# # # humansettlement_saa_direct_related = len(df_gi[df_gi['q98'] == 'Directly related'])
# # # humansettlement_saa_indirect_related = len(df_gi[df_gi['q98'] == 'Indirectly related'])
# # # humansettlement_saa_not_related = len(df_gi[df_gi['q98'] == 'Not Related'])
# # # humansettlement_saa_unsure = len(df_gi[df_gi['q98'] == 'Unsure'])
# # # st.caption("Q32. Number of partners 'Directly related': "+str(humansettlement_saa_direct_related))
# # # st.caption("Q32. Number of partners 'Indirectly related': "+str(humansettlement_saa_indirect_related))
# # # st.caption("Q32. Number of partners 'Not Related': "+str(humansettlement_saa_not_related))
# # # st.caption("Q32. Number of partners 'Unsure': "+str(humansettlement_saa_unsure))
# # # st.caption("**Q33. If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Human Settlement Systems? (OPTIONAL)**")
# # # humansettlement_saa_explanation = df_gi['q99'].count()
# # # st.caption("Q33. Number of partners responding the question: "+str(humansettlement_saa_explanation))

# # saa_taskforce = 'Human Settlements Systems'

# # with st.expander("Explore More in "+saa_taskforce ):
# # 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('Number of RtR Partners related to '+saa_taskforce)
# # 	st.write(grouped_data_all_filtered)
	
# # 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('RtR Partner relation to '+saa_taskforce)
# # 	st.write(new_df_gi_filtered_2)
	
# # 	st.write('RtR Partner relation to '+saa_taskforce+ "' Outcomes Targets" )
# # 	final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# # 	st.write(final_df_gi_outcome_target)


# # st.markdown("#### 4.5.4 - COASTAL AND OCEANIC SYSTEMS")
# # # st.caption("**Q36. How do the Coastal and Oceanic Systems relate to the work of your organization?**")
# # # coastal_oceanic_saa = df_gi['q111'].count()
# # # coastal_oceanic_saa_direct_related = len(df_gi[df_gi['q111'] == 'Directly related'])
# # # coastal_oceanic_saa_indirect_related = len(df_gi[df_gi['q111'] == 'Indirectly related'])
# # # coastal_oceanic_saa_not_related = len(df_gi[df_gi['q111'] == 'Not Related'])
# # # coastal_oceanic_saa_unsure = len(df_gi[df_gi['q111'] == 'Unsure'])
# # # st.caption("Q36. Number of partners 'Directly related': "+str(coastal_oceanic_saa_direct_related))
# # # st.caption("Q36. Number of partners 'Indirectly related': "+str(coastal_oceanic_saa_indirect_related))
# # # st.caption("Q36. Number of partners 'Not Related': "+str(coastal_oceanic_saa_not_related))
# # # st.caption("Q36. Number of partners 'Unsure': "+str(coastal_oceanic_saa_unsure))
# # # st.caption("**Q37. If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Coastal and Oceanic Systems? (OPTIONAL)**")
# # # coastal_oceanic_saa_explanation = df_gi['q112'].count()
# # # st.caption("Q37. Number of partners responding the question: "+str(coastal_oceanic_saa_explanation))

# # saa_taskforce = 'Coastal and Oceanic Systems'

# # with st.expander("Explore More in "+saa_taskforce ):
# # 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('Number of RtR Partners related to '+saa_taskforce)
# # 	st.write(grouped_data_all_filtered)
	
# # 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('RtR Partner relation to '+saa_taskforce)
# # 	st.write(new_df_gi_filtered_2)
	
# # 	st.write('RtR Partner relation to '+saa_taskforce+ "' Outcomes Targets" )
# # 	final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# # 	st.write(final_df_gi_outcome_target)



# # st.markdown("#### 4.5.5 - INFRASTRUCTURE SYSTEMS")
# # # st.caption("**Q40. How do the Infrastructure Systems relate to the work of your organization?**")
# # # infrastructure_saa = df_gi['q122'].count()
# # # infrastructure_saa_direct_related = len(df_gi[df_gi['q122'] == 'Directly related'])
# # # infrastructure_saa_indirect_related = len(df_gi[df_gi['q122'] == 'Indirectly related'])
# # # infrastructure_saa_not_related = len(df_gi[df_gi['q122'] == 'Not Related'])
# # # infrastructure_saa_unsure = len(df_gi[df_gi['q122'] == 'Unsure'])
# # # st.caption("Q40. Number of partners 'Directly related': "+str(infrastructure_saa_direct_related))
# # # st.caption("Q40. Number of partners 'Indirectly related': "+str(infrastructure_saa_indirect_related))
# # # st.caption("Q40. Number of partners 'Not Related': "+str(infrastructure_saa_not_related))
# # # st.caption("Q40. Number of partners 'Unsure': "+str(infrastructure_saa_unsure))
# # # st.caption("**Q41. If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Infrastructure Systems? (OPTIONAL)**")
# # # infrastructure_saa_explanation = df_gi['q123'].count()
# # # st.caption("Q41. Number of partners responding the question: "+str(infrastructure_saa_explanation))


# # saa_taskforce = 'Infrastructure Systems'

# # with st.expander("Explore More in "+saa_taskforce ):
# # 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('Number of RtR Partners related to '+saa_taskforce)
# # 	st.write(grouped_data_all_filtered)
	
# # 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('RtR Partner relation to '+saa_taskforce)
# # 	st.write(new_df_gi_filtered_2)
	
# # 	st.write('RtR Partner relation to '+saa_taskforce+ "' Outcomes Targets" )
# # 	final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# # 	st.write(final_df_gi_outcome_target)
	


# # st.markdown("#### 4.5.6 - HEATH SYSTEMS")
# # # st.caption("**Q44. Does the work of your organization include any health-related outcomes or activities**")
# # # # st.write(df_gi['q135'])
# # # health_saa = df_gi['q135'].count()
# # # health_saa_yes= len(df_gi[df_gi['q135'] == 'Directly related'])
# # # health_saa_no= len(df_gi[df_gi['q135'] == 'Not Related'])
# # # health_saa_not_sure = len(df_gi[df_gi['q135'] == 'Unsure'])
# # # st.caption("Q40. Number of partners 'Related': "+str(health_saa_yes))
# # # st.caption("Q40. Number of partners 'No': "+str(health_saa_no))
# # # st.caption("Q40. Number of partners 'Not Sure': "+str(health_saa_not_sure))

# # # st.caption("**Q41. If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Health Systems? (OPTIONAL)**")
# # # health_saa_explanation = df_gi['q136'].count()
# # # st.caption("Q41. Number of partners responding the question: "+str(health_saa_explanation))   

# # saa_taskforce = 'Health Systems'

# # with st.expander("Explore More in "+saa_taskforce ):
# # 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('Number of RtR Partners related to '+saa_taskforce)
# # 	st.write(grouped_data_all_filtered)
	
# # 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('RtR Partner relation to '+saa_taskforce)
# # 	st.write(new_df_gi_filtered_2)
	
# # 	st.write(saa_taskforce+ " Outcomes Targets are not yet available" )
	
# # 	# final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# # 	# st.write(final_df_gi_outcome_target)



# # st.markdown("#### 4.5.7 - CROSS CUTTING ENABLER: PLANNING")
# # # st.caption("**Q46. How do the Cross-Cutting Enabler 'Planning' relate to the work of your organization?**")
# # # planning_saa = df_gi['q137'].count()
# # # planning_saa_direct_related = len(df_gi[df_gi['q137'] == 'Directly related'])
# # # planning_saa_indirect_related = len(df_gi[df_gi['q137'] == 'Indirectly related'])
# # # planning_saa_not_related = len(df_gi[df_gi['q137'] == 'Not Related'])
# # # planning_saa_unsure = len(df_gi[df_gi['q137'] == 'Unsure'])
# # # st.caption("Q46. Number of partners 'Directly related': "+str(planning_saa_direct_related))
# # # st.caption("Q46. Number of partners 'Indirectly related': "+str(planning_saa_indirect_related))
# # # st.caption("Q46. Number of partners 'Not Related': "+str(planning_saa_not_related))
# # # st.caption("Q46. Number of partners 'Unsure': "+str(planning_saa_unsure))
# # # st.caption("**Q47. If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Cross-Cutting Enabler 'Planning'? (OPTIONAL)**")
# # # planning_saa_explanation = df_gi['q138'].count()
# # # st.caption("Q47. Number of partners responding the question: "+str(planning_saa_explanation))

# # #  'Planning & Policy: Cross-cutting'

# # saa_taskforce = 'Planning & Policy: Cross-cutting'

# # with st.expander("Explore More in "+saa_taskforce ):
# # 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('Number of RtR Partners related to '+saa_taskforce)
# # 	st.write(grouped_data_all_filtered)
	
# # 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('RtR Partner relation to '+saa_taskforce)
# # 	st.write(new_df_gi_filtered_2)
	
# # 	st.write('RtR Partner relation to '+saa_taskforce+ "' Outcomes Targets" )
# # 	final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# # 	st.write(final_df_gi_outcome_target)
	

# # st.markdown("#### 4.5.8 - CROSS CUTTING ENABLER: FINANCE")
# # # st.caption("**Q50. How do the Cross-Cutting Enabler 'Finance' relate to the work of your organization?**")
# # # finance_saa = df_gi['q147'].count()
# # # finance_saa_direct_related = len(df_gi[df_gi['q147'] == 'Directly related'])
# # # finance_saa_indirect_related = len(df_gi[df_gi['q147'] == 'Indirectly related'])
# # # finance_saa_not_related = len(df_gi[df_gi['q147'] == 'Not Related'])
# # # finance_saa_unsure = len(df_gi[df_gi['q147'] == 'Unsure'])
# # # st.caption("Q50. Number of partners 'Directly related': "+str(finance_saa_direct_related))
# # # st.caption("Q50. Number of partners 'Indirectly related': "+str(finance_saa_indirect_related))
# # # st.caption("Q50. Number of partners 'Not Related': "+str(finance_saa_not_related))
# # # st.caption("Q50. Number of partners 'Unsure': "+str(finance_saa_unsure))
# # # st.caption("**Q51. If applicable, could you provide further detail on the outcomes or activities your organization has in relation to Cross-Cutting Enabler 'Finance'? (OPTIONAL)**")
# # # finance_saa_explanation = df_gi['q148'].count()
# # # st.caption("Q51. Number of partners responding the question: "+str(finance_saa_explanation))


# # # 'Finance: Cross-cutting'

# # saa_taskforce = 'Finance: Cross-cutting'

# # with st.expander("Explore More in "+saa_taskforce ):
# # 	grouped_data_all_filtered = grouped_data_all[grouped_data_all['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('Number of RtR Partners related to '+saa_taskforce)
# # 	st.write(grouped_data_all_filtered)
	
# # 	new_df_gi_filtered_2 = new_df_gi_filtered[new_df_gi_filtered['SAA Impact Systems'] == saa_taskforce]
# # 	st.write('RtR Partner relation to '+saa_taskforce)
# # 	st.write(new_df_gi_filtered_2)
	
# # 	st.write('RtR Partner relation to '+saa_taskforce+ "' Outcomes Targets" )
# # 	final_df_gi_outcome_target = final_df_gi_outcomes[final_df_gi_outcomes['SAA Impact Systems'] == saa_taskforce]
# # 	st.write(final_df_gi_outcome_target)


# # ##_____________________________________________________________________________________________________________
# # ##_____________________________________________________________________________________________________________
# # ##_____________________________________________________________________________________________________________
# # # # ______________________________________________
# # # # ______________________________________________
# # # # END BLOCK SAA_______________________________
# # # # ______________________________________________
# # # # ______________________________________________
# # ##_____________________________________________________________________________________________________________
# # ##_____________________________________________________________________________________________________________
# # ##_____________________________________________________________________________________________________________
















# ##_______________ ______________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##__BLOCK ACTION CLUSTERS GENERAL INFORMATION _________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________

# st.header("SECTION 5. IDENTIFICATION OF RESILIENCE ACTION CLUSTERS")
# st.markdown("In this section, RtR Partners are asked to identify and elaborate on the resilience actions taken by their initiative, which are categorized into Action Clusters.")
# # st.caption("DATA: "+type_info)

# st.subheader("5.1 - RESILIENCE ACTION CLUSTERS")
# st.caption("**Q54. Please review the descriptions of each Action Cluster. Choose the ones that correspond closely to your organization work related to RtR, considering that a separate plan may be needed for each cluster in future reports. In cases where a particular action of your organization could be associated with multiple clusters, please assign it to the one that represents it best. Don't hesitate to select as many Action Clusters as required to adequately represent the wide range of activities conducted by your organization.**")

# #Creating the FULL Action CLuster Data Set
# actioncluster_list = ['q157','q158','q159','q160','q161','q162','q163','q164','q165',
# 'q166','q167','q168','q169','q170','q171','q172','q173','q174','q175',
# 'q176','q177','q178','q179','q180','q181','q182','q183','q184','q185',]

# actioncluster_how_list = ['q186','q187','q188','q189','q190','q191','q192','q193','q194','q195','q196','q197',
# 							'q198','q199','q200','q201','q202','q203','q204','q205','q206','q207','q208','q209','q210',
# 							'q211','q212','q213','q214',]

# # cleaning dataset of how
# df_gi[actioncluster_how_list] = df_gi[actioncluster_how_list].replace('x', np.nan)

# # st.caption("dataframehow")
# # df_gi_how = df_gi[actioncluster_how_list]
# # st.write(df_gi_how)

# action_cluster_to_explanation = {
# 	'q157' : 'q186','q158' : 'q187','q159' : 'q188','q160' : 'q189',
# 	'q161' : 'q190','q162' : 'q191','q163' : 'q192','q164' : 'q193',
# 	'q165' : 'q194','q166' : 'q195','q167' : 'q196','q168' : 'q197',
# 	'q169' : 'q198','q170' : 'q199','q171' : 'q200','q172' : 'q201',
# 	'q173' : 'q202','q174' : 'q203','q175' : 'q204','q176' : 'q205',
# 	'q177' : 'q206','q178' : 'q207','q179' : 'q208','q180' : 'q209',
# 	'q181' : 'q210','q182' : 'q211','q183' : 'q212','q184' : 'q213','q185' : 'q214',
# }
# partner_names = ['InitName']

# # Combine the lists
# selected_columns = partner_names + actioncluster_list + actioncluster_how_list

# # Select the specific columns from the DataFrame
# df_gi_partners_actioncluster = df_gi[selected_columns]
# # st.caption("all variables needeed")
# # st.write(df_gi_partners_actioncluster)

# # Melt the actioncluster_list columns into individual rows
# df_gi_partners_actioncluster = df_gi.melt(id_vars=partner_names, value_vars=actioncluster_list, value_name='Action Clusters')

# # Rename the 'InitName' column to 'Partner'
# df_gi_partners_actioncluster.rename(columns={'InitName': 'Partner'}, inplace=True)

# # st.caption("Melt Action Cluster")
# # st.write(df_gi_partners_actioncluster.count())

# # Remove rows where either "Partner" or "Action Clusters" is NaN or empty
# df_gi_partners_actioncluster.dropna(subset=['Partner', 'Action Clusters'], inplace=True)

# # st.caption('Melt where either "Partner" or "Action Clusters" is NaN or empty')
# # st.write(df_gi_partners_actioncluster.count())

# # Remove rows where either "Partner" or "Action Clusters" is an empty string
# df_gi_partners_actioncluster = df_gi_partners_actioncluster[df_gi_partners_actioncluster['Partner'].str.strip() != '']
# df_gi_partners_actioncluster = df_gi_partners_actioncluster[df_gi_partners_actioncluster['Action Clusters'].str.strip() != '']

# # st.caption('Remove rows where either "Partner" or "Action Clusters" is an empty string')
# # st.write(df_gi_partners_actioncluster.count())
# # st.write(df_gi_partners_actioncluster)

# #Making Status Info Data Set
# df_gi_status_unique = df_gi[['Org_Type','InitName',]]
# df_gi_status_unique = df_gi_status_unique.rename(columns={'InitName': 'Partner', 
# 													'Org_Type':'Relation to RtR',})
# #Merging Data with Status Info
# df_gi_partners_actioncluster = pd.merge(df_gi_partners_actioncluster, df_gi_status_unique, on='Partner', how='outer')
# #Merging Data with Action_Types_Info
# df_gi_partners_actioncluster = pd.merge(df_gi_partners_actioncluster, ac_atypes, on='variable', how='outer')

# #Concat to create a key
# df_gi_partners_actioncluster["key"] = df_gi_partners_actioncluster['Partner'].astype(str) + df_gi_partners_actioncluster['variable_how'].astype(str)
# # st.caption('DataFrame x Action Cluster+otra info')
# # st.write(df_gi_partners_actioncluster)

# #HOW
# # Melt the actioncluster_list columns into individual rows
# df_gi_partners_actioncluster_how = df_gi.melt(id_vars=partner_names, value_vars=actioncluster_how_list, value_name='HOW Action Clusters')

# # Rename the 'InitName' column to 'Partner'
# df_gi_partners_actioncluster_how.rename(columns={'InitName': 'Partner'}, inplace=True)
# df_gi_partners_actioncluster_how.rename(columns={'variable': 'variable_how'}, inplace=True)

# #Concat to create a key
# df_gi_partners_actioncluster_how["key"] = df_gi_partners_actioncluster_how['Partner'].astype(str) + df_gi_partners_actioncluster_how['variable_how'].astype(str)

# # st.caption("Melt Action Cluster How")
# # st.write(df_gi_partners_actioncluster_how)
# # st.write(df_gi_partners_actioncluster_how.count())

# # Remove rows where either "Partner" or "Action Clusters" is NaN or empty
# df_gi_partners_actioncluster_how.dropna(subset=['Partner', 'HOW Action Clusters'], inplace=True)

# # st.caption('HOW Melt where either "Partner" or "Action Clusters" is NaN or empty')
# # st.write(df_gi_partners_actioncluster_how.count())

# # Remove rows where either "Partner" or "Action Clusters" is an empty string
# df_gi_partners_actioncluster_how = df_gi_partners_actioncluster_how[df_gi_partners_actioncluster_how['Partner'].str.strip() != '']
# df_gi_partners_actioncluster_how = df_gi_partners_actioncluster_how[df_gi_partners_actioncluster_how['HOW Action Clusters'].str.strip() != '']

# # st.caption('HOW Remove rows where either "Partner" or "Action Clusters" is an empty string')
# # st.write(df_gi_partners_actioncluster_how.count())

# # st.caption('HOW DataFrame x Action Cluster HOW')
# # st.write(df_gi_partners_actioncluster_how)

# # st.caption('HOW DataFrame x Action Cluster HOW_edited')
# df_gi_partners_actioncluster_how = df_gi_partners_actioncluster_how[['key','HOW Action Clusters']]
# # st.write(df_gi_partners_actioncluster_how)

# #Combining datas sets
# #Merging Data with Status Info
# df_gi_partners_actioncluster_total = pd.merge(df_gi_partners_actioncluster, df_gi_partners_actioncluster_how , on='key', how='outer')
# # st.write(df_gi_partners_actioncluster_total.columns)
# new_order = ['Relation to RtR','Partner','variable','variable_how','key','Action Cluster_code','Action Clusters','HOW Action Clusters','Action Type_code','Actions Type (Marrakech)']
# df_gi_partners_actioncluster_total = df_gi_partners_actioncluster_total[new_order]

# # st.subheader('Total')
# # st.write(df_gi_partners_actioncluster_total)


# #Count of Partners with Action Cluster Infor
# actioncluster_all = df_gi[actioncluster_list].notna().any(axis=1).sum()
# st.caption("Total RtR Partners with Identified Action Clusters So Far:  "+str(actioncluster_all))
# ## Making the Chart#1: Number of RtR Partners per Action Cluster

# 	# Define the dictionary mapping for action clusters
# action_cluster_dict = {
# 	'q157': 'Monitoring and mapping of hazards and vulnerabilities',
# 	'q158': 'Climate services and modelling',
# 	'q159': 'Generation and processing of data',
# 	'q160': 'Early warning systems',
# 	'q161': 'Emergency preparedness',
# 	'q162': 'Environmental governance and resource management systems',
# 	'q163': 'Environmental laws and regulatory framework',
# 	'q164': 'Building regulations and standards for climate resilience',
# 	'q165': 'Protected areas and property rights definitions',
# 	'q166': 'Institutional-led climate adaptation planning processes',
# 	'q167': 'Measures to improve air quality and reduce pollution',
# 	'q168': 'Green infrastructure and other engineered nature-based solutions',
# 	'q169': 'Conservation and restoration of terrestrial and aquatic ecosystems',
# 	'q170': 'Critical Infrastructure and Protective Systems',
# 	'q171': 'Coastal Infrastructure Protection',
# 	'q172': 'Energy efficiency and renewable energy technologies',
# 	'q173': 'Water security and quality',
# 	'q174': 'Health services',
# 	'q175': 'Food safety and sustainable services',
# 	'q176': 'Agricultural, livestock, forestry and aquacultural practices actions',
# 	'q177': 'Social protection actions',
# 	'q178': 'Community-based inclusive and participatory risk reduction',
# 	'q179': 'Climate hazard communication, information, and technology awareness',
# 	'q180': 'Knowledge building for resilience',
# 	'q181': 'Research and collaboration actions',
# 	'q182': 'Livelihood diversification and social economy',
# 	'q183': 'Climate insurance and risk transfer',
# 	'q184': 'Financial and investment tools in case of climate disasters',
# 	'q185': 'Economic incentives'
# }
# # Calculate the count of identified action clusters
# actioncluster_all = df_gi[actioncluster_list].notna().any(axis=1).sum()

# # Create a list to store the counts of each action cluster
# actioncluster_counts = [df_gi[column].notna().sum() for column in actioncluster_list]

# # Create a list to store the action cluster names
# actioncluster_names = [action_cluster_dict[column].strip() for column in actioncluster_list]

# # Store the counts and the names in a DataFrame
# df_gi_actionclusters = pd.DataFrame({
# 	'names': actioncluster_names,
# 	'counts': actioncluster_counts
# })

# # Sort the DataFrame in descending order of 'counts'
# df_gi_actionclusters = df_gi_actionclusters.sort_values(by='counts', ascending=False)



## CHART ACTION CLUSTER

# # Chart:
# # Set the desired height for the chart
# fig_height = len(actioncluster_names) * 0.17

# # Create the horizontal bar chart
# fig, ax = plt.subplots(figsize=(8, fig_height))
# bars = ax.barh(df_gi_actionclusters['names'], df_gi_actionclusters['counts'], height=0.6)

# # Set the plot title and axis labels
# ax.set_title("Number of RtR Partners per Action Cluster", fontsize=10)
# ax.set_xlabel('Number of Partners', fontsize=8)
# ax.set_ylabel('Action Cluster', fontsize=8)
# # Invert the y-axis to display action clusters from top to bottom
# ax.invert_yaxis()
# # Set the x-axis tick range to advance by 1 point
# ax.set_xticks(range(0, max(df_gi_actionclusters['counts']) + 2, 1))

# # Set the font size for x-axis labels
# plt.xticks(fontsize=6)
# # Set the font size for action cluster names
# plt.yticks(fontsize=8)
# # Adjust the layout to prevent overlapping labels
# plt.tight_layout()
# # Add annotation text at the bottom right of the figure
# annotation_text = f"Out of {actioncluster_all} partners (Source: General Information Survey 2023)"
# # Place the text at the bottom right of the figure
# fig.text(0.95, 0.00, annotation_text, ha='right', va='bottom', fontsize=8, color='black', transform=fig.transFigure)


# # Add values on each bar
# for bar in bars:
# 	width = bar.get_width()
# 	label_y_pos = bar.get_y() + bar.get_height() / 2
# 	ax.text(width + 0.1, label_y_pos, s=f'{width}', fontsize=6, va='center', ha='left')

# # Show the plot in Streamlit
# st.markdown("#### 5.1.1 - Number of RtR Partners per Action Cluster")
# st.pyplot(fig)


# ###Chart #2: Number of Action Clusters per RtR Partner
# # Transpose the DataFrame and count non-null actions for each partner
# df_gi_transposed = df_gi.set_index('InitName')[actioncluster_list].T
# partner_action_counts = df_gi_transposed.notna().sum()

# # Convert the Series to a DataFrame and reset the index
# df_gi_partners = partner_action_counts.to_frame(name='counts').reset_index()
# # Sort the DataFrame in descending order of 'counts'
# df_gi_partners = df_gi_partners.sort_values(by='counts', ascending=False)
# # Set the desired height for the chart
# fig_height = len(df_gi_partners) * 0.2
# # Create the horizontal bar chart
# fig, ax = plt.subplots(figsize=(8, fig_height))
# bars = ax.barh(df_gi_partners['InitName'], df_gi_partners['counts'], height=0.6)
# # Set the plot title and axis labels
# ax.set_title("Number of Action Clusters per RtR Partner", fontsize=10)
# ax.set_xlabel('Number of Action Clusters', fontsize=8)
# ax.set_ylabel('RtR Partner', fontsize=8)
# # Invert the y-axis to display partners from top to bottom
# ax.invert_yaxis()
# # Set the x-axis tick range to advance by 1 point
# ax.set_xticks(range(0, max(df_gi_partners['counts']) + 2, 1))
# # Set the font size for x-axis labels
# plt.xticks(fontsize=6)
# # Set the font size for partner names
# plt.yticks(fontsize=8)
# # Adjust the layout to prevent overlapping labels
# plt.tight_layout()

# # # Add annotation text at the bottom right of the figure
# # annotation_text = f"Out of {actioncluster_all} partners (Source: General Information Survey 2023)"
# # # Place the text at the bottom right of the figure
# # fig.text(0.95, 0.00, annotation_text, ha='right', va='bottom', fontsize=8, color='black', transform=fig.transFigure)


# # Add values on each bar
# for bar in bars:
# 	width = bar.get_width()
# 	label_y_pos = bar.get_y() + bar.get_height() / 2
# 	ax.text(width + 0.1, label_y_pos, s=f'{width}', fontsize=6, va='center', ha='left')

# # Show the plot in Streamlit
# st.markdown("#### 5.1.2 - Number of Action Clusters per RtR Partner")
# st.pyplot(fig)


# ###Chart #3: Number of Action Types per RtR Partner
	
# st.subheader("5.2 - ACTION TYPES (Marrakech)")
# df_gi_to_plot_action_types = df_gi_partners_actioncluster_total[['Partner','Actions Type (Marrakech)']].drop_duplicates()
# df_gi_to_plot_action_types = df_gi_to_plot_action_types[df_gi_to_plot_action_types['Actions Type (Marrakech)'].notna()]

# # Count the values in the 'Partner' column
# partner_counts = df_gi_to_plot_action_types['Partner'].value_counts()

# # Convert it to a DataFrame for better representation
# df_gi_partner_counts = partner_counts.reset_index()
# df_gi_partner_counts.columns = ['Partner', 'counts']

# # Sort the DataFrame in descending order of 'counts'
# df_gi_partners = df_gi_partner_counts.sort_values(by='counts', ascending=False)
# # Set the desired height for the chart
# fig_height = len(df_gi_partners) * 0.2
# # Create the horizontal bar chart
# fig, ax = plt.subplots(figsize=(8, fig_height))
# bars = ax.barh(df_gi_partners['Partner'], df_gi_partners['counts'], height=0.6)

# # Set the plot title and axis labels
# ax.set_title("Number of Actions Type (Marrakech) per RtR Partner", fontsize=10)
# ax.set_xlabel('Number of Actions Type (Marrakech)', fontsize=8)
# ax.set_ylabel('RtR Partner', fontsize=8)
# # Invert the y-axis to display partners from top to bottom
# ax.invert_yaxis()
# # Set the x-axis tick range to advance by 1 point
# ax.set_xticks(range(0, max(df_gi_partners['counts']) + 2, 1))
# # Set the font size for x-axis labels
# plt.xticks(fontsize=6)
# # Set the font size for partner names
# plt.yticks(fontsize=8)
# # Adjust the layout to prevent overlapping labels
# plt.tight_layout()

# # # Add annotation text at the bottom right of the figure
# # annotation_text = f"Out of {actioncluster_all} partners (Source: General Information Survey 2023)"
# # # Place the text at the bottom right of the figure
# # fig.text(0.95, 0.00, annotation_text, ha='right', va='bottom', fontsize=8, color='black', transform=fig.transFigure)

# # Add values on each bar
# for bar in bars:
# 	width = bar.get_width()
# 	label_y_pos = bar.get_y() + bar.get_height() / 2
# 	ax.text(width + 0.1, label_y_pos, s=f'{width}', fontsize=6, va='center', ha='left')
# # Show the plot in Streamlit
# st.markdown("#### 5.2.1 - Number of Actions Type (Marrakech) per RtR Partner")
# st.pyplot(fig)



# ##NUEVO 

# st.markdown("#### 5.2.2 - Number of Partners by Marrakech Action Types")
# # df_gi_partners_actioncluster_total.columns

# # Number of Partners resporting ActionClusters
# # n_partners_actiontypes = df_gi_partners_actioncluster_total['Partner'].nunique()

# # df_gi_partners_actioncluster_total_by_actypes = df_gi_partners_actioncluster_total grouped by df_gi_to_plot_action_types['Actions Type (Marrakech)']
# df_gi_to_plot_action_types = df_gi_partners_actioncluster_total[['Partner','Actions Type (Marrakech)']].drop_duplicates()
# df_gi_to_plot_action_types = df_gi_to_plot_action_types[df_gi_to_plot_action_types['Actions Type (Marrakech)'].notna()]

# # Count the values in the 'Partner' column
# partner_counts = df_gi_to_plot_action_types['Actions Type (Marrakech)'].value_counts()
# # Convert it to a DataFrame for better representation
# df_gi_partner_counts = partner_counts.reset_index()
# df_gi_partner_counts.columns = ['Marrakech Action Types', 'counts']
# # Sort the DataFrame in descending order of 'counts'
# df_gi_partners = df_gi_partner_counts.sort_values(by='counts', ascending=False)


# #chart 4
# # Define a function to split long labels into multiple lines
# def split_label(label, max_chars=60):
#     words = label.split()
#     split_labels = ['']
#     current_line = 0

#     for word in words:
#         # Check if adding the next word would exceed the max_chars limit
#         if len(split_labels[current_line]) + len(word) > max_chars:
#             # Check if we are already at 2 lines, if so, append the rest to the last line
#             if current_line == 1:
#                 split_labels[current_line] += ' ' + word
#             else:
#                 # Move to the next line if not
#                 current_line += 1
#                 split_labels.append(word)
#         else:
#             # Add the word with a space if the line is not empty
#             if split_labels[current_line]:
#                 split_labels[current_line] += ' ' + word
#             else:
#                 split_labels[current_line] += word

#     return '\n'.join(split_labels)

# # Apply the function to the labels
# df_gi_partners['Marrakech Action Types Three Lines'] = df_gi_partners['Marrakech Action Types'].apply(split_label)


# # Create the horizontal bar chart
# fig_height = len(df_gi_partners) * 0.6  # Adjust the figure height as needed
# fig, ax = plt.subplots(figsize=(10, fig_height))  # You might need to adjust the width (first argument) as well

# # Decrease the height value to something less than 0.8 if you want to increase the space between bars
# bars = ax.barh(df_gi_partners.index, df_gi_partners['counts'], height=0.6)  # Adjust the height for better space visibility

# # Set the plot title and axis labels
# ax.set_title("Number of Partners by Marrakech Action Type Categories", fontsize=12)
# ax.set_xlabel('RtR Partners', fontsize=10)
# ax.set_ylabel('Marrakech Action Types', fontsize=10)

# # Invert the y-axis to display labels from top to bottom
# ax.invert_yaxis()

# # Set the x-axis tick range to advance by 1 point
# ax.set_xticks(range(0, max(df_gi_partners['counts']) + 2, 1))

# # Set the font size for x-axis labels
# plt.xticks(fontsize=8)

# # Set the labels for y-axis with split text
# ax.set_yticks(df_gi_partners.index)
# ax.set_yticklabels(df_gi_partners['Marrakech Action Types Three Lines'], fontsize=10)

# # Adjust the layout to prevent overlapping labels
# plt.tight_layout()

# # Add annotation text at the bottom right of the figure
# annotation_text = f"Out of {actioncluster_all} partners (Source: General Information Survey 2023)"
# # Place the text at the bottom right of the figure
# fig.text(0.95, 0.00, annotation_text, ha='right', va='bottom', fontsize=8, color='black', transform=fig.transFigure)


# # Add values on each bar
# for bar in bars:
#     width = bar.get_width()
#     label_y_pos = bar.get_y() + bar.get_height() / 2
#     ax.text(width + 0.1, label_y_pos, s=f'{width}', fontsize=8, va='center', ha='left')

# # Show the plot in Streamlit
# st.pyplot(fig)



# # # Create the horizontal bar chart
# # fig_height = len(df_gi_partners) * 0.4  # Adjust the figure height as needed
# # fig, ax = plt.subplots(figsize=(10, fig_height))  # You might need to adjust the width (first argument) as well

# # # Decrease the height value to something less than 0.8 if you want to increase the space between bars
# # bars = ax.barh(df_gi_partners.index, df_gi_partners['counts'], height=0.5)  # Adjust the height for better space visibility

# # # Set the plot title and axis labels
# # ax.set_title("Number of Partners by Marrakech Action Type Categories (RtR Action Clusters Grouped)", fontsize=12)
# # ax.set_xlabel('RtR Partners', fontsize=10)
# # ax.set_ylabel('Marrakech Action Types', fontsize=10)

# # # Invert the y-axis to display labels from top to bottom
# # ax.invert_yaxis()

# # # Set the x-axis tick range to advance by 1 point
# # ax.set_xticks(range(0, max(df_gi_partners['counts']) + 2, 1))

# # # Set the font size for x-axis labels
# # plt.xticks(fontsize=8)

# # # Set the labels for y-axis with split text
# # ax.set_yticks(df_gi_partners.index)
# # ax.set_yticklabels(df_gi_partners['Marrakech Action Types Three Lines'], fontsize=10)

# # # Adjust the layout to prevent overlapping labels
# # plt.tight_layout()

# # # Add annotation text at the bottom right of the figure
# # annotation_text = f"Out of {actioncluster_all} partners (Source: General Information Survey 2023)"
# # # Place the text at the bottom right of the figure
# # fig.text(0.95, 0.00, annotation_text, ha='right', va='bottom', fontsize=8, color='black', transform=ax.transAxes)

# # # Add values on each bar
# # for bar in bars:
# #     width = bar.get_width()
# #     label_y_pos = bar.get_y() + bar.get_height() / 2
# #     ax.text(width + 0.1, label_y_pos, s=f'{width}', fontsize=8, va='center', ha='left')

# # # Show the plot in Streamlit
# # st.pyplot(fig)




# # st.caption("**TABLE #7:** RtR Partners & Action Clusters")
# # with st.expander("Explore Data: RtR Partners & Action Clusters"):
# # 	tab1, tab2, tab3 = st.tabs([
# # 		"RtR Partners's Action Clusters", 
# # 		"N° RtR Partners per Action Cluster",
# # 		"N° Action Clusters per RtR Partner",
# # 	])
# # 	with tab1: 
# # 		st.write("RtR Partners's Action Clusters")
# # 		st.write(df_gi_partners_actioncluster_total )
# # 		# csv = convert_df_gi(df_gi_partners_actioncluster_total)
# # 		# st.download_button(
# # 		# 			label="Download data as CSV",
# # 		# 			data=csv,
# # 		# 			file_name='RtR_GeneralInfoSurvey_Action_Clusters_'+datetofiles+'.csv',
# # 		# 			mime='text/csv',
# # 		# 		)
# # 	with tab2: 
# # 		st.write("N° RtR Partners per Action Cluster")
# # 		st.write(df_gi_actionclusters)
# # 	with tab3: 
# # 		st.write("N° Action Clusters per RtR Partner")
# # 		st.write(df_gi_partners)
	

 
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##__END BLOCK ACTION CLUSTERS GENERAL INFORMATION _________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________
# ##_____________________________________________________________________________________________________________

 
 
 
 
 
 
 
 
 
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #  BLOCK HAZARDS PLEDGE
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________



# # st.subheader("Hazards where RtR Partners aim to provide resilience")

# st.write("### Hazards Pledge")
# #Sample size hazards from selection.


# # st.write(df_pledge['All_Hazards_pdg'])
# df2_sz_hazards_pledge = df_pledge['All_Hazards_pdg'].replace(['; ; ; ; '], np.nan).dropna()
# sz_hazards_pledge = str(df2_sz_hazards_pledge.count())

# #dataframe to workwith (All data from the selection).
# df2 = df_pledge['All_Hazards_pdg'].str.split(";", expand=True).apply(lambda x: x.str.strip())
# df2 = df2.stack().value_counts()
# df2 = df2.iloc[1:].sort_index().reset_index(name='Frecuency')  #Check what happend whit blank information.

# #creating new columms based in a dictionary
# flooding_list = {'Coastal flood':'Flooding','Urban flood':'Flooding', 'River flood':'Flooding'}
# drought_list = {'Drought (agriculture focus)':'Drought','Drought (other sectors)':'Drought'}
# water_list   = {'Water stress (rural focus)':'Water','Water stress (urban focus)':'Water'}
# fire_list = {'Wildfire':'Fire'}
# coastal_list = {'Oceanic Events':'Coastal / Ocean','Other coastal events':'Coastal / Ocean'}
# cold_list = {'Extreme cold':'Cold',"Snow, Ice":'Cold'}
# wind_list = {'Extreme wind':'Wind','Tropical Cyclone':'Wind'}
# heat_list = {'Extreme heat':'Heat',}
# hazard_dictionary = {**heat_list,**cold_list,**drought_list,**water_list,**fire_list,**flooding_list,**coastal_list,**wind_list}
# df2['group'] = df2['index'].map(hazard_dictionary)
# df2['% hazard']=((df2['Frecuency']/df2['Frecuency'].sum())*100)

# # sum of % country in 'mayor_area' and creating a new df
# df3 = df2.groupby(['group']).agg({'Frecuency': 'sum'})
# df3 = df3.reset_index()
# df3.rename(columns={'Frecuency': 'Sum_frec_group'}, inplace=True)
# df3['% Hazards']=((df3['Sum_frec_group']/df3['Sum_frec_group'].sum())*100)
# # st.write(df3)

# df2 = pd.merge(df2, df3, on='group', how='left')
# df2 = df2.dropna()



# # Chart Hazard Nivel Profundo

# # This line creates the treemap and sets the color. Replace your previous treemap code with this
# fig_hazard_pledge = px.treemap(
#     df2, 
#     path=[px.Constant("Hazards"), 'group', 'index'], 
#     values='Frecuency',  # or 'Sum_frec_group' if you want the grouped sums instead
#     color='% Hazards',
#     color_continuous_scale='RdPu',
#     # This is where we define what text to display in the boxes of the treemap
#     custom_data=['% hazard', '% Hazards']
# )

# # This part sets what text to show on hover
# fig_hazard_pledge.update_traces(
#     hovertemplate="<b>%{label}</b><br>%{customdata[0]:.2f}% (Group)<br>%{customdata[1]:.2f}% (Total)",
#     texttemplate="<b>%{label}</b><br>%{customdata[0]:.2f}%"
# )

# # These are your layout updates
# fig_hazard_pledge.update_layout(
#     title='Climate Hazards Targeted by RtR Initiatives (Pledge Statements)',
#     font=dict(size=16),
#     margin=dict(t=50, l=25, r=25, b=25)
# )
# fig_hazard_pledge.add_annotation(
#     x=1, y=0,
#     text='Out of ' + str(sz) + ' RtR Initiatives - All Pledges. Pledge Statement Survey 2023',
#     showarrow=False,
#     yshift=-20
# )
# fig_hazard_pledge.update_traces(marker_line_color='grey')

# fig_hazard_pledge.update_traces(root_color='#FF37D5')
# fig_hazard_pledge.update_layout(title='Climate Hazards Targeted by RtR Initiatives (Pledge Statements)', font=dict(size=16))
# fig_hazard_pledge.update_layout(margin=dict(t=50, l=25, r=25, b=25))
# # fig_hazard_pledge.add_annotation(x=1, y=0,
# #                 text = 'Out of ' + str(sz) + ' RtR Initiatives - All Pledges. Pledge Statement Survey 2023',
# #                 showarrow=False,
# #                 yshift=-20)
# fig_hazard_pledge.data[0].hovertemplate = '%{value:.1f}%'
# fig_hazard_pledge.update_traces(marker_line_color='grey')

# fig_hazard_pledge.update_traces(
#     hovertemplate='%{label}<br>%{parent}<br>%{value:.1f}%',
#     marker_line_color='grey',
#     textinfo='label+text+value'
# )
# st.plotly_chart(fig_hazard_pledge)


# # CHART HAZARD RECODED
# # Adjust the treemap code to show only two levels
# fig_hazard_pledge_recoded = px.treemap(
#     df2, 
#     path=[px.Constant("Hazards"), 'group'], 
#     values='Sum_frec_group',  # Use the aggregated sum for each group
#     color='% Hazards',  # Assuming this is the correct column for group-level percentages
#     color_continuous_scale='RdPu',
#     # Update what text to display in the boxes of the treemap
#     custom_data=['% Hazards']  # Only one percentage level is now needed
# )

# # Update the hovertemplate to reflect only one custom data value
# fig_hazard_pledge_recoded.update_traces(
#     hovertemplate="<b>%{label}</b><br>%{customdata[0]:.2f}% (Group)",
#     texttemplate="<b>%{label}</b><br>%{customdata[0]:.2f}%",
#     textfont_size=16
# )

# # Apply your layout updates
# fig_hazard_pledge_recoded.update_layout(
#     title='Groups of Climate Hazards Targeted by RtR Initiatives (Pledge Statements)',
#     font=dict(size=16),
#     margin=dict(t=50, l=25, r=25, b=25)
# )

# # fig_hazard_pledge_recoded.add_annotation(x=1, y=0,
# #                 text = 'Out of ' + str(sz) + ' RtR Initiatives - All Pledges. Pledge Statement Survey 2023',
# #                 showarrow=False,
# #                 yshift=-20)
# fig_hazard_pledge_recoded.data[0].hovertemplate = '%{value:.1f}%'
# fig_hazard_pledge_recoded.update_traces(marker_line_color='grey')


# # Display the figure
# st.plotly_chart(fig_hazard_pledge_recoded)



# # Note for hazards
# # ____________________

# # ___________________
# # gettin the highst 3 hazards and lowest 2
# # ________________

# s_df2 = df3.sort_values(by='% Hazards', ascending=False)
# s_df2_best3 = s_df2.head(3)
# list_best3 = s_df2_best3['group'].unique()
# list_best3 =  ', '.join(list_best3)
# #making a function to change the last comma with "&"
# def replace_last(string, delimiter, replacement):
#     start, _, end = string.rpartition(delimiter)
#     return start + replacement + end
# #getting a string with the first three hazards
# list_best3_hazards = replace_last(list_best3, ',', ' & ')

# # _________________________
# # #making a list of the two less presnce in hazards
# s_df2 = df3.sort_values(by='% Hazards', ascending=True)
# s_df2_best3 = s_df2.head(2)
# list_best3 = s_df2_best3['group'].unique()
# list_best3 =  ', '.join(list_best3)
# #making a function to change the last comma with "&"
# def replace_last(string, delimiter, replacement):
#     start, _, end = string.rpartition(delimiter)
#     return start + replacement + end
# #getting a string with the hazards
# list_less_2_hazards = replace_last(list_best3, ',', ' & ')


# # st.plotly_chart(fig_hazard_pledge)
# with st.expander("Navigating this treemap: A User's Guide"):
#         st.write("""
#         This chart displays the percentage of hazard mitigation efforts that RtR Partners have pledged to provide*. The chart is structured as a treemap, with the main category of "Hazards" displayed at the top. Underneath that, you can see the different groups of climate hazards, such as heat, flooding, drought, fire, and cold. Finally, each individual hazard is displayed, such as extreme heat or flooding caused by heavy rainfall.
# 	The size of each box in the treemap represents the percentage of hazard mitigation efforts pledged for that particular hazard. The color of the boxes represents the percentage of total hazard mitigation efforts pledged by RtR Partners within that hazard group. The darker the color, the higher the percentage of total mitigation efforts pledged for that group.

# To view more information about each box, simply hover your mouse over it. The tooltip will display the name of the hazard and the percentage of hazard mitigation efforts pledged for that particular hazard. The total number of RtR Partners reporting information related to hazards is displayed in the annotation at the bottom of the chart.

# *Considering all RtR Partner’s pledges, the 100% is the result of the sum of the frequencies of each specific hazard divided by the total number of hazards.

# 	""")
# st.markdown('**Note:** '+list_best3_hazards+" are the hazard groups where RtR Partners aim to provide resilience the most. Partners' pledges show less focus on resilience for "+ list_less_2_hazards+ ".")


# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# # # END BLOCK HAZARDS PLEDGE
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________









# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# # # BLOCK RA PLAN ONLY 2023
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________



# tab1, tab2, tab3= st.tabs(["Resilience Attributes", "Subcategories","Conceptual Guide"])

# ra_likert_var_list = ['Equity_&_Inclusivity','Preparedness_and_planning','Learning','Agency','Social_Collaboration','Flexibility','Assets']


# # R A

# #Sample size RA
# df2_sz = df_2023_plan[ra_likert_var_list].dropna()
# sz = str(df2_sz.count().values[0])

# #dataframe to workwith (All data from the selection).
# df2 = df_2023_plan
# df2= df2[df2[ra_likert_var_list].notna()] #cleaning na

# Equity_y_Inclusivity        = df2['Equity_&_Inclusivity'].mean() 
# Preparedness_and_planning   = df2['Preparedness_and_planning'].mean() 
# Learning                    = df2['Learning'].mean() 
# Agency                      = df2['Agency'].mean() 
# Social_Collaboration        = df2['Social_Collaboration'].mean() 
# Flexibility                 = df2['Flexibility'].mean() 
# Assets                      = df2['Assets'].mean() 

# s_df2 = pd.DataFrame(dict(
#     r=[Equity_y_Inclusivity,Preparedness_and_planning,Learning,Agency,Social_Collaboration,Flexibility,Assets],
#     theta=['Equity and Inclusivity','Preparedness and planning',
#                         'Learning','Agency','Social Collaboration','Flexibility','Assets']))
# s_fig_ra_general = px.line_polar(s_df2, r='r', theta='theta', line_close=True, title="RtR GLOBAL CAMPAIGN - RESILIENCE ATTRIBUTES (PARTNERS' PLANS) ")
# s_fig_ra_general.update_traces(line_color='#FF37D5', line_width=1)
# s_fig_ra_general.update_traces(fill='toself')
# s_fig_ra_general.update_layout(
#     polar=dict(
#         radialaxis=dict(
#             tickfont=dict(color='#112E4D'),
#             range=[0, 4]
#         )
#     )
# )
# s_fig_ra_general.add_annotation(x=1, y=0,
#             text='Out of '+ sz +' Plan',showarrow=False,
#             yshift=-60)

# # st.write(s_df2)

# #making a list of the three most important ra
# s_df2 = s_df2.sort_values(by=['r'],ascending=False)
# s_df2_best3 = s_df2.head(3)
# list_best3 = s_df2_best3.theta.unique()
# list_best3 =  ', '.join(list_best3)
# #making a function to change the last comma with "&"
# def replace_last(string, delimiter, replacement):
#     start, _, end = string.rpartition(delimiter)
#     return start + replacement + end
# #getting a string with the first three ra.
# list_best3 = replace_last(list_best3, ',', ' & ')
# #making a list of the three last important ra.
# s_df2_last3 = s_df2.tail(3)
# s_df2_last3 = s_df2_last3.sort_values(by=['r'])
# list_last3 = s_df2_last3.theta.unique()
# list_last3 =  ', '.join(list_last3)
# list_last3 = replace_last(list_last3, ',', ' & ')

# with tab1:
#     st.subheader("ASSESSING RESILIENCE ATTRIBUTES: AN OVERVIEW OF RtR PARTNERS' PLANS")
#     st.write(s_fig_ra_general)
#     st.markdown("**Note:** "+list_best3 + " are top RtR Global Campaign Resilience Attributes.")
#     st.markdown( list_last3 + " represent Resilience Attributes for enhancement.")
#     with st.expander("Description of graph polar line: Resilience Attributes"):
#         st.markdown(""" TEXT NEED TO BE UPDATED
#        This is a polar line chart that shows the different Resilience Attributes that were reported by RtR partners.

# The chart has seven different attributes, which are plotted on the vertical axis, and include Equity and Inclusivity, Preparedness and planning, Learning, Agency, Social Collaboration, Flexibility, and Assets. Each attribute is represented by a line that connects data points on the chart, and the size of the line represents the strength of the attribute*.

# The chart is designed to help the user understand the relative strength of each attribute and how they contribute to overall resilience.

# To represent the strength of the resilience attributes, RtR partners responded from 0 to 4 for each of the resilience attribute subcategories asked for in the Plan survey. The response options are as follows:

# - Value 0: We do not develop actions that contribute to this subcategory.
# - Value 1: We contribute to this subcategory but indirectly, It’s not in the core of our activities.
# - Value 2: A few of our actions contribute to this subcategory.
# - Value 3: This subcategory is in the core of an important part of our actions.
# - Value 4: This subcategory is in the core of the most of our actions.


# The value in the chart's radio is the average of all those responses.
# 	""")

# #____
# ##SUB CATEGORIES - RESILIENCE ATTRIBUTES CHART
# #_____
# sub_ra_likert_var_list = ['Preparedness','Planning','Experiential_learning','Educational_Learning','Autonomy','Leadership',
# 'Decision_making','Collective_Participation','Connectivity','Coordination','Diversity','Redundancy','Distributive_Equity','Equity_of_Access',
# 'Finance','Infrastructure','Natural_Resources','Technologies','Basic_services',]

# #Sample size SUB RA
# df2_sz = df_2023_plan[sub_ra_likert_var_list].dropna()
# sz = str(df2_sz.count().values[0])

# #dataframe to workwith (All data from the selection).
# df2 = df_2023_plan  ## AVOID DUPLICATING THE DATA SET
# df2= df2[df2[sub_ra_likert_var_list].notna()] #cleaning na


# Distributive_Equity= df2['Distributive_Equity'].mean() 
# Equity_of_Access = df2['Equity_of_Access'].mean() 
# Preparedness = df2['Preparedness'].mean() 
# Planning = df2['Planning'].mean() 
# Experiential_learning = df2['Experiential_learning'].mean() 
# Educational_learning = df2['Educational_Learning'].mean() 
# Autonomy = df2['Autonomy'].mean() 
# Leadership = df2['Leadership'].mean() 
# Decision_making = df2['Decision_making'].mean() 
# Collective_participation = df2['Collective_Participation'].mean() 
# Connectivity = df2['Connectivity'].mean() 
# Coordination = df2['Coordination'].mean() 
# Diversity = df2['Diversity'].mean() 
# Redundancy = df2['Redundancy'].mean() 
# Finance = df2['Finance'].mean() 
# Natural_resources = df2['Natural_Resources'].mean() 
# Technologies = df2['Technologies'].mean() 
# Infrastructure = df2['Infrastructure'].mean() 
# Basic_services= df2['Basic_services'].mean() 


# s_df2 = pd.DataFrame(dict(
#     r=[Distributive_Equity,Equity_of_Access,Preparedness,Planning,Experiential_learning,Educational_learning,Autonomy,
# Leadership,Decision_making,Collective_participation,Connectivity,Coordination,
# Diversity,Redundancy,Finance,Natural_resources,Technologies,Infrastructure,Basic_services],
#     theta=['Distributive Equity','Equity of Access',	'Preparedness',
# 	                        'Planning',	'Experiential Learning','Educational Learning',
#     	                    'Autonomy',	'Leadership',	'Decision Making',	'Collective Participation',
#         	                'Connectivity',	'Coordination',	'Diversity',	'Redundancy',	'Finance',
#             	            'Natural Resources',	'Technologies',	'Infrastructure',	'Basic Services']))
# s_fig_sub_ra_general = px.line_polar(s_df2, r='r', theta='theta', line_close=True, title="RtR GLOBAL CAMPAIGN - RESILIENCE ATTRIBUTES SUB CATEGORIES (PARTNERS' PLANS)")
# s_fig_sub_ra_general.update_traces(line_color='#FF37D5', line_width=1)
# s_fig_sub_ra_general.update_traces(fill='toself')
# s_fig_sub_ra_general.update_layout(
#     polar=dict(
#         radialaxis=dict(
#             tickfont=dict(color='#112E4D'),
#             range=[0, 4]
#         )
#     )
# )
# s_fig_sub_ra_general.add_annotation(x=1, y=0,
#             text='Out of '+ sz +' Plans 2023',showarrow=False,
#             yshift=-60)
# # st.write(s_fig_sub_ra_general)

# #making a list of the three most important ra-subattribute
# s_df2 = s_df2.sort_values(by=['r'],ascending=False)
# s_df2_best3 = s_df2.head(3)
# list_best3 = s_df2_best3.theta.unique()
# list_best3 =  ', '.join(list_best3)
# #making a function to change the last comma with "&"
# def replace_last(string, delimiter, replacement):
#     start, _, end = string.rpartition(delimiter)
#     return start + replacement + end
# #getting a string with the first three ra-subattribute
# list_best3 = replace_last(list_best3, ',', ' & ')
# #making a list of the three last important ra-subattribute
# s_df2_last3 = s_df2.tail(3)
# s_df2_last3 = s_df2_last3.sort_values(by=['r'])
# list_last3 = s_df2_last3.theta.unique()
# list_last3 =  ', '.join(list_last3)
# list_last3 = replace_last(list_last3, ',', ' & ')


# with tab2:
#     st.subheader("ASSESSMENT OF RESILIENCE ATTRIBUTES SUBCATEGORIES IN PARTNERS' PLANS")
#     st.write(s_fig_sub_ra_general)
#     st.markdown("**Note:** "+list_best3 + " are top Resilience Attributes Sub Categories. "+list_last3 +" represent Resilience Attributes Sub Categoriess for enhancement in Partners' Plans.")
#     with st.expander("Description of graph polar line: Resilience Attributes"):
#         st.markdown("""NEED UPDATED This chart is a polar line chart that visually represents the mean scores of 19 subcategories of resilience attributes. 
 
# Each subcategory is listed on the perimeter of the chart, and the mean scores are shown as a line connecting the points that correspond to each subcategory. The y-axis indicates the scale of the mean scores*. Thus, this chart provides an overview of the performance of the resilience attributes subcategories, as well as aspects for future improvement and areas of opportunity.   
        
# To represent the strength of the resilience attributes, RtR partners responded from 0 to 4 for each of the resilience attribute subcategories asked for in the Pledge Statement survey. The response options are as follows:

# - Value 0: We do not develop actions that contribute to this subcategory.
# - Value 1: We contribute to this subcategory but indirectly, It’s not in the core of our activities.
# - Value 2: A few of our actions contribute to this subcategory.
# - Value 3: This subcategory is in the core of an important part of our actions.
# - Value 4: This subcategory is in the core of the most of our actions.


# The value in the chart's radio is the average of all those responses.
	
	
# 	""")

# with tab3:
#     # Display the figure with Streamlit
#     col1, col2, col3 = st.columns((2.2,0.4,0.4))
#     # col1.image("RA_definitions.png")
#     st.caption('ADD EXPLANATORY ')

# st.markdown("""---""")

# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# # # END BLOCK RA ONLY 2023
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________















# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# # # BLOCK RA PLAN CONSOLIDADO
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________
# #__________________________________________________________________________________________________________________________________________________________________



# tab1, tab2, tab3= st.tabs(["Resilience Attributes", "Subcategories","Conceptual Guide"])

# ra_likert_var_list = ['Equity_&_Inclusivity','Preparedness_and_planning','Learning','Agency','Social_Collaboration','Flexibility','Assets']


# # R A

# #Sample size RA
# df2_sz = df_2022_2023_plan_consolidated[ra_likert_var_list].dropna()
# sz = str(df2_sz.count().values[0])

# #dataframe to workwith (All data from the selection).
# df2 = df_2022_2023_plan_consolidated
# df2= df2[df2[ra_likert_var_list].notna()] #cleaning na

# Equity_y_Inclusivity        = df2['Equity_&_Inclusivity'].mean() 
# Preparedness_and_planning   = df2['Preparedness_and_planning'].mean() 
# Learning                    = df2['Learning'].mean() 
# Agency                      = df2['Agency'].mean() 
# Social_Collaboration        = df2['Social_Collaboration'].mean() 
# Flexibility                 = df2['Flexibility'].mean() 
# Assets                      = df2['Assets'].mean() 

# s_df2 = pd.DataFrame(dict(
#     r=[Equity_y_Inclusivity,Preparedness_and_planning,Learning,Agency,Social_Collaboration,Flexibility,Assets],
#     theta=['Equity and Inclusivity','Preparedness and planning',
#                         'Learning','Agency','Social Collaboration','Flexibility','Assets']))
# s_fig_ra_general = px.line_polar(s_df2, r='r', theta='theta', line_close=True, title="RESILIENCE ATTRIBUTES (PARTNERS' PLANS 2022 & 2023) ")
# s_fig_ra_general.update_traces(line_color='#FF37D5', line_width=1)
# s_fig_ra_general.update_traces(fill='toself')
# s_fig_ra_general.update_layout(
#     polar=dict(
#         radialaxis=dict(
#             tickfont=dict(color='#112E4D'),
#             range=[0, 4],
#             ),
#         angularaxis=dict(
#             tickfont=dict(size=13, color='black')
#         )
#     )
# )


#     # s_fig_ra_general.add_annotation(x=1, y=0,
#     #             text='Out of '+ sz +' Plans',showarrow=False,
#     #             yshift=-60)

# # st.write(s_df2)

# #making a list of the three most important ra
# s_df2 = s_df2.sort_values(by=['r'],ascending=False)
# s_df2_best3 = s_df2.head(3)
# list_best3 = s_df2_best3.theta.unique()
# list_best3 =  ', '.join(list_best3)
# #making a function to change the last comma with "&"
# def replace_last(string, delimiter, replacement):
#     start, _, end = string.rpartition(delimiter)
#     return start + replacement + end
# #getting a string with the first three ra.
# list_best3 = replace_last(list_best3, ',', ' & ')
# #making a list of the three last important ra.
# s_df2_last3 = s_df2.tail(3)
# s_df2_last3 = s_df2_last3.sort_values(by=['r'])
# list_last3 = s_df2_last3.theta.unique()
# list_last3 =  ', '.join(list_last3)
# list_last3 = replace_last(list_last3, ',', ' & ')

# with tab1:
#     st.subheader("ASSESSING RESILIENCE ATTRIBUTES: AN OVERVIEW OF RtR PARTNERS' PLANS")
#     st.write(s_fig_ra_general)
#     st.markdown("**Note:** "+list_best3 + " are top RtR Global Campaign Resilience Attributes.")
#     st.markdown( list_last3 + " represent Resilience Attributes for enhancement.")
#     with st.expander("Description of graph polar line: Resilience Attributes"):
#         st.markdown(""" TEXT NEED TO BE UPDATED
#        This is a polar line chart that shows the different Resilience Attributes that were reported by RtR partners.

# The chart has seven different attributes, which are plotted on the vertical axis, and include Equity and Inclusivity, Preparedness and planning, Learning, Agency, Social Collaboration, Flexibility, and Assets. Each attribute is represented by a line that connects data points on the chart, and the size of the line represents the strength of the attribute*.

# The chart is designed to help the user understand the relative strength of each attribute and how they contribute to overall resilience.

# To represent the strength of the resilience attributes, RtR partners responded from 0 to 4 for each of the resilience attribute subcategories asked for in the Plan survey. The response options are as follows:

# - Value 0: We do not develop actions that contribute to this subcategory.
# - Value 1: We contribute to this subcategory but indirectly, It’s not in the core of our activities.
# - Value 2: A few of our actions contribute to this subcategory.
# - Value 3: This subcategory is in the core of an important part of our actions.
# - Value 4: This subcategory is in the core of the most of our actions.


# The value in the chart's radio is the average of all those responses.
# 	""")

# #____
# ##SUB CATEGORIES - RESILIENCE ATTRIBUTES CHART
# #_____
# sub_ra_likert_var_list = ['Preparedness','Planning','Experiential_learning','Educational_learning','Autonomy','Leadership',
# 'Decision_making','Collective_participation','Connectivity','Coordination','Diversity','Redundancy','Distributive_Equity','Equity_of_Access',
# 'Finance','Infrastructure','Natural_resources','Technologies','Basic_services',]

# #Sample size SUB RA
# df2_sz = df_2022_2023_plan_consolidated[sub_ra_likert_var_list].dropna()
# sz = str(df2_sz.count().values[0])

# #dataframe to workwith (All data from the selection).
# df2 = df_2022_2023_plan_consolidated  ## AVOID DUPLICATING THE DATA SET
# df2= df2[df2[sub_ra_likert_var_list].notna()] #cleaning na


# Distributive_Equity= df2['Distributive_Equity'].mean() 
# Equity_of_Access = df2['Equity_of_Access'].mean() 
# Preparedness = df2['Preparedness'].mean() 
# Planning = df2['Planning'].mean() 
# Experiential_learning = df2['Experiential_learning'].mean() 
# Educational_learning = df2['Educational_learning'].mean() 
# Autonomy = df2['Autonomy'].mean() 
# Leadership = df2['Leadership'].mean() 
# Decision_making = df2['Decision_making'].mean() 
# Collective_participation = df2['Collective_participation'].mean() 
# Connectivity = df2['Connectivity'].mean() 
# Coordination = df2['Coordination'].mean() 
# Diversity = df2['Diversity'].mean() 
# Redundancy = df2['Redundancy'].mean() 
# Finance = df2['Finance'].mean() 
# Natural_resources = df2['Natural_resources'].mean() 
# Technologies = df2['Technologies'].mean() 
# Infrastructure = df2['Infrastructure'].mean() 
# Basic_services= df2['Basic_services'].mean() 


# s_df2 = pd.DataFrame(dict(
#     r=[Distributive_Equity,Equity_of_Access,Preparedness,Planning,Experiential_learning,Educational_learning,Autonomy,
# Leadership,Decision_making,Collective_participation,Connectivity,Coordination,
# Diversity,Redundancy,Finance,Natural_resources,Technologies,Infrastructure,Basic_services],
#     theta=['Distributive Equity','Equity of Access',	'Preparedness',
# 	                        'Planning',	'Experiential Learning','Educational Learning',
#     	                    'Autonomy',	'Leadership',	'Decision Making',	'Collective Participation',
#         	                'Connectivity',	'Coordination',	'Diversity',	'Redundancy',	'Finance',
#             	            'Natural Resources',	'Technologies',	'Infrastructure',	'Basic Services']))
# s_fig_sub_ra_general = px.line_polar(s_df2, r='r', theta='theta', line_close=True, title="RESILIENCE ATTRIBUTES SUB CATEGORIES (PARTNERS' PLANS 2022 & 2023)")
# s_fig_sub_ra_general.update_traces(line_color='#FF37D5', line_width=1)
# s_fig_sub_ra_general.update_traces(fill='toself')
# s_fig_sub_ra_general.update_layout(
#     polar=dict(
#         radialaxis=dict(
#             tickfont=dict(color='#112E4D'),
#             range=[0, 4],
#             ),
#         angularaxis=dict(
#             tickfont=dict(size=13, color='black')
#         )
#     )
# )

# # s_fig_sub_ra_general.add_annotation(x=1, y=0,
# #             text='Out of '+ sz +' Plans',showarrow=False,
# #             yshift=-60)
# # st.write(s_fig_sub_ra_general)

# #making a list of the three most important ra-subattribute
# s_df2 = s_df2.sort_values(by=['r'],ascending=False)
# s_df2_best3 = s_df2.head(3)
# list_best3 = s_df2_best3.theta.unique()
# list_best3 =  ', '.join(list_best3)
# #making a function to change the last comma with "&"
# def replace_last(string, delimiter, replacement):
#     start, _, end = string.rpartition(delimiter)
#     return start + replacement + end
# #getting a string with the first three ra-subattribute
# list_best3 = replace_last(list_best3, ',', ' & ')
# #making a list of the three last important ra-subattribute
# s_df2_last3 = s_df2.tail(3)
# s_df2_last3 = s_df2_last3.sort_values(by=['r'])
# list_last3 = s_df2_last3.theta.unique()
# list_last3 =  ', '.join(list_last3)
# list_last3 = replace_last(list_last3, ',', ' & ')


# with tab2:
#     st.subheader("ASSESSMENT OF RESILIENCE ATTRIBUTES SUBCATEGORIES IN PARTNERS' PLANS")
#     st.write(s_fig_sub_ra_general)
#     st.markdown("**Note:** "+list_best3 + " are top Resilience Attributes Sub Categories. "+list_last3 +" represent Resilience Attributes Sub Categoriess for enhancement in Partners' Plans.")
#     with st.expander("Description of graph polar line: Resilience Attributes"):
#         st.markdown("""NEED UPDATED This chart is a polar line chart that visually represents the mean scores of 19 subcategories of resilience attributes. 
 
# Each subcategory is listed on the perimeter of the chart, and the mean scores are shown as a line connecting the points that correspond to each subcategory. The y-axis indicates the scale of the mean scores*. Thus, this chart provides an overview of the performance of the resilience attributes subcategories, as well as aspects for future improvement and areas of opportunity.   
        
# To represent the strength of the resilience attributes, RtR partners responded from 0 to 4 for each of the resilience attribute subcategories asked for in the Pledge Statement survey. The response options are as follows:

# - Value 0: We do not develop actions that contribute to this subcategory.
# - Value 1: We contribute to this subcategory but indirectly, It’s not in the core of our activities.
# - Value 2: A few of our actions contribute to this subcategory.
# - Value 3: This subcategory is in the core of an important part of our actions.
# - Value 4: This subcategory is in the core of the most of our actions.


# The value in the chart's radio is the average of all those responses.
	
	
# 	""")

# with tab3:
#     # Display the figure with Streamlit
#     col1, col2, col3 = st.columns((2.2,0.4,0.4))
#     # col1.image("RA_definitions.png")
#     st.caption('ADD EXPLANATORY ')

# st.markdown("""---""")

# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # # END BLOCK RA CONSOLIDADO
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________
# # __________________________________________________________________________________________________________________________________________________________________



















#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
# BLOCK MEMBERS INFORMATION
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________
#__________________________________________________________________________________________________________________________________________________________________

st.header("SECTION 2. MEMBERS INFORMATION")
st.markdown("This segment serves as a space for RtR Partners to detail the entities that make up their membership - implementers, collaborators, endorsers, or donors. These are the groups or individuals who actively contribute to the pledge, plan, and resilience actions reported by the initiative. Importantly, they have provided their consent to be publicly associated with the RtR through the initiative.")
# st.caption("DATA: "+type_info)

st.subheader("MEMBERS WITHIN RTR PARTNERS: CONFIRMATION BASED ON DEFINITION")
st.caption("**Q8. In RtR, a member of an RtR Partner is defined as: an implementer, collaborator (partner organization), endorser, or donor - entities that directly contribute to the pledge, plan, and resilience actions reported by your initiative to the campaign, and that have consented to be linked to the RtR through the partner initiative. According to the above, does your initiative has members?**")
n_members_yes = len(df_gi[df_gi['q56'] == 'Yes'])
n_members_no = len(df_gi[df_gi['q56'] == 'No'])
n_members_unsure = len(df_gi[df_gi['q56'] == 'Unsure (please specify)'])

# st.caption("Q8. Number of partners confirming they have members as indicated in the definition: "+str(n_members_yes))
# st.caption("Q8. Number of partners confirming they do not have members as indicated in the definition: "+str(n_members_no))
# st.caption("Q8. Number of partners unsure about having members as indicated in the definition: "+str(n_members_unsure))

#SUNBURST:
# Create a dictionary for assigning specific colors to the unique values in 'q56'
color_dict = {"Yes": "#FF37D5", "No": "#9C27B0", "Unsure": "#DCE3E5"}
# Dataset to chart
counts = df_gi.groupby(['q56', 'InitName']).size().reset_index(name='Count')
# Create sunburst chart
fig = px.sunburst(counts, path=['q56', 'InitName'], values='Count', color='q56',
				color_discrete_map=color_dict)
# Add title
fig.update_layout(title="RtR Partners' Members Confirmation")

yes_responses = counts.loc[counts['q56'] == 'Yes', 'Count'].sum()
no_responses = counts.loc[counts['q56'] == 'No', 'Count'].sum()
unsure_responses = counts.loc[counts['q56'] == 'Unsure (please specify)', 'Count'].sum()

q9_responses = counts.groupby('InitName')['Count'].sum().to_dict()

col1, col2, col3= st.columns(3)
col1.metric("Partners with members (as defined)",n_members_yes)
col2.metric("Partners without members (as defined)",n_members_no)
col3.metric("Partners unsure of members (as defined)",n_members_unsure)

# Show sunburst with Streamlit
st.plotly_chart(fig)



markdown_text = f"""
As of {formatted_date_gi}, a total of {sum([yes_responses, no_responses, unsure_responses])} responses have been received for the question about membership confirmation in RtR partners. 

In RtR, a member of an RtR Partner is defined as an implementer, collaborator (partner organization), endorser, or donor. These are entities that directly contribute to the pledge, plan, and resilience actions reported by their initiative to the campaign, and that have consented to be linked to the RtR through their partner initiative.

Based on this definition:
- {yes_responses} initiatives confirmed that they have members ('Yes').
- {no_responses} initiatives declared they do not have any members ('No').
- {unsure_responses} initiatives are uncertain about their membership status ('Unsure').

**NOTE:** We need to understand why some partners are unsure or say they don't have any members, based on the definition we provided. This is important because we can only collect member information from partners who confirmed ('Yes') that they have members.

"""
# st.markdown(markdown_text)


df_gi_members_confirmation = df_gi[['Org_Type','InitName','members_info_index_percent','q56','members_type_all','q57','q62','q63','q64','q65','q66','q67','q68',]]
# df_gi_members_confirmation = df_gi[['InitName',,'q56','q57','members_type_all']]
df_gi_members_confirmation = df_gi_members_confirmation.sort_values('q56', ascending=False)

df_gi_members_confirmation['members_type_all'] = df_gi_members_confirmation['members_type_all'].replace("nan; nan; nan; nan", "")
# df_gi_members_confirmation['members_type_all'] = df_gi_members_confirmation['members_type_all'].replace('nan;', '', regex=True).replace('; ;', ';', regex=True).str.strip('; ')
df_gi_members_confirmation['members_type_all'] = df_gi_members_confirmation['members_type_all'].replace('\s*nan\s*;?', '', regex=True)


df_gi_members_confirmation = df_gi_members_confirmation.rename(columns={"InitName": "RtR Partner Name",'Org_Type':'Relation to RtR',
																	'members_info_index_percent':'Members Index %',
																	'q56': "initiative_has_members_according_to_RtR_definition",
																	'members_type_all':'types_members',
																	'q57': "Please specify",
																	'q62':'number_of_members_engaged_in_RtR_campaign',
																	'q63':'is_initiative_open_to_members',
																	'q64':'criteria_for_new_members',
																	'q65':'process_to_connect_RtR_to_members',
																	'q66':'process_for_members_to_report_resilience',
																	'q67':'members_file_upload',
																	'q68':'online_link_to_members_spreadsheet',
																	})

st.caption('**TABLE #3** Members within RtR Partners')
with st.expander("Explore Data: Members within RtR Partners"):
	st.write(df_gi_members_confirmation)
	# csv = convert_df_gi(df_gi_members_confirmation)
	# st.download_button(
	# 			label="Download data as CSV",
	# 			data=csv,
	# 			file_name='RtR_GeneralInfoSurvey_MembersInfo_'+datetofiles+'.csv',
	# 			mime='text/csv',
	# 		)


st.subheader("TYPES OF MEMBERS WITHIN RTR PARTNERS: IMPLEMENTER, COLLABORATOR, ENDORSER, DONOR")
st.caption("**Q9 Please indicate the types of members your initiative has. You may select multiple options.**")
# list of columns with Regions
typemembers_list = ['q58', 'q59', 'q60', 'q61']
typemembers_all = df_gi[typemembers_list].notna().any(axis=1).sum()
st.caption("Q9. Number of partners indicating types of members: "+str(typemembers_all))

## CHART HORIZONTAL
# Summing the counts for each type of member
typemembers_counts = df_gi[typemembers_list].notna().sum()
# Defining the types of members
member_types = ['Implementer', 'Collaborator', 'Endorser', 'Donor']
# Defining the colors using the RdPu colormap
colors = cm.get_cmap('RdPu')(typemembers_counts / typemembers_counts.max())
fig_type_members, ax = plt.subplots(figsize=(6, 2))

# Create horizontal bar chart
bars = ax.barh(member_types, typemembers_counts, color=colors)

# Increase the text size in the bar chart
for bar in bars:
    width = bar.get_width()
    ax.annotate(f'{width} Partners',
                xy=(width, bar.get_y() + bar.get_height() / 2),
                xytext=(20, 0),  # 3 points horizontal offset
                textcoords="offset points",
                ha='center', va='center',
                fontsize=6)  # Increase the font size for the labels here
    
# ax.set_title('Types of Members within RtR Partners', fontweight='bold', fontsize=12)
# ax.set_title('Diversity of Member Types within RtR Partners', fontsize=8)
# ax.set_title('Types of Members within RtR Partners')
# plt.title('Types of Members within RtR Partners')
ax.tick_params(axis='y', labelsize=6)

# Remove x-tick labels and ticks
ax.set_xticklabels([])
ax.tick_params(axis='x', length=0)
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)
# ax.xaxis.set_visible(False)
# If you need to remove the frame (spines) on the left and bottom as well, do:
ax.spines['left'].set_visible(False)
ax.spines['bottom'].set_visible(False)
# Remove y-tick marks but keep y-tick labels if necessary
ax.tick_params(axis='y', which='both', left=False)
ax.tick_params(axis='x', which='both', left=False)
ax.set_xlabel("Number of Partners", fontsize=6)
ax.set_ylabel("Types of  Partners' Members", fontsize=6)

# Adding the source text below the chart
# Since you're using Streamlit to display the plot, we need to ensure enough space below the figure to display additional text.
plt.gcf().subplots_adjust(bottom=0.2)  # Adjust the bottom of the figure to make room for the text.

# The figtext coordinates are in figure units, which go from 0 to 1 across the figure
plt.figtext(0.6, 0.00, f'Out of {typemembers_all} partners (Source: General Information Survey 2023).', 
            ha='center', fontsize=6, wrap=True)  # Increase the font size for the source text here
st.write("Chart No. 1: Diversity of RtR Partners' Members")
st.pyplot(fig_type_members)



# CHART VERTICAL OPCION 1 DE 2

# Assuming df_gi and typemembers_list are already defined
# Summing the counts for each type of member
typemembers_counts = df_gi[typemembers_list].notna().sum()
# Defining the types of members
member_types = ['Implementer', 'Collaborator', 'Endorser', 'Donor']

# Combine member_types and typemembers_counts into tuples
combined = list(zip(member_types, typemembers_counts))

# Sort the combined list by member types (alphabetically)
combined_sorted = sorted(combined)

# Separate the sorted list back into member_types and typemembers_counts
member_types_sorted, typemembers_counts_sorted = zip(*combined_sorted)

# Defining the colors using the RdPu colormap (adjust to the sorted counts)
# colors = cm.get_cmap('RdPu')(np.array(typemembers_counts_sorted) / max(typemembers_counts_sorted))

# Set the color for all bars to be #FF37D5
bar_color = '#FF37D5'

fig_type_members, ax = plt.subplots(figsize=(11, 5))

# Create vertical bar chart with sorted data
# bars = ax.bar(member_types_sorted, typemembers_counts_sorted, color=colors)
bars = ax.bar(member_types_sorted, typemembers_counts_sorted, color=bar_color)

# Increase the text size in the bar chart
for bar in bars:
    height = bar.get_height()
    ax.annotate(f'{height} Partners',
                xy=(bar.get_x() + bar.get_width() / 2, height),
                xytext=(0, 3),  # 3 points vertical offset
                textcoords="offset points",
                ha='center', va='bottom',
                fontsize=12)  # Increase the font size for the labels here

ax.tick_params(axis='y', labelsize=12)
ax.tick_params(axis='x', labelsize=12)
ax.tick_params(axis='x', length=5)
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# X axis: Setting integer ticks
ax.yaxis.set_major_locator(ticker.MaxNLocator(integer=True)) 
ax.set_ylim(0, 29)
    
ax.spines['top'].set_visible(False)
# ax.spines['left'].set_visible(False)
# ax.spines['bottom'].set_visible(False)

# Remove x-tick marks but keep x-tick labels if necessary
ax.tick_params(axis='x', which='both', bottom=False)
ax.set_ylabel("% of Partners", fontsize=14)
ax.set_xlabel("Member Categories", fontsize=14)

# # Adding the source text below the chart
# plt.gcf().subplots_adjust(bottom=0.17)  # Adjust the bottom of the figure to make room for the text.

# plt.figtext(0.90, 0.00, f'Out of {typemembers_all} partners (Source: General Information Survey 2023)', 
#             ha='right', fontsize=12, wrap=True)

# Streamlit display commands
st.write("Chart No. 1: Diversity of RtR Partners' Members")
st.pyplot(fig_type_members)





## CHART IN %

typemembers_list = ['q58', 'q59', 'q60', 'q61']
typemembers_all = df_gi[typemembers_list].notna().any(axis=1).sum()
total = n_reporting_partners
# st.caption("Q9. Number of partners indicating types of members: "+str(typemembers_all))

# Summing the counts for each type of member
typemembers_counts = df_gi[typemembers_list].notna().sum()
# Defining the types of members
member_types = ['Implementer', 'Collaborator', 'Endorser', 'Donor']

# Combine member_types and typemembers_counts into tuples
combined = list(zip(member_types, typemembers_counts))

# Sort the combined list by member types (alphabetically)
combined_sorted = sorted(combined)

# Separate the sorted list back into member_types and typemembers_counts
member_types_sorted, typemembers_counts_sorted = zip(*combined_sorted)

# Convert absolute counts to percentages
typemembers_counts_percent = [(count / total) * 100 for count in typemembers_counts_sorted]

# Set the color for all bars to be #FF37D5
# bar_color = '#FF37D5'
n_colors = 4
colors = cm.RdPu(np.linspace(0.2, 1, n_colors))

fig_type_members, ax = plt.subplots(figsize=(11, 2))

# Create vertical bar chart with sorted data and percentages
bars = ax.bar(member_types_sorted, typemembers_counts_percent, color=colors)

# Increase the text size in the bar chart
for bar in bars:
    height = bar.get_height()
    ax.annotate(f'{height:.1f}%',
                xy=(bar.get_x() + bar.get_width() / 2, height),
                xytext=(0, 3),  # 3 points vertical offset
                textcoords="offset points",
                ha='center', va='bottom',
                fontsize=12)  # Increase the font size for the labels here

ax.tick_params(axis='y', labelsize=12)
ax.tick_params(axis='x', labelsize=12)
ax.tick_params(axis='x', length=5)
ax.spines['right'].set_visible(False)
ax.spines['top'].set_visible(False)

# Adjust the y-axis limits and ticks to accommodate percentages
ax.set_ylim(0, 100)
ax.yaxis.set_major_formatter(ticker.PercentFormatter())

ax.spines['top'].set_visible(False)

# Remove x-tick marks but keep x-tick labels if necessary
ax.tick_params(axis='x', which='both', bottom=False)
ax.set_ylabel("% of Partners", fontsize=14)
ax.set_xlabel("Member Categories", fontsize=14)

# Streamlit display commands
st.write("Chart No.3: Diversity of RtR Partners' Members %")
st.pyplot(fig_type_members)





st.subheader("NUMBER OF MEMBERS")
st.caption("**Q10 Among the members that constitute your initiative, how many are explicitly engaged in and actively committed to the RtR campaign?**")
number_members = df_gi['q62'].count()
# Convert 'q62' column to numeric type
df_gi['q62'] = pd.to_numeric(df_gi['q62'], errors='coerce')
# Perform the sum
sum_members = int(df_gi['q62'].sum())
st.caption("Q10. Number of partners indicating a number of members: "+str(number_members)+". The sum of members reported here is: "+str(sum_members))
st.metric("Number of members",sum_members)


st.subheader("OPENNESS TO RECEIVING NEW MEMBERS")
st.caption("**Q11. Is your initiative open to receiving members?**")
new_members_question = df_gi['q63'].count()
st.caption("Q11. Number of partners responding the question: "+str(new_members_question))

df_gi_members_data = df_gi[['InitName','q63']]
df_gi_members_data = df_gi_members_data.sort_values('q63', ascending=False)
df_gi_members_data = df_gi_members_data.rename(columns={"q9": "RtR Partner Name", "q63": "is_initiative_open_to_members"})

with st.expander("Explore Answers"):
	st.write(df_gi_members_data)


st.subheader("ACCEPTANCE CRITERIA FOR NEW MEMBERS")
st.caption("**Q12 What criteria it uses to accept new members?**")
new_members_criteria = df_gi['q64'].count()
st.caption("Q12. Number of partners responding the question: "+str(new_members_criteria))

df_gi_members_confirmation = df_gi[['InitName','q64']]
df_gi_members_confirmation = df_gi_members_confirmation.sort_values('q64', ascending=False)
df_gi_members_confirmation = df_gi_members_confirmation.rename(columns={"q9": "RtR Partner Name", "q64": "Criteria"})

with st.expander("Explore Answers"):
	st.table(df_gi_members_confirmation)

st.subheader("ALIGNMENT PROCESS BETWEEN MEMBERS AND RTR")
st.caption("**Q13 What is the process your initiative takes to connect RtR to members  (how members acknowledge their actions are part of the campaign)?**")
members_rtr_connect = df_gi['q65'].count()
st.caption("Q13. Number of partners responding the question: "+str(members_rtr_connect))

df_gi_members_confirmation = df_gi[['InitName','q65']]
df_gi_members_confirmation = df_gi_members_confirmation.sort_values('q65', ascending=False)
df_gi_members_confirmation = df_gi_members_confirmation.rename(columns={"q9": "RtR Partner Name", "q65": "Aligment to RtR"})

with st.expander("Explore Answers"):
	st.table(df_gi_members_confirmation)

st.subheader("MEMBER REPORTING STRATEGIES")
st.caption("**Q14 How is the process for members to report resilience actions to your initiative?**")
members_reporting = df_gi['q66'].count()
st.caption("Q14. Number of partners responding the question: "+str(members_reporting))

df_gi_members_confirmation = df_gi[['InitName','q66']]
df_gi_members_confirmation = df_gi_members_confirmation.sort_values('q66', ascending=False)
df_gi_members_confirmation = df_gi_members_confirmation.rename(columns={"q9": "RtR Partner Name", "q66": "Reporting Strategies"})

with st.expander("Explore Answers"):
	st.table(df_gi_members_confirmation)


st.subheader("UPLOAD FILE CONTAINING MEMBER INFORMATION")
st.caption("**Q15 We kindly request that you upload a detailed file containing information about your current members. If possible, please clarify the types of members included.**")
members_upload_file = df_gi['q67'].count()
st.caption("Q15. Number of partners uploading a file: "+str(members_upload_file))

st.caption("**Q16. If before you could not save your spreadsheet with members' details as a Pdf_gi, DOC, or DOCX file, please paste an online link to your spreadsheet (in XLS, XLSX, CSV, TXT, or any other preferred format) in the space provided below.**")
members_spreadsheets_link = df_gi['q68'].count()
st.caption("Q16. Number of partners responding the question: "+str(members_spreadsheets_link))

# ______________________________________________
# ______________________________________________
# ______________________________________________
# ______________________________________________
# ______________________________________________
# END MEMBERS INFORMATION GI
# ______________________________________________
# ______________________________________________
# ______________________________________________
# ______________________________________________
# ______________________________________________
# ______________________________________________





# #####____________________________________________
# #####____________________________________________
# ##### DOWNLOAD DATA _____________________________
# #####____________________________________________
# #####____________________________________________


# # st.markdown("# DOWNLOAD ALL DATA")

# namegi = 'GI' + str(formatted_date_gi)
# namepledge = 'PLEDGE' + str(formatted_date_pledge)
# nameplan = 'PLAN' + str(formatted_date_plan)
# namegi = namegi.replace('/', '')
# namepledge = namepledge.replace('/', '')
# nameplan = nameplan.replace('/', '')

# # st.write("Last Update: ", namegi,", ",namepledge,", ",nameplan)
# st.sidebar.write("Last Update: ", namegi,", ",namepledge,", ",nameplan)

# def add_borders_to_range(ws, start_col, end_col, start_row, end_row):
#     thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
#     for row in ws.iter_rows(min_col=start_col, max_col=end_col, min_row=start_row, max_row=end_row):
#         for cell in row:
#             cell.border = thin_border

# @st.cache_data
# def convert_dfs_to_excel(df1, df2,df3):
#     # Convert dataframes to BytesIO object
#     output = BytesIO()
#     with pd.ExcelWriter(output, engine='openpyxl') as writer:
#         df1.to_excel(writer, sheet_name=namegi, index=False)
#         df2.to_excel(writer, sheet_name=namepledge, index=False)
#         df3.to_excel(writer, sheet_name=nameplan, index=False)
        
#         # Auto size columns and add borders
#         for sheet_name in writer.sheets:
#             worksheet = writer.sheets[sheet_name]
            
#             # Auto size columns
            
#             for column in worksheet.columns:
#                 max_length = 0
#                 column = [cell for cell in column]
#                 for cell in column:
#                     try:
#                         if len(str(cell.value)) > max_length:
#                             max_length = len(cell.value)
#                     except:
#                         pass
#                 adjusted_width = (max_length + 2)
#                 worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
#             # Add borders to all filled cells
#             max_row = worksheet.max_row
#             max_col = worksheet.max_column
#             add_borders_to_range(worksheet, start_col=1, end_col=max_col, start_row=1, end_row=max_row)
                
#     return output.getvalue()

# excel_data = convert_dfs_to_excel(df_gi_summary, df_pledge_summary,df_2023_plan_summary)

# st.download_button(
#     label="Download GI, Pledge & Plan Summary Data as Excel",
#     data=excel_data,
#     file_name='DataAllSurvey_summary.xlsx',
#     mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
# )

# #####____________________________________________
# #####____________________________________________
# ##### END DOWNLOAD DATA _____________________________
# #####____________________________________________
# #####____________________________________________




